"""Generate SAP490 testing deliverables by filling copied school templates."""

from __future__ import annotations

from copy import copy
from datetime import date
from pathlib import Path
import shutil

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.pagebreak import Break


ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_DIR = ROOT / "docs" / "sap490" / "templates" / "Deliverable_template"
OUTPUT_DIR = ROOT / "docs" / "sap490" / "generated"
VERSION = "0.1"
DOCUMENT_DATE = date(2026, 6, 15)
COMMIT = "4d6c05cbb924db7c1d0ffee3d6006ddfc4c3d686"
QA_SCRIPT_URL = (
    "https://github.com/DarwinDO/IDTS-SAP01/blob/"
    f"{COMMIT}/scripts/qa/test-idts6-programmatic.js"
)
QA_CHECKLIST_URL = (
    "https://github.com/DarwinDO/IDTS-SAP01/blob/"
    f"{COMMIT}/docs/qa/idts6-happy-flow-checklist.md"
)


TEST_CASES = [
    {
        "id": "SC-01a",
        "scenario": "SC-01",
        "en": "Create a bug with all mandatory fields.",
        "vi": "Tạo bug với đầy đủ trường bắt buộc.",
        "data_en": "Valid title, description, classification, priority, severity, environment and reproduction details.",
        "data_vi": "Title, mô tả, phân loại, priority, severity, environment và thông tin tái hiện hợp lệ.",
        "expected_en": "Create succeeds; equivalent HTTP status is 201.",
        "expected_vi": "Tạo thành công; trạng thái HTTP tương đương là 201.",
    },
    {
        "id": "SC-01b",
        "scenario": "SC-01",
        "en": "Create a bug without a title.",
        "vi": "Tạo bug không có title.",
        "data_en": "All mandatory fields except title.",
        "data_vi": "Có tất cả trường bắt buộc ngoại trừ title.",
        "expected_en": "Request is rejected with HTTP 400 and a title-required message.",
        "expected_vi": "Request bị từ chối với HTTP 400 và thông báo title là bắt buộc.",
    },
    {
        "id": "SC-02a",
        "scenario": "SC-02",
        "en": "Assign BUG-0001 to DatDT.",
        "vi": "Assign BUG-0001 cho DatDT.",
        "data_en": "Valid assignee ID and assignment note.",
        "data_vi": "Assignee ID hợp lệ và ghi chú assign.",
        "expected_en": "Action succeeds and status becomes ASSIGNED.",
        "expected_vi": "Action thành công và status chuyển thành ASSIGNED.",
    },
    {
        "id": "SC-02b",
        "scenario": "SC-02",
        "en": "Assign a bug without assigneeID.",
        "vi": "Assign bug nhưng không truyền assigneeID.",
        "data_en": "Assignment note only.",
        "data_vi": "Chỉ có ghi chú assign.",
        "expected_en": "Request is rejected with HTTP 400.",
        "expected_vi": "Request bị từ chối với HTTP 400.",
    },
    {
        "id": "SC-03a",
        "scenario": "SC-03",
        "en": "Move an assigned bug to In Review.",
        "vi": "Chuyển bug đã assign sang In Review.",
        "data_en": "BUG-0001 in ASSIGNED status.",
        "data_vi": "BUG-0001 đang ở status ASSIGNED.",
        "expected_en": "Action succeeds and status becomes IN_REVIEW.",
        "expected_vi": "Action thành công và status chuyển thành IN_REVIEW.",
    },
    {
        "id": "SC-04a",
        "scenario": "SC-04",
        "en": "Start work on a bug in review.",
        "vi": "Bắt đầu xử lý bug đang review.",
        "data_en": "BUG-0001 in IN_REVIEW status with an optional note.",
        "data_vi": "BUG-0001 đang ở IN_REVIEW kèm ghi chú tùy chọn.",
        "expected_en": "Action succeeds and status becomes IN_PROGRESS.",
        "expected_vi": "Action thành công và status chuyển thành IN_PROGRESS.",
    },
    {
        "id": "SC-05a",
        "scenario": "SC-05",
        "en": "Request more information with a reason.",
        "vi": "Yêu cầu thêm thông tin kèm lý do.",
        "data_en": "Reason: Need error logs.",
        "data_vi": "Lý do: Cần error log.",
        "expected_en": "Action succeeds and status becomes NEED_MORE_INFORMATION.",
        "expected_vi": "Action thành công và status chuyển thành NEED_MORE_INFORMATION.",
    },
    {
        "id": "SC-05b",
        "scenario": "SC-05",
        "en": "Request more information without a reason.",
        "vi": "Yêu cầu thêm thông tin nhưng không có lý do.",
        "data_en": "Empty reason.",
        "data_vi": "Reason rỗng.",
        "expected_en": "Request is rejected with HTTP 400.",
        "expected_vi": "Request bị từ chối với HTTP 400.",
    },
    {
        "id": "SC-06a",
        "scenario": "SC-06",
        "en": "Reject a bug with a classification reason.",
        "vi": "Reject bug kèm lý do phân loại sai.",
        "data_en": "Reason: Wrong classification - UI category.",
        "data_vi": "Lý do: Phân loại sai - UI category.",
        "expected_en": "Action succeeds and status becomes REJECTED.",
        "expected_vi": "Action thành công và status chuyển thành REJECTED.",
    },
    {
        "id": "SC-06b",
        "scenario": "SC-06",
        "en": "Reject a bug without a reason.",
        "vi": "Reject bug nhưng không có lý do.",
        "data_en": "Empty reason.",
        "data_vi": "Reason rỗng.",
        "expected_en": "Request is rejected with HTTP 400.",
        "expected_vi": "Request bị từ chối với HTTP 400.",
    },
    {
        "id": "SC-07a",
        "scenario": "SC-07",
        "en": "Move a rejected bug to Pending Assignment.",
        "vi": "Chuyển bug bị reject sang Pending Assignment.",
        "data_en": "Rejected bug; reason: Awaiting reclassification.",
        "data_vi": "Bug đang Rejected; lý do: Chờ phân loại lại.",
        "expected_en": "Action succeeds and status becomes PENDING_ASSIGNMENT.",
        "expected_vi": "Action thành công và status chuyển thành PENDING_ASSIGNMENT.",
    },
    {
        "id": "SC-08P1",
        "scenario": "SC-08",
        "en": "Reassign BUG-0003 to SangVN before resolve testing.",
        "vi": "Reassign BUG-0003 cho SangVN trước khi test resolve.",
        "unit_en": "Assign BUG-0003 to SangVN.",
        "unit_vi": "Assign BUG-0003 cho SangVN.",
        "data_en": "Valid SangVN developer ID.",
        "data_vi": "Developer ID hợp lệ của SangVN.",
        "expected_en": "Action succeeds and status becomes ASSIGNED.",
        "expected_vi": "Action thành công và status chuyển thành ASSIGNED.",
    },
    {
        "id": "SC-08P2",
        "scenario": "SC-08",
        "en": "Move BUG-0003 to In Review before resolve testing.",
        "vi": "Chuyển BUG-0003 sang In Review trước khi test resolve.",
        "unit_en": "Mark BUG-0003 In Review.",
        "unit_vi": "Chuyển BUG-0003 sang In Review.",
        "data_en": "BUG-0003 in ASSIGNED status.",
        "data_vi": "BUG-0003 đang ở status ASSIGNED.",
        "expected_en": "Action succeeds and status becomes IN_REVIEW.",
        "expected_vi": "Action thành công và status chuyển thành IN_REVIEW.",
    },
    {
        "id": "SC-08P3",
        "scenario": "SC-08",
        "en": "Move BUG-0003 to In Progress before resolve testing.",
        "vi": "Chuyển BUG-0003 sang In Progress trước khi test resolve.",
        "unit_en": "Start progress on BUG-0003.",
        "unit_vi": "Bắt đầu xử lý BUG-0003.",
        "data_en": "BUG-0003 in IN_REVIEW status.",
        "data_vi": "BUG-0003 đang ở status IN_REVIEW.",
        "expected_en": "Action succeeds and status becomes IN_PROGRESS.",
        "expected_vi": "Action thành công và status chuyển thành IN_PROGRESS.",
    },
    {
        "id": "SC-08b",
        "scenario": "SC-08",
        "en": "Resolve an in-progress bug without a developer note.",
        "vi": "Resolve bug đang xử lý nhưng không có developer note.",
        "data_en": "Empty resolve note.",
        "data_vi": "Resolve note rỗng.",
        "expected_en": "Request is rejected with HTTP 400 because resolve requires an explanation.",
        "expected_vi": "Request bị từ chối với HTTP 400 vì resolve bắt buộc có giải thích.",
    },
    {
        "id": "SC-08a",
        "scenario": "SC-08",
        "en": "Resolve an in-progress bug with a developer note.",
        "vi": "Resolve bug đang xử lý kèm developer note.",
        "data_en": "Note: Root cause fixed - corrected category mapping.",
        "data_vi": "Note: Đã sửa nguyên nhân gốc - điều chỉnh category mapping.",
        "expected_en": "Action succeeds and status becomes RESOLVED.",
        "expected_vi": "Action thành công và status chuyển thành RESOLVED.",
    },
    {
        "id": "SC-09a",
        "scenario": "SC-09",
        "en": "Send a resolved bug to retest.",
        "vi": "Chuyển bug đã resolve sang retest.",
        "data_en": "Resolved bug; note: Please retest in QAS.",
        "data_vi": "Bug đang Resolved; note: Vui lòng retest trên QAS.",
        "expected_en": "Action succeeds and status becomes RETEST_REQUIRED.",
        "expected_vi": "Action thành công và status chuyển thành RETEST_REQUIRED.",
    },
    {
        "id": "SC-10a",
        "scenario": "SC-10",
        "en": "Reopen a bug with a reason.",
        "vi": "Reopen bug kèm lý do.",
        "data_en": "Reason: Issue is still reproducible.",
        "data_vi": "Lý do: Lỗi vẫn tái hiện được.",
        "expected_en": "Action succeeds and status becomes REOPENED.",
        "expected_vi": "Action thành công và status chuyển thành REOPENED.",
    },
    {
        "id": "SC-10b",
        "scenario": "SC-10",
        "en": "Reopen a bug without a reason.",
        "vi": "Reopen bug nhưng không có lý do.",
        "data_en": "Empty reason.",
        "data_vi": "Reason rỗng.",
        "expected_en": "Request is rejected with HTTP 400.",
        "expected_vi": "Request bị từ chối với HTTP 400.",
    },
    {
        "id": "SC-11a",
        "scenario": "SC-11",
        "en": "Close a resolved bug after QA verification.",
        "vi": "Close bug đã resolve sau khi QA xác nhận.",
        "data_en": "Resolved bug; note: QA verified and closed.",
        "data_vi": "Bug đang Resolved; note: QA đã xác nhận và đóng.",
        "expected_en": "Action succeeds and status becomes CLOSED.",
        "expected_vi": "Action thành công và status chuyển thành CLOSED.",
    },
    {
        "id": "SC-12a",
        "scenario": "SC-12",
        "en": "Verify immutable history entries for lifecycle and next processor changes.",
        "vi": "Kiểm tra history bất biến cho thay đổi lifecycle và next processor.",
        "data_en": "Query the latest five HistoryLogs for BUG-0003.",
        "data_vi": "Query năm HistoryLogs mới nhất của BUG-0003.",
        "expected_en": "At least one entry exists; the run returns five entries including status and nextProcessor changes.",
        "expected_vi": "Có ít nhất một entry; lần chạy trả về năm entry gồm status và thay đổi nextProcessor.",
    },
]


SCENARIOS = [
    ("SC-01", "Create Bug and mandatory-field validation", "Tạo Bug và kiểm tra trường bắt buộc"),
    ("SC-02", "Assign Developer", "Assign Developer"),
    ("SC-03", "Developer review", "Developer review"),
    ("SC-04", "Start progress", "Bắt đầu xử lý"),
    ("SC-05", "Request more information", "Yêu cầu thêm thông tin"),
    ("SC-06", "Reject with follow-up reason", "Reject kèm lý do follow-up"),
    ("SC-07", "Move to Pending Assignment", "Chuyển sang Pending Assignment"),
    ("SC-08", "Resolve with mandatory note", "Resolve với note bắt buộc"),
    ("SC-09", "Send to retest", "Chuyển sang retest"),
    ("SC-10", "Reopen with mandatory reason", "Reopen với reason bắt buộc"),
    ("SC-11", "Close after QA verification", "Close sau khi QA xác nhận"),
    ("SC-12", "Verify HistoryLogs", "Kiểm tra HistoryLogs"),
]


LABELS = {
    "en": {
        "unit_title": "Unit Test",
        "scenario_title": "Test Scenario",
        "module": "Module",
        "module_name": "Module Name",
        "function_id": "Function ID",
        "function_name": "Function Name",
        "created_date": "Created Date",
        "updated_date": "Last Update Date",
        "approver": "Approver",
        "reviewer": "Reviewer",
        "creator": "Creator",
        "history": ["No", "Version", "Description", "Sheet", "Modified date", "Modified by"],
        "ut_headers": [
            "NO.",
            "Test Contents",
            "Test Results",
            "Evidence",
            "Test Cases",
            "Predicted Test Results",
            "Tester",
            "Test Date",
            "Result",
        ],
        "scenario_headers": ["No", "Step Name", "TEST CASE"],
        "case_headers": [
            "Business Flow",
            "Function Name",
            "Created/Updated by",
            "Created date",
            "Reviewed by",
            "Reviewed Date",
            "NO.",
            "Test Contents",
            "Test Cases",
            "Test Data",
            "Predicted Test Results",
        ],
        "history_description": "Initial IDTS Sprint 02 backend happy-flow test deliverable.",
        "unit_function": "Backend happy-flow lifecycle validation",
        "scenario_function": "Bug lifecycle happy-flow scenarios",
        "pass": "PASS",
        "tester": "NhanT / DonHV",
        "evidence": "QA script",
        "bug_headers": ["No", "Bug", "Details", "Expected result", "Fix"],
        "bug_title": "IDTS-5 - SC-01a direct CAP CREATE verification failed",
        "bug_details": (
            "The merged QA harness dispatched CREATE without an INSERT query. "
            "After adding the query, the test also exposed a missing explicit UUID."
        ),
        "bug_expected": "A valid bug request is created successfully and the complete backend suite passes.",
        "bug_fix": "Added the INSERT query and explicit UUIDs; rerun result: 21/21 PASS.",
        "page_document_id": "Document ID: F400",
        "unit_page_title": "Unit Test Specification - UT",
        "case_page_title": "Test Scenario Specification - Test Cases",
    },
    "vi": {
        "unit_title": "Kiểm thử đơn vị",
        "scenario_title": "Kịch bản kiểm thử",
        "module": "Mô-đun",
        "module_name": "Tên mô-đun",
        "function_id": "Mã chức năng",
        "function_name": "Tên chức năng",
        "created_date": "Ngày tạo",
        "updated_date": "Ngày cập nhật",
        "approver": "Người phê duyệt",
        "reviewer": "Người rà soát",
        "creator": "Người tạo",
        "history": ["STT", "Phiên bản", "Mô tả", "Sheet", "Ngày sửa", "Người sửa"],
        "ut_headers": [
            "STT",
            "Nội dung kiểm thử",
            "Kết quả kiểm thử",
            "Bằng chứng",
            "Test case",
            "Kết quả dự kiến",
            "Người test",
            "Ngày test",
            "Kết quả",
        ],
        "scenario_headers": ["STT", "Tên bước", "TEST CASE"],
        "case_headers": [
            "Luồng nghiệp vụ",
            "Tên chức năng",
            "Người tạo/cập nhật",
            "Ngày tạo",
            "Người rà soát",
            "Ngày rà soát",
            "STT",
            "Nội dung kiểm thử",
            "Test case",
            "Dữ liệu test",
            "Kết quả dự kiến",
        ],
        "history_description": "Khởi tạo tài liệu kiểm thử backend happy flow Sprint 02 của IDTS.",
        "unit_function": "Kiểm tra lifecycle backend theo happy flow",
        "scenario_function": "Kịch bản happy flow vòng đời bug",
        "pass": "PASS",
        "tester": "NhanT / DonHV",
        "evidence": "QA script",
        "bug_headers": ["STT", "Lỗi", "Chi tiết", "Kết quả mong đợi", "Cách sửa"],
        "bug_title": "IDTS-5 - SC-01a lỗi verify CAP CREATE trực tiếp",
        "bug_details": (
            "QA harness sau khi merge gọi CREATE nhưng thiếu INSERT query. "
            "Sau khi thêm query, test tiếp tục phát hiện thiếu UUID tường minh."
        ),
        "bug_expected": "Request tạo bug hợp lệ thành công và toàn bộ backend suite phải pass.",
        "bug_fix": "Đã thêm INSERT query và UUID tường minh; kết quả chạy lại: 21/21 PASS.",
        "page_document_id": "Mã tài liệu: F400",
        "unit_page_title": "Đặc tả kiểm thử đơn vị - UT",
        "case_page_title": "Đặc tả kịch bản kiểm thử - Test Cases",
    },
}


def copy_template(template_name: str, output_name: str) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output = OUTPUT_DIR / output_name
    shutil.copy2(TEMPLATE_DIR / template_name, output)
    return output


def write(ws, cell: str, value, *, hyperlink: str | None = None, wrap: bool = False):
    target = ws[cell]
    target.value = value
    if isinstance(value, date):
        target.number_format = "yyyy-mm-dd"
    if hyperlink:
        target.hyperlink = hyperlink
    if wrap:
        target.alignment = copy(target.alignment)
        target.alignment = Alignment(
            horizontal=target.alignment.horizontal,
            vertical=target.alignment.vertical or "top",
            text_rotation=target.alignment.text_rotation,
            wrap_text=True,
            shrink_to_fit=target.alignment.shrink_to_fit,
            indent=target.alignment.indent,
        )


def fill_cover(ws, labels: dict, title_key: str, function_name: str):
    write(ws, "B8", labels[title_key])
    write(ws, "I11", labels["module"])
    write(ws, "N11", "IDTS")
    write(ws, "U11", labels["module_name"])
    write(ws, "Z11", "Issue and Defect Tracking System in SAP")
    write(ws, "I12", labels["function_id"])
    write(ws, "N12", "IDTS-SP2-BE-HF")
    write(ws, "I13", labels["function_name"])
    write(ws, "N13", function_name)
    write(ws, "I14", labels["created_date"])
    write(ws, "N14", DOCUMENT_DATE)
    write(ws, "U14", labels["updated_date"])
    write(ws, "Z14", DOCUMENT_DATE)
    write(ws, "U18", labels["approver"])
    write(ws, "Z18", labels["reviewer"])
    write(ws, "AE18", labels["creator"])
    write(ws, "AE19", "DonHV / NhanT")


def fill_history(ws, labels: dict, sheets: str):
    for cell, value in zip(("B2", "C2", "D2", "E2", "F2", "G2"), labels["history"]):
        write(ws, cell, value)
    write(ws, "B3", 1)
    write(ws, "C3", VERSION)
    write(ws, "D3", labels["history_description"])
    write(ws, "E3", sheets)
    write(ws, "F3", DOCUMENT_DATE)
    write(ws, "G3", "DonHV")


def copy_single_row_template(ws, source_row: int, target_row: int, max_column: int):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for column in range(1, max_column + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)

    source_merges = [
        merged
        for merged in list(ws.merged_cells.ranges)
        if merged.min_row == source_row and merged.max_row == source_row
    ]
    for merged in source_merges:
        row_offset = target_row - source_row
        ws.merge_cells(
            start_row=merged.min_row + row_offset,
            start_column=merged.min_col,
            end_row=merged.max_row + row_offset,
            end_column=merged.max_col,
        )


def fill_unit_test(language: str):
    labels = LABELS[language]
    output = copy_template(
        "Unit_Test.xlsx", f"Unit_Test_IDTS_SAP01_{language}_v{VERSION}.xlsx"
    )
    wb = load_workbook(output)
    wb.properties.title = f"IDTS SAP490 Unit Test {language.upper()} v{VERSION}"
    wb.properties.subject = "Sprint 02 backend happy-flow verification"
    wb.properties.creator = "IDTS SAP01 Team"

    fill_cover(wb["Cover"], labels, "unit_title", labels["unit_function"])
    fill_history(wb["Histories"], labels, "UT")
    ws = wb["UT"]

    header_cells = ("B2", "L2", "AO2", "AX2", "BE2", "BL2")
    header_values = (
        labels["function_id"],
        labels["function_name"],
        labels["creator"],
        labels["created_date"],
        labels["reviewer"],
        labels["updated_date"],
    )
    for cell, value in zip(header_cells, header_values):
        write(ws, cell, value)

    write(ws, "B3", "IDTS-SP2-BE-HF")
    write(ws, "L3", labels["unit_function"])
    write(ws, "AO3", "NhanT / DonHV")
    write(ws, "AX3", DOCUMENT_DATE)

    unit_headers = labels["ut_headers"]
    for cell, value in zip(
        ("B6", "E6", "AX6", "BL6", "E7", "Y7", "AX7", "BD7", "BJ7"),
        unit_headers,
    ):
        write(ws, cell, value)

    ws.oddHeader.left.text = labels["page_document_id"]
    ws.oddHeader.right.text = labels["unit_page_title"]

    for row in range(17, 29):
        copy_single_row_template(ws, 16, row, 70)
    ws.row_breaks.append(Break(id=19))

    for index, test_case in enumerate(TEST_CASES, start=8):
        write(ws, f"B{index}", test_case["id"])
        write(ws, f"E{index}", test_case.get(f"unit_{language}", test_case[language]))
        write(ws, f"Y{index}", test_case[f"expected_{language}"], wrap=True)
        write(ws, f"AX{index}", labels["tester"])
        write(ws, f"BD{index}", DOCUMENT_DATE)
        write(ws, f"BJ{index}", labels["pass"])
        write(ws, f"BL{index}", labels["evidence"], hyperlink=QA_SCRIPT_URL)
        ws.row_dimensions[index].height = max(ws.row_dimensions[index].height or 15, 30)

    wb.save(output)
    return output


def fill_test_scenario(language: str):
    labels = LABELS[language]
    output = copy_template(
        "Test_Scenario.xlsx", f"Test_Scenario_IDTS_SAP01_{language}_v{VERSION}.xlsx"
    )
    wb = load_workbook(output)
    wb.properties.title = f"IDTS SAP490 Test Scenario {language.upper()} v{VERSION}"
    wb.properties.subject = "Sprint 02 backend happy-flow scenarios"
    wb.properties.creator = "IDTS SAP01 Team"

    fill_cover(wb["Cover"], labels, "scenario_title", labels["scenario_function"])
    fill_history(wb["Histories"], labels, "Test Scenario, Test Cases")

    matrix = wb["Test Scenario"]
    write(matrix, "A2", labels["scenario_headers"][0])
    write(matrix, "B2", labels["scenario_headers"][1])
    write(matrix, "C1", labels["scenario_headers"][2])

    for column in range(3, 22):
        matrix.cell(2, column).value = None
    for row in range(3, 20):
        matrix.cell(row, 2).value = None
        for column in range(3, 22):
            matrix.cell(row, column).value = None

    for column, (scenario_id, _, _) in enumerate(SCENARIOS, start=3):
        write(matrix, matrix.cell(2, column).coordinate, scenario_id)
    for row, (_, en_name, vi_name) in enumerate(SCENARIOS, start=3):
        write(matrix, f"A{row}", row - 2)
        write(matrix, f"B{row}", en_name if language == "en" else vi_name)
        write(matrix, matrix.cell(row, row).coordinate, "X")
    matrix.print_area = "A1:N19"
    matrix.sheet_properties.pageSetUpPr.fitToPage = True
    matrix.page_setup.fitToWidth = 1
    matrix.page_setup.fitToHeight = 1

    cases = wb["Test Cases"]
    metadata_cells = ("B2", "L2", "BF2", "BO2", "BV2", "CC2")
    metadata_labels = labels["case_headers"][:6]
    for cell, value in zip(metadata_cells, metadata_labels):
        write(cases, cell, value)
    write(cases, "B3", "Bug lifecycle happy flow" if language == "en" else "Happy flow vòng đời bug")
    write(cases, "L3", labels["scenario_function"])
    write(cases, "BF3", "DonHV / NhanT")
    write(cases, "BO3", DOCUMENT_DATE)

    for cell, value in zip(
        ("B6", "E6", "E7", "Y7", "AP7"), labels["case_headers"][6:]
    ):
        write(cases, cell, value)

    cases.oddHeader.left.text = labels["page_document_id"]
    cases.oddHeader.right.text = labels["case_page_title"]

    for row, test_case in enumerate(TEST_CASES, start=8):
        write(cases, f"B{row}", test_case["id"])
        write(cases, f"E{row}", test_case[language], wrap=True)
        write(cases, f"Y{row}", test_case[f"data_{language}"], wrap=True)
        write(cases, f"AP{row}", test_case[f"expected_{language}"], wrap=True)
        cases.row_dimensions[row].height = max(cases.row_dimensions[row].height or 15, 32)
    write(cases, "B29", None)
    write(cases, "B30", None)

    wb.save(output)
    return output


def fill_bug_log(language: str):
    labels = LABELS[language]
    output = copy_template(
        "Test_And_Fix_Bug.xlsx",
        f"Test_And_Fix_Bug_IDTS_SAP01_{language}_v{VERSION}.xlsx",
    )
    wb = load_workbook(output)
    wb.properties.title = f"IDTS SAP490 Test and Fix Bug {language.upper()} v{VERSION}"
    wb.properties.subject = "Sprint 02 QA defects and fixes"
    wb.properties.creator = "IDTS SAP01 Team"
    ws = wb["Fix and bugs"]

    for cell, value in zip(("A1", "B1", "C1", "D1", "E1"), labels["bug_headers"]):
        write(ws, cell, value)
    write(ws, "A2", 1)
    write(ws, "B2", labels["bug_title"], wrap=True)
    write(ws, "C2", labels["bug_details"], wrap=True)
    write(ws, "D2", labels["bug_expected"], wrap=True)
    write(ws, "E2", labels["bug_fix"], wrap=True)
    ws["F2"]._style = copy(ws["F3"]._style)
    write(ws, "F2", "IDTS-5 / QA evidence", hyperlink=QA_CHECKLIST_URL)
    ws.row_dimensions[2].height = 72

    for cell in ("F3", "F6"):
        ws[cell].value = None
        ws[cell].hyperlink = None

    wb.save(output)
    return output


def main():
    outputs = []
    for language in ("en", "vi"):
        outputs.extend(
            [
                fill_bug_log(language),
                fill_unit_test(language),
                fill_test_scenario(language),
            ]
        )
    for output in outputs:
        print(output.relative_to(ROOT))


if __name__ == "__main__":
    main()
