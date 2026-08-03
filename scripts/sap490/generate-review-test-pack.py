"""Fill the official SAP490 test templates from the canonical test catalog.

The generator intentionally preserves the official workbook sheet structure,
layout, styles, merges, images, and print setup. It only fills the existing
template regions, extends repeated data rows where necessary, and removes
known-broken sample references or explicitly unused template remnants.
"""

from __future__ import annotations

import json
import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink


ROOT = Path(__file__).resolve().parents[2]
CATALOG_PATH = ROOT / "docs" / "qa" / "test-catalog.json"
TEMPLATE_DIR = ROOT / "docs" / "sap490" / "templates" / "Deliverable_template"
REPORT_TEMPLATE = (
    ROOT / "docs" / "sap490" / "templates" / "2_SAP490_Test Report Template (1).xlsx"
)
OUTPUT_DIR = ROOT / "docs" / "sap490" / "generated"

FILES = {
    "scenario": (
        TEMPLATE_DIR / "Test_Scenario.xlsx",
        "Test_Scenario_IDTS_SAP01_{lang}_v0.3.xlsx",
        "0.3",
    ),
    "unit": (
        TEMPLATE_DIR / "Unit_Test.xlsx",
        "Unit_Test_IDTS_SAP01_{lang}_v0.4.xlsx",
        "0.4",
    ),
    "functional": (
        TEMPLATE_DIR / "Functional_Test.xlsx",
        "Functional_Test_IDTS_SAP01_{lang}_v0.3.xlsx",
        "0.3",
    ),
    "report": (
        REPORT_TEMPLATE,
        "Test_Report_IDTS_SAP01_{lang}_v0.4.xlsx",
        "0.4",
    ),
    "uat": (
        TEMPLATE_DIR / "UAT.xlsx",
        "UAT_IDTS_SAP01_{lang}_prepared_v0.2.xlsx",
        "0.2",
    ),
    "defect": (
        TEMPLATE_DIR / "Test_And_Fix_Bug.xlsx",
        "Test_And_Fix_Bug_IDTS_SAP01_{lang}_v0.5.xlsx",
        "0.5",
    ),
}

STATUS_DISPLAY = {
    "PASSED": "Passed",
    "FAILED": "Failed",
    "BLOCKED": "Blocked",
    "NOT_RUN": "Pending",
    "PENDING": "Pending",
    "PREPARED": "Pending",
    "N/A": "N/A",
}

FLOW_VI = {
    "AI safe data boundary": "Ranh giới dữ liệu an toàn cho AI",
    "AI cannot mutate bug": "AI không được tự thay đổi bug",
    "Create bug": "Tạo bug",
    "Required-field validation": "Validation trường bắt buộc",
    "Bearer authorization": "Phân quyền bearer",
    "Login and session creation": "Đăng nhập và tạo session",
    "Duplicate checking support": "Hỗ trợ kiểm tra trùng",
    "Classification concepts": "Khái niệm phân loại",
    "Pending Assignment": "Chờ phân công",
    "Reassignment history": "Lịch sử phân công lại",
    "Developer review": "Developer review",
    "Controlled rejection": "Từ chối có kiểm soát",
    "In Progress and Resolved": "Đang xử lý và Đã giải quyết",
    "Comments": "Bình luận",
    "Attachment persistence": "Lưu bền attachment",
    "In-app notifications": "Thông báo trong ứng dụng",
    "PM filters": "Bộ lọc PM",
}

BLUE = "0B5CAD"
LIGHT_BLUE = "DDEBF7"
LIGHT_GREEN = "E2F0D9"
LIGHT_YELLOW = "FFF2CC"
LIGHT_RED = "FCE4D6"
THIN_GRAY = Side(style="thin", color="B7B7B7")

VI_TEMPLATE_TRANSLATIONS = {
    "Module": "Phân hệ",
    "Module Name": "Tên phân hệ",
    "Function ID": "ID chức năng",
    "Function Name": "Tên chức năng",
    "Created Date": "Ngày tạo",
    "Last Update Date": "Ngày cập nhật",
    "Approver": "Người phê duyệt",
    "Reviewer": "Người review",
    "Creator": "Người tạo",
    "No": "STT",
    "NO.": "STT",
    "Version": "Phiên bản",
    "Description": "Mô tả",
    "Sheet": "Trang tính",
    "Modified date": "Ngày sửa",
    "Modified by": "Người sửa",
    "Step Name": "Tên bước",
    " TEST SCENARIOS": " KỊCH BẢN KIỂM THỬ",
    "Business Flow": "Luồng nghiệp vụ",
    "Created/Updated by": "Người tạo/cập nhật",
    "Created date": "Ngày tạo",
    "Reviewed by": "Người review",
    "Reviewed Date": "Ngày review",
    "Test Contents": "Nội dung kiểm thử",
    "Test Cases": "Ca kiểm thử",
    "Test Data": "Dữ liệu test",
    "Predicted Test Results": "Kết quả dự kiến",
    "Test Results": "Kết quả kiểm thử",
    "Tester": "Người thực thi",
    "Test Date": "Ngày test",
    "Result": "Kết quả",
    "Evidence": "Bằng chứng",
    "Remarks": "Ghi chú",
    "Test Case": "Ca kiểm thử",
    "TEST REPORT DOCUMENT": "TÀI LIỆU BÁO CÁO KIỂM THỬ",
    "TEST CASE LIST": "DANH SÁCH CA KIỂM THỬ",
    "TEST STATISTICS": "THỐNG KÊ KIỂM THỬ",
    "Project Name": "Tên dự án",
    "Project Code": "Mã dự án",
    "Issue Date": "Ngày phát hành",
    "Document Code": "Mã tài liệu",
    "Record of change": "Lịch sử thay đổi",
    "Effective Date": "Ngày hiệu lực",
    "Change Item": "Hạng mục thay đổi",
    "Change description": "Mô tả thay đổi",
    "Reference": "Tham chiếu",
    "Test Environment Setup Description": "Mô tả thiết lập môi trường test",
    "Sheet Name": "Tên trang tính",
    "Pre-Condition": "Điều kiện trước",
    "Reviewer/Approver": "Người review/phê duyệt",
    "Notes": "Ghi chú",
    "Module code": "Mã phân hệ",
    "Number of  test cases": "Số ca kiểm thử",
    "Sub total": "Tổng phụ",
    "Feature": "Chức năng",
    "Test requirement": "Yêu cầu kiểm thử",
    "Number of TCs": "Số ca kiểm thử",
    "Testing Round": "Vòng kiểm thử",
    "Round 1": "Vòng 1",
    "Round 2": "Vòng 2",
    "Round 3": "Vòng 3",
    "Test Case ID": "Mã ca kiểm thử",
    "Test Case Description": "Mô tả ca kiểm thử",
    "Test Case Procedure": "Quy trình kiểm thử",
    "Expected Results": "Kết quả dự kiến",
    "Pre-conditions": "Điều kiện trước",
    "Test date": "Ngày test",
    "Note": "Ghi chú",
    "Menu path": "Đường dẫn menu",
    "Scenario": "Kịch bản",
    "Test Case ": "Mã ca kiểm thử",
    "Test Data Description": "Mô tả dữ liệu test",
}


def load_catalog():
    return json.loads(CATALOG_PATH.read_text(encoding="utf-8"))


def local(value, lang):
    if isinstance(value, dict):
        return value.get(lang) or value.get("en") or ""
    if isinstance(value, list):
        return "\n".join(str(item) for item in value)
    return "" if value is None else str(value)


def joined(values):
    return ", ".join(str(value) for value in values) if values else ""


def numbered(value, lang):
    items = value.get(lang, []) if isinstance(value, dict) else value
    return "\n".join(f"{index}. {item}" for index, item in enumerate(items or [], 1))


def inline_steps(value, lang):
    items = value.get(lang, []) if isinstance(value, dict) else value
    return " | ".join(
        f"{index}) {item}" for index, item in enumerate(items or [], 1)
    )


def label(en, vi, lang):
    return en if lang == "en" else vi


def anchor(ws, coordinate):
    target = ws[coordinate]
    if not isinstance(target, MergedCell):
        return target
    for merged in ws.merged_cells.ranges:
        if coordinate in merged:
            return ws.cell(merged.min_row, merged.min_col)
    raise ValueError(f"No merge anchor for {ws.title}!{coordinate}")


def write(
    ws,
    coordinate,
    value,
    *,
    wrap=True,
    center=False,
    bold=None,
    shrink=False,
    font_size=None,
):
    target = anchor(ws, coordinate)
    target.value = value
    # A copied SAP490 template cell may carry an old sample hyperlink. Generated
    # text must start clean; callers add a verified link explicitly when needed.
    target.hyperlink = None
    old = copy(target.alignment)
    target.alignment = Alignment(
        horizontal="center" if center else (old.horizontal or "left"),
        vertical="center" if center else "top",
        wrap_text=wrap,
        text_rotation=0,
        shrink_to_fit=shrink,
        indent=old.indent,
    )
    if bold is not None or font_size is not None:
        new_font = copy(target.font)
        if bold is not None:
            new_font.bold = bold
        if font_size is not None:
            new_font.sz = font_size
        target.font = new_font


def add_hyperlink(cell, target, *, display=None):
    """Add a real hyperlink without changing the template font family/size."""
    if display is not None:
        cell.value = display
    cell.hyperlink = target
    link_font = copy(cell.font)
    link_font.color = "0563C1"
    link_font.underline = "single"
    cell.font = link_font


def add_internal_hyperlink(wb, cell, sheet_name, coordinate="A1"):
    """Fail generation when an internal target is missing or blank."""
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Missing hyperlink target sheet: {sheet_name}")
    if wb[sheet_name][coordinate].value in (None, ""):
        raise ValueError(f"Blank hyperlink target: {sheet_name}!{coordinate}")
    # Use an OOXML location-only hyperlink. Assigning a string beginning with
    # ``#`` makes OpenPyXL create an external relationship; LibreOffice then
    # rewrites it to a relative file URL that Google Sheets cannot follow.
    cell.hyperlink = Hyperlink(
        ref=cell.coordinate,
        location=f"'{sheet_name}'!{coordinate}",
        display=str(cell.value or sheet_name),
    )
    link_font = copy(cell.font)
    link_font.color = "0563C1"
    link_font.underline = "single"
    cell.font = link_font


def evidence_artifact_path(case, catalog):
    """Return the strongest reviewable artifact, never a command or source file."""
    case_id = case["caseId"]
    if case_id.startswith("UT-AI-"):
        return "docs/pm/evidence/idts-100/shared-qa-ai/idts72-ai-acceptance.json"
    for path in case.get("evidenceLinks", []):
        if not path.startswith("scripts/") and not path.endswith("/"):
            return path
    return catalog["evidenceBaseline"]["defaultEvidencePath"]


def evidence_url(catalog, path):
    baseline = catalog["evidenceBaseline"]
    return (
        f"{baseline['repositoryUrl']}/blob/"
        f"{baseline['evidenceCommitSha']}/{path}"
    )


def source_url(catalog, path):
    baseline = catalog["evidenceBaseline"]
    return (
        f"{baseline['repositoryUrl']}/blob/"
        f"{baseline['runtimeCommitSha']}/{path}"
    )


def run_for_case(case, catalog):
    command = case.get("automationCommand") or ""
    for run in catalog["executionRuns"]:
        if command and command == run["command"]:
            return run
    return None


def evidence_record(case, catalog):
    baseline = catalog["evidenceBaseline"]
    run = run_for_case(case, catalog)
    executed = case["status"] in {"PASSED", "FAILED", "BLOCKED"}
    artifact_path = evidence_artifact_path(case, catalog)
    is_shared_qa = "shared-qa" in artifact_path
    return {
        "evidenceId": f"EVD-{case['caseId']}",
        "caseId": case["caseId"],
        "runId": run["runId"] if run else (
            "SHARED-QA-20260724" if executed else "UAT-NOT-EXECUTED"
        ),
        "command": case.get("automationCommand") or "Manual UAT pending",
        "baseline": baseline["runtimeCommitSha"],
        "deploy": (
            f"{baseline['renderDeployId']}\n{baseline['runtimeCommitSha']}"
            if is_shared_qa
            else "N/A — local/programmatic evidence"
        ),
        "context": (
            f"{case.get('environment') or baseline['environment']} | "
            f"{case.get('executor') or 'Pending human executor'} | "
            f"{case.get('executionDate') or 'Not executed'}"
        ),
        "result": (
            f"{case['status']}"
            + (
                f" | P/F/S {run['passed']}/{run['failed']}/{run['skipped']} | exit {run['exitCode']}"
                if run else ""
            )
        ),
        "actual": local(case.get("actualResult", {}), "en") or "Not executed; prepared only.",
        "limitation": local(case.get("limitations", {}), "en") or "None recorded.",
        "artifactPath": artifact_path,
        "artifactUrl": evidence_url(catalog, artifact_path),
    }


def clear_values(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None


def copy_row_style(ws, source_row, target_row, max_col):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for column in range(1, max_col + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format


def clone_merged_row_block(ws, source_start, source_end, target_start, max_col):
    """Clone an official template row block, including shifted merged ranges."""
    offset = target_start - source_start
    for source_row in range(source_start, source_end + 1):
        copy_row_style(ws, source_row, source_row + offset, max_col)
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= source_start and merged.max_row <= source_end:
            shifted = (
                f"{get_column_letter(merged.min_col)}{merged.min_row + offset}:"
                f"{get_column_letter(merged.max_col)}{merged.max_row + offset}"
            )
            ensure_merge(ws, shifted)


def ensure_merge(ws, cell_range):
    if cell_range not in {str(item) for item in ws.merged_cells.ranges}:
        ws.merge_cells(cell_range)


def remove_broken_names(wb):
    for item in list(wb.defined_names.values()):
        text = item.attr_text or ""
        if "#REF!" in text or "[" in text or text.startswith("{"):
            del wb.defined_names[item.name]


def localize_template_labels(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if (
                    not isinstance(cell, MergedCell)
                    and isinstance(cell.value, str)
                    and cell.value in VI_TEMPLATE_TRANSLATIONS
                ):
                    cell.value = VI_TEMPLATE_TRANSLATIONS[cell.value]


def trim_sheet(
    ws,
    last_row,
    last_col,
    *,
    freeze=None,
    auto_filter=None,
    print_area=None,
    orientation=None,
    fit_width=1,
    title_rows=None,
):
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row > last_row or merged.min_col > last_col:
            ws.unmerge_cells(str(merged))
    if ws.max_row > last_row:
        ws.delete_rows(last_row + 1, ws.max_row - last_row)
    if ws.max_column > last_col:
        ws.delete_cols(last_col + 1, ws.max_column - last_col)
    ws.freeze_panes = freeze
    ws.auto_filter.ref = auto_filter
    ws.print_area = print_area or f"A1:{ws.cell(last_row, last_col).coordinate}"
    if orientation:
        ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = fit_width
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = title_rows


def prepare(kind, lang):
    template, filename, version = FILES[kind]
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output = OUTPUT_DIR / filename.format(lang=lang)
    shutil.copy2(template, output)
    wb = load_workbook(output)
    remove_broken_names(wb)
    if lang == "vi":
        localize_template_labels(wb)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    return wb, output, version


def cover(ws, title, function_id, function_name, catalog):
    evidence_date = datetime.fromisoformat(catalog["asOf"]).date()
    for coordinate, value in (
        ("B8", title),
        ("N11", "IDTS"),
        ("Z11", "Issue and Defect Tracking System in SAP"),
        ("N12", function_id),
        ("N13", function_name),
        ("N14", evidence_date),
        ("Z14", evidence_date),
        ("AE19", "DonHV"),
    ):
        write(ws, coordinate, value)
    for cell_range in ("N11:T11", "Z11:AI11", "N12:AI12", "N13:AI13"):
        ensure_merge(ws, cell_range)
    ws.column_dimensions["N"].width = 28
    ws.column_dimensions["Z"].width = 52
    for row, height in (
        (8, 34),
        (11, 34),
        (12, 34),
        (13, 48),
        (14, 26),
        (18, 24),
        (19, 28),
    ):
        ws.row_dimensions[row].height = height


def history(ws, version, description, sheets, catalog):
    evidence_date = datetime.fromisoformat(catalog["asOf"]).date()
    for coordinate, value in (
        ("B3", 1),
        ("C3", version),
        ("D3", description),
        ("E3", sheets),
        ("F3", evidence_date),
        ("G3", "DonHV"),
    ):
        write(ws, coordinate, value)
    clear_values(ws, 4, 27, 2, 7)
    for column, width in {
        "B": 7,
        "C": 11,
        "D": 72,
        "E": 38,
        "F": 16,
        "G": 18,
    }.items():
        ws.column_dimensions[column].width = width
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 48


def execution_note(case, lang):
    return " | ".join(
        [
            f"{label('Actual', 'Thực tế', lang)}: "
            f"{local(case['actualResult'], lang) or label('Not run', 'Chưa chạy', lang)}",
            f"{label('Env', 'Môi trường', lang)}: {case['environment'] or 'N/A'}",
            f"{label('Command', 'Lệnh', lang)}: {case['automationCommand'] or 'N/A'}",
            f"{label('Evidence', 'Bằng chứng', lang)}: {joined(case['evidenceLinks']) or 'N/A'}",
            f"{label('Defects', 'Defect', lang)}: {joined(case['defectIds']) or 'None'}",
            f"{label('Limit', 'Giới hạn', lang)}: {local(case['limitations'], lang) or 'None'}",
        ]
    )


def case_definition(case, lang, *, unit=False, uat=False):
    if unit:
        return "\n".join(
            [
                f"{local(case['title'], lang)} — Req: {joined(case['requirementIds'])}",
                f"{label('Objective', 'Mục tiêu', lang)}: {local(case['objective'], lang)}",
                f"{label('Setup / input', 'Setup / đầu vào', lang)}: "
                f"{local(case['preconditions'], lang)} | {local(case['testData'], lang)}",
                f"{label('Steps', 'Các bước', lang)}: {inline_steps(case['steps'], lang)}",
            ]
        )
    lines = [
        f"{local(case['title'], lang)} — {case['testType']} / "
        f"{case['classification']} / {case['priority']} / {case['role']}",
        f"Req: {joined(case['requirementIds'])} | "
        f"{label('Objective', 'Mục tiêu', lang)}: {local(case['objective'], lang)}",
        f"{label('Preconditions', 'Điều kiện trước', lang)}: "
        f"{local(case['preconditions'], lang)}",
        f"{label('Steps', 'Các bước', lang)}: {inline_steps(case['steps'], lang)}",
    ]
    if uat:
        lines.append(
            f"{label('Execution state', 'Trạng thái thực thi', lang)}: "
            f"PREPARED — {label('not executed; no sign-off', 'chưa thực thi; chưa sign-off', lang)}"
        )
    return "\n".join(lines)


def expected_block(case, lang, *, unit=False, uat=False):
    lines = []
    lines.extend(
        [
            f"{label('Assertion / expected', 'Assertion / dự kiến', lang)}: "
            f"{local(case['expectedResult'], lang)}",
            f"{label('Postcondition / persistence', 'Điều kiện sau / persistence', lang)}: "
            f"{local(case['postcondition'], lang)}",
        ]
    )
    if unit:
        lines.append(
            f"{label('Actual output', 'Kết quả thực tế', lang)}: "
            f"{local(case['actualResult'], lang)}"
        )
    if uat:
        lines.append(
            f"{label('Pass/Fail rule', 'Quy tắc Pass/Fail', lang)}: "
            f"{label('Pass only when every acceptance checkpoint matches; otherwise Fail and record a defect.', 'Chỉ Pass khi mọi checkpoint chấp nhận khớp; nếu không thì Fail và ghi defect.', lang)}"
        )
    return "\n".join(lines)


def set_case_row_height(ws, row, height=84):
    current = ws.row_dimensions[row].height or 15
    ws.row_dimensions[row].height = min(max(current, height), 90)


def generate_scenario(catalog, lang):
    wb, output, version = prepare("scenario", lang)
    cover(
        wb["Cover"],
        label("Test Scenario", "Kịch bản kiểm thử", lang),
        "IDTS-SAP490-SCENARIO",
        label(
            "Requirement-linked current test scenarios",
            "Kịch bản kiểm thử hiện hành liên kết requirement",
            lang,
        ),
        catalog,
    )
    history(
        wb["Histories"],
        version,
        label(
            "Filled the official template from the canonical test catalog; predicted results are separate from test data.",
            "Điền template chính thức từ test catalog chuẩn; kết quả dự kiến tách khỏi dữ liệu test.",
            lang,
        ),
        "Test Scenario, Test Cases",
        catalog,
    )

    requirements = {item["id"]: item for item in catalog["requirements"]}
    cases = [case for case in catalog["cases"] if case["testType"] != "UAT"]
    flows = {}
    for case in cases:
        flow = requirements[case["requirementIds"][0]]["flow"]
        flows.setdefault(flow, []).append(case["caseId"])

    matrix = wb["Test Scenario"]
    clear_values(matrix, 2, 19, 2, 23)
    write(matrix, "C1", label("TEST CASE", "CA KIỂM THỬ", lang), center=True, bold=True)
    for column, case in enumerate(cases, 3):
        source_column = min(column, 21)
        if column > 21:
            matrix.cell(2, column)._style = copy(matrix.cell(2, source_column)._style)
        write(
            matrix,
            matrix.cell(2, column).coordinate,
            case["caseId"],
            wrap=False,
            center=True,
            shrink=True,
        )
        matrix.column_dimensions[matrix.cell(2, column).column_letter].width = 13
    matrix.row_dimensions[2].height = 30
    for row, (flow, case_ids) in enumerate(flows.items(), 3):
        write(matrix, f"A{row}", row - 2, center=True)
        write(matrix, f"B{row}", FLOW_VI.get(flow, flow) if lang == "vi" else flow)
        matrix.row_dimensions[row].height = 30
        for column, case in enumerate(cases, 3):
            if case["caseId"] in case_ids:
                write(matrix, matrix.cell(row, column).coordinate, "X", center=True)

    ws = wb["Test Cases"]
    for coordinate, value in (
        ("B3", label("IDTS SAP490 current test scenarios", "Kịch bản test IDTS SAP490 hiện hành", lang)),
        ("L3", label("CAP/Fiori issue and defect lifecycle", "Vòng đời issue/defect CAP/Fiori", lang)),
        ("BF3", "DonHV"),
        ("BO3", datetime.fromisoformat(catalog["asOf"]).date()),
        ("BV3", label("Mentor / Supervisor", "Mentor / Giảng viên", lang)),
        ("CC3", label("Pending review", "Chờ review", lang)),
        ("BO7", label("Traceability / classification", "Truy vết / phân loại", lang)),
        ("BU7", label("Executor / Date", "Người thực thi / Ngày", lang)),
        ("CA7", label("Status", "Trạng thái", lang)),
        ("CC7", label("Evidence / Defect", "Bằng chứng / Defect", lang)),
    ):
        write(ws, coordinate, value, center=coordinate.endswith("7"))
    ws.row_dimensions[7].height = 40
    for row in range(8, 31):
        add_group_merges(ws, row, (("CC", "CI"),))
        for coordinate in (f"B{row}", f"E{row}", f"Y{row}", f"AP{row}", f"BO{row}", f"BU{row}", f"CA{row}", f"CC{row}"):
            write(ws, coordinate, None)
    for row, case in enumerate(cases, 8):
        write(ws, f"B{row}", case["caseId"], center=True)
        write(
            ws,
            f"E{row}",
            f"{local(case['title'], lang)}\n"
            f"{label('Objective', 'Mục tiêu', lang)}: {local(case['objective'], lang)}\n"
            f"Req: {joined(case['requirementIds'])}\n"
            f"{label('Postcondition', 'Điều kiện sau', lang)}: "
            f"{local(case['postcondition'], lang)}",
        )
        write(
            ws,
            f"Y{row}",
            f"{label('Role', 'Vai trò', lang)}: {case['role']} | "
            f"{label('Preconditions', 'Điều kiện trước', lang)}: "
            f"{local(case['preconditions'], lang)}\n"
            f"{label('Test data', 'Dữ liệu test', lang)}: {local(case['testData'], lang)}",
        )
        write(
            ws,
            f"AP{row}",
            f"{label('Steps', 'Các bước', lang)}: {inline_steps(case['steps'], lang)}\n"
            f"{label('Expected', 'Dự kiến', lang)}: {local(case['expectedResult'], lang)}",
        )
        write(
            ws,
            f"BO{row}",
            f"{case['testType']} / {case['classification']} / {case['priority']}",
        )
        write(
            ws,
            f"BU{row}",
            f"{case['executor'] or ''}\n{case['executionDate'] or ''}".strip(),
        )
        write(ws, f"CA{row}", case["status"], center=True)
        write(
            ws,
            f"CC{row}",
            f"{joined(case['evidenceLinks'])}\n{joined(case['defectIds'])}".strip(),
        )
        ws.row_dimensions[row].height = 130

    trim_sheet(wb["Cover"], 20, 43, print_area="B8:AO20", orientation="landscape")
    trim_sheet(wb["Histories"], 3, 7, freeze="B3", auto_filter="B2:G3", print_area="B2:G3", orientation="landscape", title_rows="2:2")
    trim_sheet(matrix, 19, 23, freeze="C3", print_area="A1:W19", orientation="landscape", fit_width=2, title_rows="1:2")
    trim_sheet(ws, 28, 87, freeze="B8", auto_filter="B7:CI28", print_area="B2:CI28", orientation="landscape", fit_width=1, title_rows="6:7")
    save(wb, output, "Test Scenario", lang, version)
    return output


def generate_unit(catalog, lang):
    wb, output, version = prepare("unit", lang)
    cover(
        wb["Cover"],
        label("Unit Test", "Kiểm thử đơn vị", lang),
        "IDTS-SAP490-UNIT",
        label("AI boundary unit tests", "Unit test ranh giới AI", lang),
        catalog,
    )
    history(
        wb["Histories"],
        version,
        label(
            "Classified true unit cases separately from programmatic regression runs.",
            "Phân loại unit case đúng nghĩa tách khỏi programmatic regression run.",
            lang,
        ),
        "UT, Evidence",
        catalog,
    )
    cases = [case for case in catalog["cases"] if case["testType"] == "UNIT"]
    ws = wb["UT"]
    for coordinate, value in (
        ("B3", "IDTS-SAP490-UNIT"),
        ("L3", label("AI boundary unit tests", "Unit test ranh giới AI", lang)),
        ("AO3", "DonHV"),
        ("AX3", datetime.fromisoformat(catalog["asOf"]).date()),
    ):
        write(ws, coordinate, value)
    for row in range(8, 17):
        add_group_merges(
            ws,
            row,
            (
                ("B", "D"),
                ("E", "X"),
                ("Y", "AW"),
                ("AX", "BC"),
                ("BD", "BI"),
                ("BJ", "BK"),
            ),
        )
        old_evidence_merge = f"BL{row}:BR{row}"
        if old_evidence_merge in {
            str(item) for item in ws.merged_cells.ranges
        }:
            ws.unmerge_cells(old_evidence_merge)
        ensure_merge(ws, f"BL{row}:BV{row}")
        for coordinate in (f"B{row}", f"E{row}", f"Y{row}", f"AX{row}", f"BD{row}", f"BJ{row}", f"BL{row}"):
            write(ws, coordinate, None)
    for row, case in enumerate(cases, 8):
        evidence_row = row - 6
        record = evidence_record(case, catalog)
        write(ws, f"B{row}", case["caseId"], center=True)
        write(
            ws,
            f"E{row}",
            f"{local(case['title'], lang)} — Req: {joined(case['requirementIds'])}\n"
            f"{label('Steps', 'Các bước', lang)}: {inline_steps(case['steps'], lang)}",
            font_size=12,
        )
        write(
            ws,
            f"Y{row}",
            f"{label('Setup / input', 'Setup / đầu vào', lang)}: "
            f"{local(case['preconditions'], lang)} | {local(case['testData'], lang)}\n"
            f"{expected_block(case, lang, unit=True)}",
            font_size=12,
        )
        write(ws, f"AX{row}", case["executor"], center=True)
        write(ws, f"BD{row}", case["executionDate"], center=True)
        write(ws, f"BJ{row}", STATUS_DISPLAY[case["status"]], center=True)
        write(
            ws,
            f"BL{row}",
            f"{label('Command', 'Lệnh', lang)}: {case['automationCommand']}\n"
            f"{label('Evidence', 'Bằng chứng', lang)}: {record['evidenceId']}",
            font_size=12,
        )
        ws.row_dimensions[row].height = 125

    evidence = wb["Evidence"]
    clear_values(evidence, 1, max(evidence.max_row, 20), 1, 12)
    headers = [
        label("Evidence ID", "Mã bằng chứng", lang),
        label("Case ID", "Mã ca kiểm thử", lang),
        label("Run ID", "Run ID", lang),
        label("Exact command", "Lệnh chính xác", lang),
        label("Baseline commit", "Commit baseline", lang),
        label("Deploy ID / SHA", "Deploy ID / SHA", lang),
        label("Environment / executor / time", "Môi trường / người chạy / thời điểm", lang),
        label("Exit / result", "Exit / kết quả", lang),
        label("Actual result", "Kết quả thực tế", lang),
        label("Limitation", "Giới hạn", lang),
        label("Artifact", "Artifact", lang),
    ]
    for column, header_text in enumerate(headers, 1):
        cell = evidence.cell(1, column, header_text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
    for row, case in enumerate(cases, 2):
        record = evidence_record(case, catalog)
        values = [
            record["evidenceId"],
            record["caseId"],
            record["runId"],
            record["command"],
            record["baseline"],
            record["deploy"],
            record["context"],
            record["result"],
            local(case.get("actualResult", {}), lang) or label("Not executed", "Chưa thực thi", lang),
            local(case.get("limitations", {}), lang) or label("None recorded", "Không ghi nhận", lang),
            record["artifactPath"],
        ]
        for column, value in enumerate(values, 1):
            cell = evidence.cell(row, column, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True, text_rotation=0)
            cell.border = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
            cell.font = Font(name="Arial", size=12)
        add_hyperlink(evidence.cell(row, 11), record["artifactUrl"])
        evidence.row_dimensions[row].height = 72
    for row, _case in enumerate(cases, 8):
        evidence_row = row - 6
        add_internal_hyperlink(
            wb,
            anchor(ws, f"BL{row}"),
            "Evidence",
            f"A{evidence_row}",
        )
    for column, width in enumerate((22, 18, 24, 42, 52, 54, 44, 28, 56, 48, 58), 1):
        evidence.column_dimensions[evidence.cell(1, column).column_letter].width = width
    evidence.freeze_panes = "A2"
    evidence.auto_filter.ref = f"A1:K{len(cases) + 1}"
    evidence.print_area = f"A1:K{len(cases) + 1}"
    evidence.page_setup.orientation = "landscape"
    evidence.page_setup.fitToWidth = 1
    evidence.page_setup.fitToHeight = 0

    trim_sheet(wb["Cover"], 20, 43, print_area="B8:AO20", orientation="landscape")
    trim_sheet(wb["Histories"], 3, 7, freeze="B3", auto_filter="B2:G3", print_area="B2:G3", orientation="landscape", title_rows="2:2")
    trim_sheet(ws, 12, 74, freeze="B8", auto_filter="B7:BV12", print_area="B2:BV12", orientation="landscape", fit_width=1, title_rows="6:7")
    trim_sheet(evidence, len(cases) + 1, 11, freeze="A2", auto_filter=f"A1:K{len(cases) + 1}", print_area=f"A1:K{len(cases) + 1}", orientation="landscape", fit_width=2, title_rows="1:1")
    save(wb, output, "Unit Test", lang, version)
    return output


def add_group_merges(ws, row, groups):
    for start, end in groups:
        ensure_merge(ws, f"{start}{row}:{end}{row}")


def generate_functional(catalog, lang):
    wb, output, version = prepare("functional", lang)
    cover(
        wb["Cover"],
        label("Functional Test", "Kiểm thử chức năng", lang),
        "IDTS-SAP490-FUNCTIONAL",
        label("Current CAP/Fiori functional verification", "Xác minh chức năng CAP/Fiori hiện hành", lang),
        catalog,
    )
    history(
        wb["Histories"],
        version,
        label(
            "Filled functional cases, exact programmatic runs and concrete test-data descriptions in the official template.",
            "Điền case chức năng, programmatic run chính xác và mô tả dữ liệu cụ thể trong template chính thức.",
            lang,
        ),
        "Test Cases, Test Result, Test Data Description",
        catalog,
    )
    cases = [
        case
        for case in catalog["cases"]
        if case["testType"] not in {"UNIT", "UAT"}
    ]
    ws = wb["Test Cases"]
    for coordinate, value in (
        ("B3", label("IDTS SAP490 functional verification", "Xác minh chức năng IDTS SAP490", lang)),
        ("L3", label("CAP/Fiori current baseline", "Baseline CAP/Fiori hiện hành", lang)),
        ("AO3", "DonHV"),
        ("AX3", datetime.fromisoformat(catalog["asOf"]).date()),
        ("BE3", label("Mentor / Supervisor", "Mentor / Giảng viên", lang)),
        ("BL3", label("Pending review", "Chờ review", lang)),
        ("BL6", label("Evidence / limitations", "Bằng chứng / giới hạn", lang)),
    ):
        write(ws, coordinate, value)
    for row in range(8, 34):
        old_evidence_merge = f"BL{row}:BR{row}"
        if old_evidence_merge in {
            str(item) for item in ws.merged_cells.ranges
        }:
            ws.unmerge_cells(old_evidence_merge)
        ensure_merge(ws, f"BL{row}:BV{row}")
        for coordinate in (f"B{row}", f"E{row}", f"Y{row}", f"AX{row}", f"BD{row}", f"BJ{row}", f"BL{row}"):
            write(ws, coordinate, None)
    for row, case in enumerate(cases, 8):
        write(ws, f"B{row}", case["caseId"], center=True)
        write(
            ws,
            f"E{row}",
            f"{local(case['title'], lang)} — {case['role']}\n"
            f"Req: {joined(case['requirementIds'])}\n"
            f"{label('Steps', 'Các bước', lang)}: {inline_steps(case['steps'], lang)}",
            font_size=12,
        )
        write(
            ws,
            f"Y{row}",
            f"{label('Preconditions', 'Điều kiện trước', lang)}: "
            f"{local(case['preconditions'], lang)}\n"
            f"{label('Data', 'Dữ liệu', lang)}: {local(case['testData'], lang)}\n"
            f"{label('Expected', 'Dự kiến', lang)}: {local(case['expectedResult'], lang)}\n"
            f"{label('Actual / postcondition', 'Thực tế / điều kiện sau', lang)}: "
            f"{local(case['actualResult'], lang) or label('Not run', 'Chưa chạy', lang)} | "
            f"{local(case['postcondition'], lang)}",
            font_size=12,
        )
        write(ws, f"AX{row}", case["executor"], center=True)
        write(ws, f"BD{row}", case["executionDate"], center=True)
        write(ws, f"BJ{row}", STATUS_DISPLAY[case["status"]], center=True)
        write(
            ws,
            f"BL{row}",
            f"{label('Command', 'Lệnh', lang)}: {case['automationCommand'] or 'N/A'}\n"
            f"{label('Evidence', 'Bằng chứng', lang)}: {joined(case['evidenceLinks']) or 'N/A'}\n"
            f"{label('Defects / limit', 'Defect / giới hạn', lang)}: "
            f"{joined(case['defectIds']) or 'None'} | "
            f"{local(case['limitations'], lang) or 'None'}",
            font_size=12,
        )
        ws.row_dimensions[row].height = 150

    runs = wb["Test Result"]
    for coordinate, value in (
        ("B4", label("Run ID / exact command", "Run ID / lệnh chính xác", lang)),
        ("L4", label("Observed result / limitation", "Kết quả quan sát / giới hạn", lang)),
        ("AO4", label("Executor", "Người thực thi", lang)),
        ("AX4", label("Execution date", "Ngày thực thi", lang)),
        ("BE4", label("Status / counts", "Trạng thái / số lượng", lang)),
        ("BL4", label("Evidence", "Bằng chứng", lang)),
    ):
        write(runs, coordinate, value, center=True)
    runs.row_dimensions[4].height = 36
    # Rows 5-6 are the official two-row Test Result block. Reuse that exact
    # merge/style signature for every run instead of placing data below it.
    clear_values(runs, 5, 40, 2, 70)
    for merged in list(runs.merged_cells.ranges):
        if merged.min_row >= 7:
            runs.unmerge_cells(str(merged))
    for index, run in enumerate(catalog["executionRuns"]):
        row = 5 + index * 2
        if index:
            clone_merged_row_block(runs, 5, 6, row, 70)
        write(runs, f"B{row}", f"{run['runId']}\n{run['command']}")
        write(
            runs,
            f"L{row}",
            f"{local(run['actualResult'], lang)}\n"
            f"{label('Limitation', 'Giới hạn', lang)}: {local(run['limitations'], lang)}",
        )
        write(runs, f"AO{row}", run["executor"], center=True)
        write(runs, f"AX{row}", run["executionDate"], center=True)
        write(
            runs,
            f"BE{row}",
            f"{run['status']}\nP/F/S: {run['passed']}/{run['failed']}/{run['skipped']}",
            center=True,
        )
        matching_case = next(
            (
                case for case in catalog["cases"]
                if case.get("automationCommand") == run["command"]
            ),
            None,
        )
        artifact_path = (
            evidence_artifact_path(matching_case, catalog)
            if matching_case else catalog["evidenceBaseline"]["defaultEvidencePath"]
        )
        write(runs, f"BL{row}", artifact_path)
        add_hyperlink(anchor(runs, f"BL{row}"), evidence_url(catalog, artifact_path))

    data = wb["Test Data Description"]
    for coordinate, value in (
        ("B3", "IDTS-SAP490-FUNCTIONAL"),
        ("L3", label("Concrete test data and environment", "Dữ liệu và môi trường test cụ thể", lang)),
        ("AO3", "DonHV"),
        ("AX3", datetime.fromisoformat(catalog["asOf"]).date()),
        ("BE3", label("Mentor / Supervisor", "Mentor / Giảng viên", lang)),
        ("BL3", label("Pending review", "Chờ review", lang)),
    ):
        write(data, coordinate, value)
    data_rows = [
        (
            "TD-01",
            label("Local CAP/SQLite regression", "Regression CAP/SQLite local", lang),
            label(
                "Local seeded users and in-memory/file SQLite as stated per case; no production environment claim.",
                "User seed local và SQLite in-memory/file theo từng case; không tuyên bố môi trường production.",
                lang,
            ),
        ),
        (
            "TD-02",
            label("Role and authorization data", "Dữ liệu role và authorization", lang),
            label(
                "PM, Tester and Developer seeded identities; no credentials are embedded in the workbook.",
                "Identity seed PM, Tester và Developer; workbook không chứa credential.",
                lang,
            ),
        ),
        (
            "TD-03",
            label("Classification and assignment data", "Dữ liệu phân loại và phân công", lang),
            label(
                "Concrete component/category/developer relationships are defined inside each linked case.",
                "Quan hệ component/category/developer cụ thể nằm trong từng case liên kết.",
                lang,
            ),
        ),
        (
            "TD-04",
            label("Attachment and notification data", "Dữ liệu attachment và notification", lang),
            label(
                "Small safe files, comment text, notification rows and outbox records; live provider delivery may be skipped.",
                "File nhỏ an toàn, comment, notification và outbox; live provider có thể bị skip.",
                lang,
            ),
        ),
        (
            "TD-05",
            label("AI advisory boundary", "Ranh giới AI advisory", lang),
            label(
                "Sanitized prompt/input, mocked or disabled provider modes, mandatory human review, no automatic bug mutation.",
                "Prompt/input đã sanitize, provider mock/disabled, bắt buộc human review, không tự đổi bug.",
                lang,
            ),
        ),
    ]
    groups = (("B", "K"), ("L", "AN"), ("AO", "AW"), ("AX", "BD"), ("BE", "BK"), ("BL", "BR"))
    for row, (data_id, title, description) in enumerate(data_rows, 8):
        add_group_merges(data, row, groups)
        write(data, f"B{row}", data_id, center=True)
        write(data, f"L{row}", f"{title}\n{description}")
        write(data, f"AO{row}", "DonHV", center=True)
        write(data, f"AX{row}", datetime.fromisoformat(catalog["asOf"]).date(), center=True)
        write(data, f"BE{row}", label("Not applicable", "Không áp dụng", lang), center=True)
        write(data, f"BL{row}", label("See linked case evidence", "Xem evidence của case liên kết", lang))
        set_case_row_height(data, row, 60)

    trim_sheet(wb["Cover"], 20, 43, print_area="B8:AO20", orientation="landscape")
    trim_sheet(wb["Histories"], 3, 7, freeze="B3", auto_filter="B2:G3", print_area="B2:G3", orientation="landscape", title_rows="2:2")
    trim_sheet(ws, 23, 74, freeze="B8", auto_filter="B7:BV23", print_area="B2:BV23", orientation="landscape", fit_width=1, title_rows="6:7")
    last_run_row = 4 + len(catalog["executionRuns"]) * 2
    trim_sheet(runs, last_run_row, 70, freeze="B5", auto_filter=None, print_area=f"B4:BR{last_run_row}", orientation="landscape", fit_width=1, title_rows="4:4")
    trim_sheet(data, 12, 70, freeze="B8", auto_filter="B7:BR12", print_area="B2:BR12", orientation="landscape", fit_width=1, title_rows="2:4")
    save(wb, output, "Functional Test", lang, version)
    return output


def feature_case_row(case, lang):
    procedure = " | ".join(
        [
            f"Req: {joined(case['requirementIds'])}",
            f"{label('Role', 'Vai trò', lang)}: {case['role']}",
            f"{label('Data', 'Dữ liệu', lang)}: {local(case['testData'], lang)}",
            inline_steps(case["steps"], lang),
        ]
    )
    note = " | ".join(
        [
            f"{label('Actual', 'Thực tế', lang)}: "
            f"{local(case['actualResult'], lang) or label('Not run', 'Chưa chạy', lang)}",
            f"{label('Evidence', 'Bằng chứng', lang)}: {joined(case['evidenceLinks']) or 'N/A'}",
            f"{label('Defects', 'Defect', lang)}: {joined(case['defectIds']) or 'None'}",
            f"{label('Limitations', 'Giới hạn', lang)}: {local(case['limitations'], lang) or 'None'}",
            f"{label('Postcondition', 'Điều kiện sau', lang)}: {local(case['postcondition'], lang)}",
        ]
    )
    return [
        case["caseId"],
        local(case["title"], lang),
        procedure,
        local(case["expectedResult"], lang),
        local(case["preconditions"], lang),
        STATUS_DISPLAY[case["status"]],
        case["executionDate"] or "",
        case["executor"] or "",
        "N/A",
        "",
        "",
        "N/A",
        "",
        "",
        note,
    ]


def fill_feature_sheet(ws, cases, feature_name, requirement_text, lang, header_row):
    name_row = 1 if ws.title == "Feature 1" else 2
    description_row = 2 if ws.title == "Feature 1" else 3
    count_row = 3 if ws.title == "Feature 1" else 4
    round_header_row = 4 if ws.title == "Feature 1" else 5
    round_one_row = 5 if ws.title == "Feature 1" else 6
    # Test Report links always target A1; keep that anchor non-empty and stable.
    write(ws, "A1", feature_name)
    write(ws, f"B{name_row}", feature_name)
    write(ws, f"B{description_row}", requirement_text)
    write(ws, "A11", label("Cases", "Ca kiểm thử", lang), bold=True)
    write(ws, f"B{count_row}", f"=COUNTA(A12:A{11 + len(cases)})")
    write(ws, f"B{round_header_row}", label("Passed", "Passed", lang), center=True)
    write(ws, f"C{round_header_row}", label("Failed", "Failed", lang), center=True)
    write(ws, f"D{round_header_row}", label("Pending", "Pending", lang), center=True)
    write(ws, f"E{round_header_row}", label("Blocked", "Blocked", lang), center=True)
    write(ws, f"B{round_one_row}", f'=COUNTIF($F$12:$F${11 + len(cases)},"Passed")')
    write(ws, f"C{round_one_row}", f'=COUNTIF($F$12:$F${11 + len(cases)},"Failed")')
    write(ws, f"D{round_one_row}", f'=COUNTIF($F$12:$F${11 + len(cases)},"Pending")')
    write(ws, f"E{round_one_row}", f'=COUNTIF($F$12:$F${11 + len(cases)},"Blocked")')
    clear_values(ws, 12, 200, 1, 15)
    for row, case in enumerate(cases, 12):
        if row > 19:
            copy_row_style(ws, 12, row, 18)
        for column, value in enumerate(feature_case_row(case, lang), 1):
            write(ws, ws.cell(row, column).coordinate, value, center=column in {1, 6, 7, 8})
        set_case_row_height(ws, row)
    for column, width in {
        "A": 18,
        "B": 45,
        "C": 85,
        "D": 48,
        "E": 46,
        "F": 12,
        "G": 18,
        "H": 16,
        "I": 10,
        "J": 14,
        "K": 14,
        "L": 10,
        "M": 14,
        "N": 14,
        "O": 62,
    }.items():
        ws.column_dimensions[column].width = width
    ws.row_dimensions[10].height = 42
    ws.freeze_panes = "A11"
    ws.auto_filter.ref = f"A10:O{11 + len(cases)}"
    ws.print_area = f"A1:O{11 + len(cases)}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 2
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:10"


def add_report_metrics(ws, catalog, lang):
    cases = catalog["cases"]
    planned = len(cases)
    passed = sum(case["status"] == "PASSED" for case in cases)
    failed = sum(case["status"] == "FAILED" for case in cases)
    blocked = sum(case["status"] == "BLOCKED" for case in cases)
    prepared = sum(case["status"] == "PREPARED" for case in cases)
    not_run = sum(case["status"] in {"NOT_RUN", "PENDING"} for case in cases)
    executed = passed + failed
    linked = {rid for case in cases for rid in case["requirementIds"]}
    executed_requirements = {
        rid
        for case in cases
        if case["status"] in {"PASSED", "FAILED"}
        for rid in case["requirementIds"]
    }
    assertions = sum(
        run["passed"] + run["failed"] + run["skipped"]
        for run in catalog["executionRuns"]
    )
    rows = [
        (label("Planned cases", "Case kế hoạch", lang), planned, f"{planned} canonical cases"),
        (label("Executed cases", "Case đã thực thi", lang), executed, f"{passed} Passed + {failed} Failed"),
        (label("Passed", "Đạt", lang), passed, "Fresh evidence and required execution metadata"),
        (label("Failed", "Không đạt", lang), failed, "Executed failed outcomes"),
        (label("Blocked", "Bị chặn", lang), blocked, "Execution blockers"),
        (
            label("Not Run / Pending", "Chưa chạy / Đang chờ", lang),
            not_run + prepared,
            f"{not_run} NOT_RUN + {prepared} PREPARED",
        ),
        (
            label("Planned requirement traceability", "Truy vết requirement theo kế hoạch", lang),
            len(linked) / len(catalog["requirements"]),
            f"{len(linked)} / {len(catalog['requirements'])}",
        ),
        (
            label("Executed requirement coverage", "Coverage requirement đã thực thi", lang),
            len(executed_requirements) / len(catalog["requirements"]),
            f"{len(executed_requirements)} / {len(catalog['requirements'])}",
        ),
        (
            label("Automated assertion/check count", "Số assertion/check tự động", lang),
            assertions,
            "Fresh suite-reported checks; skipped checks are not Pass",
        ),
        (
            label("Defect count", "Số defect", lang),
            len(catalog["defects"]),
            "Real tracked defects; no synthetic defect",
        ),
    ]
    for row in range(19, 30):
        clear_values(ws, row, row, 2, 8)
    for column, value in enumerate(
        [
            label("Metric", "Metric", lang),
            label("Value", "Giá trị", lang),
            label("Calculation / scope", "Cách tính / phạm vi", lang),
        ],
        2,
    ):
        cell = ws.cell(19, column, value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
    ensure_merge(ws, "D19:H19")
    for row, (metric, value, calculation) in enumerate(rows, 20):
        copy_row_style(ws, 18, row, 8)
        ensure_merge(ws, f"D{row}:H{row}")
        write(ws, f"B{row}", metric)
        write(ws, f"C{row}", value, center=True)
        write(ws, f"D{row}", calculation)
        if isinstance(value, float) and 0 <= value <= 1:
            ws[f"C{row}"].number_format = "0.00%"
        set_case_row_height(ws, row, 34)


def generate_report(catalog, lang):
    wb, output, version = prepare("report", lang)
    evidence_date = datetime.fromisoformat(catalog["asOf"]).date()
    cover_ws = wb["Cover"]
    for coordinate, value in (
        ("B4", "Issue and Defect Tracking System in SAP"),
        ("B5", "IDTS-SAP01"),
        ("B6", f"IDTS-SAP01_Test Report_v{version}"),
        ("F4", "DonHV"),
        ("F5", evidence_date),
        ("F6", version),
        ("A11", evidence_date),
        ("B11", version),
        ("C11", label("Full SAP490 test pack", "Toàn bộ test pack SAP490", lang)),
        ("D11", "M"),
        (
            "E11",
            label(
                "Reconciled from canonical cases and fresh exact npm runs; Pending/Prepared is not Passed.",
                "Đối soát từ case chuẩn và lệnh npm chính xác; Pending/Prepared không phải Passed.",
                lang,
            ),
        ),
        ("F11", "docs/qa/test-catalog.json"),
    ):
        write(cover_ws, coordinate, value)
    cover_ws.row_dimensions[4].height = 32
    cover_ws.row_dimensions[11].height = 48

    cases = catalog["cases"]
    executed_cases = [
        case for case in cases if case["status"] in {"PASSED", "FAILED", "BLOCKED"}
    ]
    pending_cases = [case for case in cases if case not in executed_cases]
    fill_feature_sheet(
        wb["Feature 1"],
        executed_cases,
        label("Executed unit/API/integration verification", "Xác minh unit/API/integration đã chạy", lang),
        label(
            "Freshly executed cases with exact command or direct execution evidence.",
            "Case đã thực thi mới với lệnh chính xác hoặc evidence trực tiếp.",
            lang,
        ),
        lang,
        10,
    )
    fill_feature_sheet(
        wb["Feature 2"],
        pending_cases,
        label("Manual functional and UAT preparation", "Chuẩn bị functional manual và UAT", lang),
        label(
            "Manual UI flows remain NOT_RUN and UAT remains PREPARED; no sign-off claim.",
            "Flow UI manual vẫn NOT_RUN và UAT vẫn PREPARED; không claim sign-off.",
            lang,
        ),
        lang,
        10,
    )

    test_cases = wb["Test Cases"]
    clear_values(test_cases, 9, 100, 2, 6)
    write(test_cases, "D3", "Issue and Defect Tracking System in SAP")
    write(test_cases, "D4", "IDTS-SAP01")
    write(
        test_cases,
        "D5",
        label(
            "Local Node.js/CAP/SQLite; seeded roles; exact environment is recorded per case.",
            "Node.js/CAP/SQLite local; role seed; môi trường chính xác ghi theo từng case.",
            lang,
        ),
    )
    for row, case in enumerate(cases, 9):
        if row > 13:
            copy_row_style(test_cases, 13, row, 6)
        sheet_name = "Feature 1" if case in executed_cases else "Feature 2"
        write(test_cases, f"B{row}", row - 8, center=True)
        write(test_cases, f"C{row}", f"{case['caseId']} — {local(case['title'], lang)}")
        write(test_cases, f"D{row}", sheet_name)
        test_cases[f"D{row}"].font = copy(test_cases[f"F{row}"].font)
        add_internal_hyperlink(wb, test_cases[f"D{row}"], sheet_name)
        write(
            test_cases,
            f"E{row}",
            f"{joined(case['requirementIds'])}\n{case['testType']} / {case['status']}",
        )
        write(test_cases, f"F{row}", local(case["preconditions"], lang))
        # Requirement/status text is plain text unless it has a real target.
        # Copy the neighboring template font so it cannot look clickable.
        normal_font = copy(test_cases[f"F{row}"].font)
        normal_font.color = "000000"
        normal_font.underline = None
        test_cases[f"E{row}"].font = normal_font
        set_case_row_height(test_cases, row, 66)
    for column, width in {"B": 7, "C": 48, "D": 16, "E": 40, "F": 55}.items():
        test_cases.column_dimensions[column].width = width
    test_cases.freeze_panes = "B9"
    test_cases.auto_filter.ref = f"B8:F{8 + len(cases)}"
    test_cases.print_area = f"B1:F{8 + len(cases)}"
    test_cases.print_title_rows = "1:8"

    stats = wb["Test Statistics"]
    passed = sum(case["status"] == "PASSED" for case in cases)
    failed = sum(case["status"] == "FAILED" for case in cases)
    blocked = sum(case["status"] == "BLOCKED" for case in cases)
    pending = len(cases) - passed - failed - blocked
    for coordinate, value in (
        ("C3", "Issue and Defect Tracking System in SAP"),
        ("E3", "DonHV"),
        ("C4", "IDTS-SAP01"),
        ("E4", label("Mentor / Supervisor", "Mentor / Giảng viên", lang)),
        ("C5", f"IDTS-SAP01_Test Report_v{version}"),
        ("H5", evidence_date),
        (
            "C6",
            label(
                f"{len(cases)} planned; {passed + failed} executed; {passed} passed; "
                f"{failed} failed; {blocked} blocked; {pending} Not Run/Prepared. "
                "Pending is not Passed.",
                f"{len(cases)} kế hoạch; {passed + failed} đã chạy; {passed} đạt; "
                f"{failed} không đạt; {blocked} bị chặn; {pending} Chưa chạy/Prepared. "
                "Pending không phải Passed.",
                lang,
            ),
        ),
        ("C11", wb["Feature 1"]["B1"].value),
        ("C12", wb["Feature 2"]["B2"].value),
        ("G10", label("Blocked", "Bị chặn", lang)),
        ("C16", label("Case execution rate", "Tỷ lệ thực thi case", lang)),
        ("E16", "=IFERROR((D14+E14)/(H14),0)"),
        ("F16", None),
        ("C17", label("Pass rate over executed cases", "Tỷ lệ đạt trên case đã chạy", lang)),
        ("E17", "=IFERROR(D14/(D14+E14),0)"),
        ("F17", None),
    ):
        write(stats, coordinate, value)
    stats["D11"] = f'=COUNTIF(\'Feature 1\'!$F$12:$F${11 + len(executed_cases)},"Passed")'
    stats["E11"] = f'=COUNTIF(\'Feature 1\'!$F$12:$F${11 + len(executed_cases)},"Failed")'
    stats["F11"] = f'=COUNTIF(\'Feature 1\'!$F$12:$F${11 + len(executed_cases)},"Pending")'
    stats["G11"] = f'=COUNTIF(\'Feature 1\'!$F$12:$F${11 + len(executed_cases)},"Blocked")'
    stats["H11"] = f"=SUM(D11:G11)"
    stats["D12"] = f'=COUNTIF(\'Feature 2\'!$F$12:$F${11 + len(pending_cases)},"Passed")'
    stats["E12"] = f'=COUNTIF(\'Feature 2\'!$F$12:$F${11 + len(pending_cases)},"Failed")'
    stats["F12"] = f'=COUNTIF(\'Feature 2\'!$F$12:$F${11 + len(pending_cases)},"Pending")'
    stats["G12"] = f'=COUNTIF(\'Feature 2\'!$F$12:$F${11 + len(pending_cases)},"Blocked")'
    stats["H12"] = f"=SUM(D12:G12)"
    for column in range(4, 9):
        letter = stats.cell(14, column).column_letter
        stats.cell(14, column, f"=SUM({letter}11:{letter}12)")
    stats["E16"].number_format = "0.00%"
    stats["E17"].number_format = "0.00%"
    add_report_metrics(stats, catalog, lang)
    for row, height in {
        3: 30,
        4: 24,
        5: 24,
        6: 36,
        10: 34,
        16: 28,
        17: 28,
        19: 32,
        26: 44,
        27: 44,
        28: 44,
    }.items():
        stats.row_dimensions[row].height = height
    for column, width in {
        "B": 34,
        "C": 34,
        "D": 20,
        "E": 20,
        "F": 18,
        "G": 18,
        "H": 22,
    }.items():
        stats.column_dimensions[column].width = width
    stats.print_area = "B1:H29"
    stats.page_setup.fitToWidth = 1
    stats.page_setup.fitToHeight = 0

    trim_sheet(cover_ws, 17, 6, print_area="A1:F17", orientation="landscape")
    trim_sheet(test_cases, 35, 6, freeze="B9", auto_filter="B8:F35", print_area="B1:F35", orientation="landscape", title_rows="1:8")
    trim_sheet(stats, 29, 8, freeze="B10", auto_filter="B10:H14", print_area="B1:H29", orientation="landscape", title_rows="1:10")
    # The executed/pending split changes as evidence matures. Keep every
    # canonical case instead of trimming Feature 1 to the old 12-case baseline.
    feature_one_last_row = max(23, 11 + len(executed_cases))
    feature_two_last_row = max(26, 11 + len(pending_cases))
    trim_sheet(
        wb["Feature 1"], feature_one_last_row, 18,
        freeze="A11", auto_filter=f"A10:O{feature_one_last_row}",
        print_area=f"A1:O{feature_one_last_row}", orientation="landscape",
        fit_width=2, title_rows="1:10",
    )
    trim_sheet(
        wb["Feature 2"], feature_two_last_row, 18,
        freeze="A11", auto_filter=f"A10:O{feature_two_last_row}",
        print_area=f"A1:O{feature_two_last_row}", orientation="landscape",
        fit_width=2, title_rows="1:10",
    )
    save(wb, output, "Test Report", lang, version)
    return output


def generate_uat(catalog, lang):
    wb, output, version = prepare("uat", lang)
    cover(
        wb["Cover"],
        label("User Acceptance Test (UAT)", "Kiểm thử chấp nhận người dùng (UAT)", lang),
        "IDTS-SAP490-UAT",
        label("Prepared mentor/user acceptance cases", "Case UAT đã chuẩn bị cho mentor/user", lang),
        catalog,
    )
    history(
        wb["Histories"],
        version,
        label(
            "Prepared executable UAT cases; actual result, defect, decision, tester, date and sign-off remain blank.",
            "Chuẩn bị case UAT có thể thực thi; actual result, defect, decision, tester, ngày và sign-off để trống.",
            lang,
        ),
        "Test Scenario, Test Cases, Test Result",
        catalog,
    )
    cases = [case for case in catalog["cases"] if case["testType"] == "UAT"]
    scenario = wb["Test Scenario"]
    if "Q4:S21" in {str(item) for item in scenario.merged_cells.ranges}:
        scenario.unmerge_cells("Q4:S21")
    clear_values(scenario, 4, 22, 2, 19)
    for coordinate, value in (
        ("B2", label("Case ID", "Mã case", lang)),
        ("C2", label("Persona / role", "Persona / vai trò", lang)),
        ("E2", label("Scenario and acceptance checkpoints", "Kịch bản và checkpoint chấp nhận", lang)),
        ("Q2", label("State / limitation", "Trạng thái / giới hạn", lang)),
    ):
        write(scenario, coordinate, value, center=True)
    for row, case in enumerate(cases, 4):
        for start, end in (("E", "G"), ("H", "J"), ("K", "M"), ("N", "P"), ("Q", "S")):
            ensure_merge(scenario, f"{start}{row}:{end}{row}")
        write(scenario, f"A{row}", row - 3, center=True)
        write(scenario, f"B{row}", case["caseId"], center=True)
        write(scenario, f"C{row}", case["role"])
        write(scenario, f"E{row}", local(case["title"], lang))
        write(scenario, f"H{row}", local(case["preconditions"], lang))
        write(scenario, f"K{row}", local(case["testData"], lang))
        write(scenario, f"N{row}", local(case["expectedResult"], lang))
        write(
            scenario,
            f"Q{row}",
            label(
                "PREPARED — not executed; mentor/user sign-off pending",
                "PREPARED — chưa thực thi; chờ mentor/user sign-off",
                lang,
            ),
        )
        set_case_row_height(scenario, row, 72)
    scenario.row_dimensions[3].height = 54

    ws = wb["Test Cases"]
    for coordinate, value in (
        ("B3", label("IDTS SAP490 prepared UAT", "UAT IDTS SAP490 đã chuẩn bị", lang)),
        ("L3", label("Mentor/user acceptance execution", "Thực thi chấp nhận mentor/user", lang)),
        ("BF3", "DonHV"),
        ("BO3", datetime.fromisoformat(catalog["asOf"]).date()),
        ("BV3", label("Mentor / User", "Mentor / Người dùng", lang)),
        ("CC3", label("Pending execution", "Chờ thực thi", lang)),
        ("BO7", label("Actual Result", "Kết quả thực tế", lang)),
        ("BU7", label("Defect ID", "Defect ID", lang)),
        ("CA7", label("Decision", "Quyết định", lang)),
        ("CC6", label("Tester / Date / Sign-off", "Tester / Ngày / Sign-off", lang)),
    ):
        write(ws, coordinate, value, center=coordinate in {"BO7", "BU7", "CA7"})
    ws.row_dimensions[7].height = 32
    add_group_merges(
        ws,
        13,
        (("B", "D"), ("E", "X"), ("Y", "AO"), ("AP", "BN"), ("BO", "BT"), ("BU", "BZ"), ("CA", "CB"), ("CC", "CI")),
    )
    copy_row_style(ws, 12, 13, 91)
    for row in range(8, 14):
        for coordinate in (f"B{row}", f"E{row}", f"Y{row}", f"AP{row}", f"BO{row}", f"BU{row}", f"CA{row}", f"CC{row}"):
            write(ws, coordinate, None)
    for row, case in enumerate(cases, 8):
        write(ws, f"B{row}", case["caseId"], center=True)
        write(
            ws,
            f"E{row}",
            f"{local(case['title'], lang)} — {case['role']}\n"
            f"Req: {joined(case['requirementIds'])}\n"
            f"{label('Objective', 'Mục tiêu', lang)}: {local(case['objective'], lang)}\n"
            f"PREPARED — {label('not executed; no sign-off', 'chưa thực thi; chưa sign-off', lang)}",
            font_size=12,
        )
        write(
            ws,
            f"Y{row}",
            f"{label('Preconditions', 'Điều kiện trước', lang)}: "
            f"{local(case['preconditions'], lang)}\n"
            f"{label('Data', 'Dữ liệu', lang)}: {local(case['testData'], lang)}\n"
            f"{label('Steps', 'Các bước', lang)}: {inline_steps(case['steps'], lang)}",
            font_size=12,
        )
        write(
            ws,
            f"AP{row}",
            f"{expected_block(case, lang, uat=True)}",
            font_size=12,
        )
        # Actual, defect, decision, tester/date/sign-off are intentionally blank.
        ws.row_dimensions[row].height = 125

    result = wb["Test Result"]
    for coordinate, value in (
        ("A2", label("Test Case ID", "Mã test", lang)),
        ("C2", label("Actual Result", "Kết quả thực tế", lang)),
        ("I2", label("Defect ID", "Defect ID", lang)),
        ("K2", label("Decision", "Quyết định", lang)),
        ("M2", label("Tester / Date", "Tester / Ngày", lang)),
        ("O2", label("Sign-off", "Sign-off", lang)),
    ):
        write(result, coordinate, value, center=True)
    clear_values(result, 3, 20, 1, 16)
    for index, case in enumerate(cases):
        start = 3 + index * 2
        end = start + 1
        if index:
            copy_row_style(result, 3, start, 16)
            copy_row_style(result, 4, end, 16)
            for first, last in (("A", "B"), ("C", "H"), ("I", "J"), ("K", "L"), ("M", "N"), ("O", "P")):
                ensure_merge(result, f"{first}{start}:{last}{end}")
        write(result, f"A{start}", case["caseId"], center=True)
        result.row_dimensions[start].height = 30
        result.row_dimensions[end].height = 30
    result.freeze_panes = "A3"
    result.print_area = f"A1:P{2 + len(cases) * 2}"
    result.page_setup.orientation = "landscape"
    result.page_setup.fitToWidth = 1
    result.page_setup.fitToHeight = 0

    trim_sheet(wb["Cover"], 20, 43, print_area="B8:AO20", orientation="landscape")
    trim_sheet(wb["Histories"], 3, 7, freeze="B3", auto_filter="B2:G3", print_area="B2:G3", orientation="landscape", title_rows="2:2")
    trim_sheet(scenario, 9, 19, freeze="E4", auto_filter="A3:S9", print_area="A1:S9", orientation="landscape", fit_width=1, title_rows="1:3")
    trim_sheet(ws, 13, 87, freeze="B8", auto_filter="B7:CI13", print_area="B2:CI13", orientation="landscape", fit_width=1, title_rows="6:7")
    trim_sheet(result, 14, 16, freeze="A3", auto_filter="A2:P14", print_area="A1:P14", orientation="landscape", title_rows="1:2")
    save(wb, output, "UAT Prepared", lang, version)
    return output


def defect_block(defect, lang):
    summary = " | ".join(
        [
            f"{defect['defectId']} — {local(defect['summary'], lang)}",
            f"{label('Classification', 'Phân loại', lang)}: {defect['classification']}",
            f"{label('Severity / priority / status', 'Severity / priority / trạng thái', lang)}: "
            f"{defect['severity']} / {defect['priority']} / {defect['status']}",
            f"{label('Reporter / owner', 'Reporter / owner', lang)}: "
            f"{defect['reporter']} / {defect['owner']}",
            f"{label('Discovered / closed', 'Phát hiện / đóng', lang)}: "
            f"{defect['discoveredDate']} / {defect['closedDate'] or 'Open'}",
        ]
    )
    details = " | ".join(
        [
            f"{label('Preconditions', 'Điều kiện trước', lang)}: {local(defect['preconditions'], lang)}",
            f"{label('Steps to reproduce', 'Các bước tái hiện', lang)}: {local(defect['stepsToReproduce'], lang)}",
            f"{label('Actual result', 'Kết quả thực tế', lang)}: {local(defect['actualResult'], lang)}",
            f"{label('Environment / build', 'Môi trường / build', lang)}: {defect['environment']}",
        ]
    )
    expected = " | ".join(
        [
            f"{label('Expected result', 'Kết quả dự kiến', lang)}: {local(defect['expectedResult'], lang)}",
            f"{label('Requirement IDs', 'Requirement IDs', lang)}: {joined(defect['requirementIds'])}",
            f"{label('Related test case IDs', 'Test case IDs liên quan', lang)}: {joined(defect['testCaseIds'])}",
        ]
    )
    fix = " | ".join(
        [
            f"{label('Root cause', 'Nguyên nhân gốc', lang)}: {local(defect['rootCause'], lang)}",
            f"{label('Fix description', 'Mô tả sửa lỗi', lang)}: {local(defect['fixDescription'], lang)}",
            f"{label('Retest command / steps', 'Lệnh / bước retest', lang)}: {defect['retestCommand']}",
            f"{label('Retest result / date', 'Kết quả / ngày retest', lang)}: "
            f"{local(defect['retestResult'], lang)} / {defect['retestDate'] or 'N/A'}",
            f"{label('Remaining action', 'Việc còn lại', lang)}: {local(defect['remainingAction'], lang) or 'None'}",
        ]
    )
    evidence = "\n".join(defect["evidenceLinks"])
    return summary, details, expected, fix, evidence


def generate_defect(catalog, lang):
    wb, output, version = prepare("defect", lang)
    for unused in ("Issue 2", "Issue 4"):
        if unused in wb.sheetnames:
            wb.remove(wb[unused])
    ws = wb["Fix and bugs"]
    for coordinate, value in (
        ("A1", label("No", "STT", lang)),
        ("B1", label("Defect / ownership", "Defect / ownership", lang)),
        ("C1", label("Reproduction / actual", "Tái hiện / thực tế", lang)),
        ("D1", label("Expected / traceability", "Dự kiến / truy vết", lang)),
        ("E1", label("Root cause / fix / retest", "Nguyên nhân / sửa / retest", lang)),
        ("F1", f"Evidence\nv{version} — {catalog['asOf']}"),
    ):
        write(ws, coordinate, value, center=True, bold=True)
    for column, width in {
        "A": 7,
        "B": 48,
        "C": 58,
        "D": 52,
        "E": 78,
        "F": 50,
    }.items():
        ws.column_dimensions[column].width = width
    ws.row_dimensions[1].height = 48
    clear_values(ws, 2, 40, 1, 6)
    rows = [2, 3, 5, 6, 8] + list(range(9, 16))
    for row in rows[5:]:
        copy_row_style(ws, 8, row, 6)
    for sequence, (row, defect) in enumerate(zip(rows, catalog["defects"]), 1):
        summary, details, expected, fix, evidence = defect_block(defect, lang)
        write(ws, f"A{row}", sequence, center=True)
        write(ws, f"B{row}", summary, font_size=12)
        write(ws, f"C{row}", details, font_size=12)
        write(ws, f"D{row}", expected, font_size=12)
        write(ws, f"E{row}", fix, font_size=12)
        write(ws, f"F{row}", evidence, font_size=12)
        ws.row_dimensions[row].height = 90
    ws.row_dimensions[4].height = 45
    ws.row_dimensions[7].height = 45
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:F15"
    ws.print_area = "A1:F15"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 2
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:1"
    trim_sheet(ws, 15, 6, freeze="A2", auto_filter="A1:F15", print_area="A1:F15", orientation="landscape", fit_width=2, title_rows="1:1")
    save(wb, output, "Test And Fix Bug", lang, version)
    return output


def enforce_mentor_readability(wb, minimum_font_size=12):
    """Make 12pt the workbook default without discarding template styling."""
    for named_style in wb._named_styles:
        if named_style.name == "Normal":
            normal_font = copy(named_style.font)
            normal_font.sz = minimum_font_size
            named_style.font = normal_font
            break

    for ws in wb.worksheets:
        for row in ws.iter_rows(
            min_row=1,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column,
        ):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if not cell.has_style and cell.value in (None, ""):
                    continue
                current_size = float(cell.font.sz or 11)
                if current_size < minimum_font_size:
                    readable_font = copy(cell.font)
                    readable_font.sz = minimum_font_size
                    cell.font = readable_font

        if ws.title == "Test Cases" and ws.max_column == 74:
            for column in ("AX", "AY", "AZ", "BA", "BB", "BC", "BD", "BE", "BF", "BG", "BH", "BI"):
                ws.column_dimensions[column].width = 4
            for column in ("BJ", "BK"):
                ws.column_dimensions[column].width = 10
            for column in ("BL", "BM", "BN", "BO", "BP", "BQ", "BR", "BS", "BT", "BU", "BV"):
                ws.column_dimensions[column].width = 4
            ws.row_dimensions[7].height = max(
                ws.row_dimensions[7].height or 15,
                54,
            )
            data_height = 165 if ws.max_row <= 12 else 180
            for row_number in range(8, ws.max_row + 1):
                ws.row_dimensions[row_number].height = max(
                    ws.row_dimensions[row_number].height or 15,
                    data_height,
                )

        if ws.title == "Test Cases" and ws.max_column == 87:
            for column_index in range(5, 67):
                column = get_column_letter(column_index)
                ws.column_dimensions[column].width = max(
                    ws.column_dimensions[column].width or 0,
                    4,
                )
            for column in ("BO", "BP", "BQ", "BR", "BS", "BT", "BU", "BV", "BW", "BX", "BY", "BZ"):
                ws.column_dimensions[column].width = max(
                    ws.column_dimensions[column].width or 0,
                    4,
                )
            ws.column_dimensions["CA"].width = 12
            for column in ("CC", "CD", "CE", "CF", "CG", "CH", "CI"):
                ws.column_dimensions[column].width = max(
                    ws.column_dimensions[column].width or 0,
                    4,
                )
            ws.row_dimensions[7].height = max(
                ws.row_dimensions[7].height or 15,
                54,
            )
            data_height = 165 if ws.max_row <= 13 else 130
            for row_number in range(8, ws.max_row + 1):
                ws.row_dimensions[row_number].height = max(
                    ws.row_dimensions[row_number].height or 15,
                    data_height,
                )

        if ws.title == "Evidence":
            for row_number in range(2, ws.max_row + 1):
                ws.row_dimensions[row_number].height = max(
                    ws.row_dimensions[row_number].height or 15,
                    90,
                )

        if ws.title == "Test Result" and ws.max_column == 70:
            for column_index in range(12, 40):
                column = get_column_letter(column_index)
                ws.column_dimensions[column].width = max(
                    ws.column_dimensions[column].width or 0,
                    3,
                )
            # Preserve the official two-row result block heights cloned from
            # rows 5-6; changing every generated row to one uniform height
            # would break the template signature again.

        if ws.title == "Cover" and ws.max_column == 6:
            ws.row_dimensions[6].height = max(
                ws.row_dimensions[6].height or 15,
                36,
            )
            ws.row_dimensions[11].height = max(
                ws.row_dimensions[11].height or 15,
                60,
            )

        if ws.title == "Test Statistics":
            for row_number in (16, 17):
                ws.row_dimensions[row_number].height = max(
                    ws.row_dimensions[row_number].height or 15,
                    36,
                )

        if ws.title in {"Feature 1", "Feature 2"}:
            for row_number in range(1, 5):
                ws.row_dimensions[row_number].height = max(
                    ws.row_dimensions[row_number].height or 15,
                    36,
                )
            for row_number in range(12, ws.max_row + 1):
                ws.row_dimensions[row_number].height = max(
                    ws.row_dimensions[row_number].height or 15,
                    110,
                )

        if ws.title == "Fix and bugs":
            for row_number in range(2, ws.max_row + 1):
                if any(
                    ws.cell(row_number, column).value not in (None, "")
                    for column in range(1, ws.max_column + 1)
                ):
                    ws.row_dimensions[row_number].height = max(
                        ws.row_dimensions[row_number].height or 15,
                        120,
                    )


def save(wb, output, title, lang, version):
    wb.properties.title = f"IDTS SAP490 {title} {lang.upper()} v{version}"
    wb.properties.creator = "IDTS SAP01 Team"
    wb.properties.lastModifiedBy = "IDTS SAP01 Team"
    wb.properties.description = (
        "Official-template-filled current mentor-review artifact generated from "
        "docs/qa/test-catalog.json."
    )
    enforce_mentor_readability(wb)
    remove_all_filters(wb)
    for ws in wb.worksheets:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    wb.save(output)


def remove_all_filters(wb):
    """Remove filter buttons that clutter the official merged templates."""
    for ws in wb.worksheets:
        ws.auto_filter.ref = None
        ws.sheet_properties.filterMode = False
        for table in ws.tables.values():
            table.autoFilter = None


def generate_integration_evidence_index(catalog):
    """Create the mentor-facing cross-workbook evidence catalog."""
    output = OUTPUT_DIR / "Integration_Evidence_Index_IDTS_SAP01_v0.1.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Evidence Index"
    headers = [
        "Evidence ID / Mã bằng chứng",
        "Case ID / Mã ca",
        "Run ID",
        "Command / Lệnh",
        "Baseline commit",
        "Deploy ID / SHA",
        "Environment / executor / time",
        "Result / Kết quả",
        "Actual result / Kết quả thực tế",
        "Limitation / Giới hạn",
        "Evidence artifact / Bằng chứng",
        "Source or test at baseline / Nguồn hoặc test",
    ]
    for column, value in enumerate(headers, 1):
        cell = ws.cell(1, column, value)
        cell.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
    baseline = catalog["evidenceBaseline"]
    commit_url = f"{baseline['repositoryUrl']}/commit/{baseline['runtimeCommitSha']}"
    for row, case in enumerate(catalog["cases"], 2):
        record = evidence_record(case, catalog)
        source_path = next(
            (
                path for path in case.get("evidenceLinks", [])
                if path.startswith("scripts/") and not path.endswith("/")
            ),
            None,
        )
        values = [
            record["evidenceId"],
            record["caseId"],
            record["runId"],
            record["command"],
            record["baseline"],
            record["deploy"],
            record["context"],
            record["result"],
            record["actual"],
            record["limitation"],
            record["artifactPath"],
            source_path or "N/A — manual or artifact-based verification",
        ]
        for column, value in enumerate(values, 1):
            cell = ws.cell(row, column, value)
            cell.font = Font(name="Arial", size=12)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
        add_hyperlink(ws.cell(row, 5), commit_url)
        add_hyperlink(ws.cell(row, 11), record["artifactUrl"])
        if source_path:
            add_hyperlink(ws.cell(row, 12), source_url(catalog, source_path))
        ws.row_dimensions[row].height = 78
    widths = (23, 22, 25, 43, 52, 54, 48, 32, 62, 52, 62, 58)
    for column, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{len(catalog['cases']) + 1}"
    ws.print_area = f"A1:L{len(catalog['cases']) + 1}"
    ws.print_title_rows = "1:1"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 2
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    wb.properties.title = "IDTS SAP490 Integration Evidence Index v0.1"
    wb.properties.creator = "IDTS SAP01 Team"
    wb.properties.lastModifiedBy = "IDTS SAP01 Team"
    wb.save(output)
    return output


def main():
    catalog = load_catalog()
    outputs = []
    generators = (
        generate_scenario,
        generate_unit,
        generate_functional,
        generate_report,
        generate_uat,
        generate_defect,
    )
    for lang in ("en",):
        for generator in generators:
            outputs.append(generator(catalog, lang))
    outputs.append(generate_integration_evidence_index(catalog))
    for output in outputs:
        print(output.relative_to(ROOT))


if __name__ == "__main__":
    main()
