"""Generate the mentor-facing Technical Specification from the official template.

This module deliberately fills the template's existing sections instead of
replacing them with a new workbook design.  Technical detail that does not fit
an existing screen/message block is placed in Technical Implementation.
"""

from __future__ import annotations

from copy import copy
from datetime import date
from pathlib import Path
import csv
import re
import shutil

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment

from specification_catalog import AI_FUNCTIONS, LIFECYCLE_ACTIONS


ROOT = Path(__file__).resolve().parents[2]
TEMPLATE = ROOT / "docs" / "sap490" / "templates" / "Deliverable_template" / "Technical_Specification.xlsx"
OUTPUT_DIR = ROOT / "docs" / "sap490" / "generated"
VERSION = "0.8"
DOCUMENT_DATE = date(2026, 8, 4)

IDTS107_DIR = ROOT / "docs" / "pm" / "evidence" / "idts-107" / "technical-spec"
IDTS108_README = ROOT / "docs" / "pm" / "evidence" / "idts-108" / "README.md"
IDTS109_DIR = ROOT / "docs" / "pm" / "evidence" / "idts-109" / "technical-spec"


def _markdown_table(path, heading):
    """Read one governed Markdown table by heading without adding a parser dependency."""
    lines = Path(path).read_text(encoding="utf-8").splitlines()
    start = next(i for i, line in enumerate(lines) if line.strip() == heading)
    table = []
    for line in lines[start + 1:]:
        if line.startswith("##") and table:
            break
        if line.startswith("|"):
            table.append([cell.strip() for cell in line.strip().strip("|").split("|")])
        elif table and line.strip():
            break
    if len(table) < 2:
        raise ValueError(f"No table found below {heading!r} in {path}")
    return table[0], table[2:]


def _database_dictionary_rows():
    with (IDTS107_DIR / "database-dictionary.en.csv").open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _technical_implementation_sections():
    """Parse IDTS-109's approved 14-part traces into formal vertical blocks."""
    text = (IDTS109_DIR / "technical-implementation.md").read_text(encoding="utf-8")
    sections = []
    current = None
    for line in text.splitlines():
        match = re.match(r"^###\s+(.+)$", line)
        if match:
            if current:
                sections.append(current)
            current = {"title": match.group(1), "trace": "", "items": []}
            continue
        trace = re.match(r"^Technical trace ID:\s+(.+)$", line)
        if current and trace:
            current["trace"] = trace.group(1).replace("`", "")
            continue
        item = re.match(r"^(\d+)\.\s+\*\*(.+?):\*\*\s*(.*)$", line)
        if current and item:
            current["items"].append((item.group(1), item.group(2), item.group(3)))
    if current:
        sections.append(current)
    return [section for section in sections if section["items"]]


def _supplemental_implementation_sections():
    """Add full 14-part traces for core Bug writes and exact lifecycle actions.

    IDTS-109 owns the platform, collaboration, monitoring and AI traces.  The
    lifecycle source package deliberately groups those actions, while the
    mentor requires one independently reviewable block per action.  These
    deterministic blocks preserve that distinction without inventing runtime
    behavior.
    """
    sections = []

    def add(title, trace, actor, trigger, request, contract, handler, validation,
            transaction, side_effect, response, failure, evidence):
        sections.append({
            "title": title,
            "trace": trace,
            "items": [
                ("1", "Function name", title.split(" ", 1)[-1]),
                ("2", "Purpose", f"Execute {title.split(' ', 1)[-1]} through the governed Bug workflow."),
                ("3", "Actor/precondition", actor),
                ("4", "UI trigger", trigger),
                ("5", "Frontend source", "Fiori Elements Object Page actions defined in app/bug-management-ui/annotations/actions.cds and the supported controller extensions."),
                ("6", "HTTP/OData request", request),
                ("7", "Service contract", contract),
                ("8", "CAP handler/helper", handler),
                ("9", "Validation/authorization", validation),
                ("10", "Transaction", transaction),
                ("11", "Database/provider side effect", side_effect),
                ("12", "Response/UI refresh", response),
                ("13", "Failure/rollback", failure),
                ("14", "Test/evidence", evidence),
            ],
        })

    add(
        "7.1 Create and activate a Bug", "FLOW-DRAFT-CREATE",
        "Tester or PM with an active IDTS identity; required catalogs and fields must be available.",
        "Choose Create, complete the draft form and choose Create/Save.",
        "OData V4 draft NEW/PATCH/SAVE followed by active CREATE.",
        "srv/service.cds::Bugs draft projection and CAP draft events.",
        "srv/service.js draft registration; srv/bug-service/drafts.js; srv/bug-service/bug-write.js::prepareBugWrite.",
        "Server-owned reporter, required fields, active code lists, component/category mapping and role authorization.",
        "CAP request transaction covers active Bug, history and in-app notification persistence.",
        "Insert the active Bug and audit/notification rows in SAP HANA Cloud/HDI; no attachment binary is stored in HANA.",
        "Return the active Bug context and refresh the Object Page.",
        "Any validation or persistence failure rolls back the active write and returns a sanitized field/action message.",
        "IDTS-110 Unit Test candidate evidence and IDTS-111 UAT review evidence; blocked/held cases retain their recorded disposition.",
    )
    add(
        "7.2 Edit and save an active Bug", "FLOW-ACTIVE-EDIT",
        "Authorized participant editing an existing active Bug through its draft.",
        "Choose Edit, change permitted fields and choose Save.",
        "OData V4 EDIT/PATCH/SAVE followed by active UPDATE.",
        "srv/service.cds::Bugs draft projection and active UPDATE contract.",
        "srv/bug-service/drafts.js::prepareDraftPatch; srv/bug-service/bug-write.js::prepareBugWrite.",
        "Role, immutable/server-owned fields, catalogs, assignment responsibility and lifecycle rules.",
        "CAP request transaction covers the active update and its audit/notification side effects.",
        "Update the Bug in SAP HANA Cloud/HDI and append the applicable history/notification rows.",
        "Refresh the active Object Page with committed values.",
        "Failure rolls back the update; no partial Bug/history/notification state is accepted.",
        "IDTS-110 Unit Test candidate evidence and IDTS-111 UAT review matrix at the frozen integration baseline.",
    )

    for index, (action, actors, before, after) in enumerate(LIFECYCLE_ACTIONS, start=1):
        trace = f"FLOW-{action.upper()}"
        add(
            f"8.{index} {action}", trace,
            f"{actors}; current status must be one of: {before}.",
            f"Choose the {action} Object Page action and confirm any required parameters.",
            f"POST the bound OData V4 action BugService.{action} for the selected Bug.",
            f"srv/service.cds::Bugs.{action} bound action.",
            f"srv/service.js action registration; srv/bug-service/actions.js::{action if action in ('assignToDeveloper', 'resubmitToDeveloper') else 'transitionBug'}.",
            f"Validate actor role/ownership, current status {before}, required reason/note and assignee rules before deriving {after}.",
            "One CAP request transaction covers Bug state, HistoryEvents/HistoryLogs and in-app notification changes.",
            f"Update the Bug to {after} when permitted and append governed audit/notification rows in SAP HANA Cloud/HDI.",
            "Return the committed Bug and let Fiori Elements refresh the affected context and side-effect sections.",
            "Authorization, stale state, missing parameter or persistence failure rolls back the entire action; no partial lifecycle mutation remains.",
            "IDTS-89 lifecycle programmatic evidence plus the reviewed IDTS-110/111 case disposition; commands alone are not the acceptance artifact.",
        )

    add(
        "9.1 Read Bug history", "FLOW-HISTORY-READ",
        "An authorized Bug participant opens an active Bug with stored history.",
        "Open History or choose Show More.",
        "GET the Bug HistoryEvents/HistoryLogs composition through BugService.",
        "srv/service.cds history projections.",
        "srv/bug-service/history.js read and formatting helpers.",
        "Role-scoped Bug visibility and safe paging/order rules.",
        "Read-only request; no business transaction mutation.",
        "Read append-only history rows from SAP HANA Cloud/HDI; no provider side effect.",
        "Return chronological audit rows and refresh the History section.",
        "Read failure returns sanitized feedback and leaves Bug/history unchanged.",
        "IDTS-108 History screen evidence and reviewed IDTS-110/111 history cases.",
    )
    return sections


def _anchor(sheet, coordinate):
    cell = sheet[coordinate]
    if not isinstance(cell, MergedCell):
        return cell
    for merged in sheet.merged_cells.ranges:
        if coordinate in merged:
            return sheet.cell(merged.min_row, merged.min_col)
    raise ValueError(f"No merge anchor for {sheet.title}!{coordinate}")


def _write(sheet, coordinate, value, *, wrap=True, vertical="top"):
    cell = _anchor(sheet, coordinate)
    cell.value = value
    if wrap:
        old = cell.alignment
        cell.alignment = Alignment(
            horizontal=old.horizontal,
            vertical=vertical,
            text_rotation=old.text_rotation,
            wrap_text=True,
            shrink_to_fit=False,
            indent=old.indent,
        )


def _clear_values(sheet, start_row, end_row, start_col=2, end_col=59):
    for row in sheet.iter_rows(
        min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col
    ):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None


def _copy_style(source, target):
    target._style = copy(source._style)
    target.number_format = source.number_format
    target.protection = copy(source.protection)


def _merge_once(sheet, cell_range):
    if cell_range not in {str(item) for item in sheet.merged_cells.ranges}:
        sheet.merge_cells(cell_range)


def _unmerge_rows(sheet, start_row, end_row):
    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row >= start_row and merged.max_row <= end_row:
            sheet.unmerge_cells(str(merged))


def _set_metadata(sheet, title, language):
    labels = {
        "en": ("Created by:", "Created date:", "Modified by:", "Modified date:", "Reviewed by:", "Reviewed date:", "Pending"),
        "vi": ("Người tạo:", "Ngày tạo:", "Người cập nhật:", "Ngày cập nhật:", "Người review:", "Ngày review:", "Chờ duyệt"),
    }[language]
    def replace_template_value(row, start_col, end_col, value, fallback_col):
        candidates = []
        for column in range(start_col, end_col + 1):
            cell = sheet.cell(row, column)
            if isinstance(cell, MergedCell):
                continue
            if cell.value not in (None, ""):
                candidates.append(cell)
        target = candidates[0] if candidates else sheet.cell(row, fallback_col)
        for cell in candidates:
            cell.value = None
        _write(sheet, target.coordinate, value, wrap=False)

    _write(sheet, "B2", title, wrap=False)
    # Template tabs do not use one fixed coordinate for these values. Locate
    # the populated sample/formula cell in each official metadata block and
    # replace it in place so no WS/Credit-Memo sample survives.
    replace_template_value(3, 3, 15, "IDTS-TECH", 6)
    replace_template_value(3, 17, 42, "IDTS CAP/Fiori Technical Baseline" if language == "en" else "Nền tảng kỹ thuật IDTS CAP/Fiori", 21)
    _write(sheet, "AQ2", labels[0], wrap=False)
    replace_template_value(2, 44, 50, "DonHV", 46)
    _write(sheet, "AY2", labels[1], wrap=False)
    replace_template_value(2, 52, 59, DOCUMENT_DATE.isoformat(), 55)
    _write(sheet, "AQ3", labels[2], wrap=False)
    replace_template_value(3, 44, 50, "DonHV", 46)
    _write(sheet, "AY3", labels[3], wrap=False)
    replace_template_value(3, 52, 59, DOCUMENT_DATE.isoformat(), 55)
    _write(sheet, "AQ4", labels[4], wrap=False)
    replace_template_value(4, 44, 50, "Mentor / Supervisor" if language == "en" else "Mentor / người hướng dẫn", 46)
    _write(sheet, "AY4", labels[5], wrap=False)
    replace_template_value(4, 52, 59, labels[6], 55)


def _set_print(sheet, area, title_rows=None, *, fit_height=0):
    sheet.print_area = area
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = fit_height
    # Preserve the official template page shell while repeating the populated
    # table header on continuation pages.  This is a print-only aid; it does
    # not change sheet rows, styles, merges, tab structure or visible content.
    if title_rows:
        sheet.print_title_rows = title_rows


def _write_template_table(sheet, start_row, groups, headers, rows, style_sheet):
    """Create a record table only in template tabs that provide no record grid."""
    header_style = style_sheet["B5"]
    data_style = style_sheet["B6"]
    for offset, values in enumerate([headers, *rows]):
        row = start_row + offset
        is_header = offset == 0
        for (start, end), value in zip(groups, values):
            start_col = sheet[f"{start}1"].column
            end_col = sheet[f"{end}1"].column
            for col in range(start_col, end_col + 1):
                target = sheet.cell(row, col)
                _copy_style(header_style if is_header else data_style, target)
                target.alignment = Alignment(
                    horizontal="center" if is_header else "left",
                    vertical="center" if is_header else "top",
                    wrap_text=True,
                    shrink_to_fit=False,
                )
            _merge_once(sheet, f"{start}{row}:{end}{row}")
            sheet[f"{start}{row}"] = value
        sheet.row_dimensions[row].height = 32 if is_header else 60
    return start_row + len(rows)


def _fill_cover_and_history(workbook, language):
    cover = workbook["Cover"]
    title = "TECHNICAL SPECIFICATION" if language == "en" else "ĐẶC TẢ KỸ THUẬT"
    name = "IDTS CAP/Fiori Technical Baseline" if language == "en" else "Nền tảng kỹ thuật IDTS CAP/Fiori"
    for coordinate, value in {
        "B8": title, "N11": "IDTS", "Z11": "Issue and Defect Tracking System in SAP",
        "N12": "IDTS-TECH", "N13": name, "N14": DOCUMENT_DATE.isoformat(),
        "Z14": DOCUMENT_DATE.isoformat(), "AE19": "DonHV", "Z19": "Pending" if language == "en" else "Chờ duyệt",
        "U19": "Pending" if language == "en" else "Chờ duyệt",
    }.items():
        _write(cover, coordinate, value, wrap=False)
    _set_print(cover, "B1:BG24")

    history = workbook["Histories"]
    _clear_values(history, 3, 10, 2, 7)
    rows = {
        "en": [
            ("0.1", "Initial technical baseline", "Architecture, service and persistence scope", "2026-06-21"),
            ("0.2", "Shared QA integrations", "PostgreSQL, S3, Render and email outbox", "2026-07-02"),
            ("0.3", "Security and AI boundary", "Authentication, safe error and advisory AI", "2026-07-24"),
            ("0.4", "Official-template remediation", "All 12 sheets and exact action inventory", "2026-07-25"),
            ("0.5", "Runtime trace correction", "Exact source symbols, test truth and failure paths", "2026-07-25"),
            ("0.6", "Formal table remediation", "Structured catalogs and bilingual parity", "2026-07-25"),
            ("0.7", "Template-fidelity remediation", "Restore official inner layouts and exact per-action trace", "2026-07-26"),
            ("0.8", "BTP production-truth integration", "Integrate IDTS-107/108/109, HANA dictionary, current screens/messages and 14-part traces", DOCUMENT_DATE.isoformat()),
        ],
        "vi": [
            ("0.1", "Nền kỹ thuật ban đầu", "Kiến trúc, service và phạm vi lưu trữ", "2026-06-21"),
            ("0.2", "Tích hợp Shared QA", "PostgreSQL, S3, Render và email outbox", "2026-07-02"),
            ("0.3", "Ranh giới bảo mật và AI", "Xác thực, lỗi an toàn và AI tư vấn", "2026-07-24"),
            ("0.4", "Khắc phục theo template chính thức", "Đủ 12 sheet và danh mục action chính xác", "2026-07-25"),
            ("0.5", "Sửa truy vết runtime", "Source symbol, sự thật kiểm thử và failure path chính xác", "2026-07-25"),
            ("0.6", "Chuẩn hóa bảng formal", "Catalog có cấu trúc và parity song ngữ", "2026-07-25"),
            ("0.7", "Khôi phục độ trung thành với template", "Dùng đúng layout bên trong và truy vết riêng từng action", DOCUMENT_DATE.isoformat()),
        ],
    }[language]
    for row_number, (version, summary, affected, changed) in enumerate(rows, 3):
        for coordinate, value in {
            f"B{row_number}": row_number - 2, f"C{row_number}": version,
            f"D{row_number}": summary, f"E{row_number}": affected,
            f"F{row_number}": changed, f"G{row_number}": "DonHV",
        }.items():
            _write(history, coordinate, value, wrap=False)
        history.row_dimensions[row_number].height = 60
    history.row_dimensions[2].height = 30
    _set_print(history, "B1:G10", "1:2")


def _fill_intro_scope_assumptions(workbook, language):
    titles = {
        "en": ("Introduction", "Scope", "Assumptions"),
        "vi": ("Giới thiệu", "Phạm vi", "Giả định"),
    }[language]
    intro = workbook["Introduction"]
    _set_metadata(intro, titles[0], language)
    processing = {
        "H6": "Online", "P6": "Dialog / OData service", "V6": "IDTS SRS and approved Jira baseline",
        "H10": "Yes — English-only mentor submission",
    }
    if language == "vi":
        processing.update({"H6": "Trực tuyến", "P6": "Dialog / OData service", "V6": "IDTS SRS và Jira baseline đã duyệt", "H10": "Có — đặc tả tiếng Anh và tiếng Việt"})
    for coordinate, value in processing.items():
        _write(intro, coordinate, value, wrap=False)
    intro_text = {
        "en": "IDTS is implemented with SAP CAP Node.js and OData V4, consumed by SAP Fiori Elements/SAPUI5. CDS defines the service contract; Node.js handlers enforce authentication, authorization, validation, transactions, history, notifications, attachments and human-reviewed AI boundaries.",
        "vi": "IDTS được triển khai bằng SAP CAP Node.js và OData V4, được SAP Fiori Elements/SAPUI5 sử dụng. CDS định nghĩa service contract; handler Node.js kiểm soát xác thực, phân quyền, validation, transaction, lịch sử, thông báo, tệp đính kèm và ranh giới AI có con người review.",
    }[language]
    _write(intro, "C15", intro_text)
    intro.row_dimensions[15].height = 88
    _write(intro, "B16", "No additional supplement." if language == "en" else "Không có phụ lục bổ sung.")
    _merge_once(intro, "B16:BG16")
    intro.row_dimensions[16].height = 24
    _set_print(intro, "B1:BG16", "1:4")

    scope = workbook["Scope"]
    _set_metadata(scope, titles[1], language)
    _clear_values(scope, 6, 30)
    scope_rows = {
        "en": [
            "1. IN SCOPE — AuthService, BugService, draft/active writes and eleven exact lifecycle actions.",
            "1.1 IN SCOPE — Comments, standard CAP attachments, history, notifications, monitoring and role authorization.",
            "1.2 IN SCOPE — Human-reviewed AI suggestions, decisions, explicit apply/confirm and sanitized metrics.",
            "2. OUT OF SCOPE — ABAP/RAP, SAP Transport Requests and autonomous AI workflow decisions.",
            "3. LIMITATION — BTP is the current mentor/demo deployment; external S3, Brevo and AI providers remain governed integrations rather than native BTP services.",
            "3.1 EVIDENCE STATUS — Unit Test review: 38 accepted candidates, 2 held, 135 mapping-only and 13 blocked. UAT review: 22 MEETS, 12 DOES_NOT_MEET and 23 BLOCKED. These are review dispositions, not final acceptance claims.",
        ],
        "vi": [
            "SCP-01 — TRONG PHẠM VI — AuthService, BugService, ghi draft/active và mười một action vòng đời chính xác.",
            "SCP-02 — TRONG PHẠM VI — Bình luận, tệp đính kèm, lịch sử, thông báo, giám sát và phân quyền.",
            "SCP-03 — TRONG PHẠM VI — Suggestion AI có review, quyết định, apply/confirm rõ ràng và metrics đã làm sạch.",
            "SCP-04 — NGOÀI PHẠM VI — ABAP/RAP, SAP Transport Request và AI tự động quyết định workflow.",
            "SCP-05 — GIỚI HẠN — Shared QA phục vụ mentor/demo; OpenAI live vẫn TẮT / CHƯA NGHIỆM THU.",
        ],
    }[language]
    for row, value in enumerate(scope_rows, 6):
        _merge_once(scope, f"B{row}:BG{row}")
        _write(scope, f"B{row}", value)
        scope.row_dimensions[row].height = 34
    _write(scope, "B31", "No additional supplement." if language == "en" else "Không có phụ lục bổ sung.")
    _merge_once(scope, "B31:BG31")
    scope.row_dimensions[31].height = 24
    _set_print(scope, "B1:BG31", "1:5")

    assumptions = workbook["Assumptions"]
    _set_metadata(assumptions, titles[2], language)
    _clear_values(assumptions, 6, 41)
    assumption_rows = {
        "en": [
            "1. Local development uses SQLite; local records are not Shared QA evidence.",
            "1.1 Shared QA uses SAP HANA Cloud through the idts-sap01-db HDI container and must retain data after application restart/redeploy.",
            "1.2 Schema migrations are additive and idempotent; broad seed reload is forbidden.",
            "2. S3, Brevo and database credentials remain private and never appear in evidence.",
            "2.1 AI is advisory and feature-routed through the configured gateway; primary-provider, provider-fallback and deterministic-fallback evidence are reported separately.",
            "2.2 Shared QA evidence is valid only for the frozen Git SHA and matching SAP BTP deployment baseline.",
        ],
        "vi": [
            "ASM-01 — Phát triển local dùng SQLite; dữ liệu local không phải evidence Shared QA.",
            "ASM-02 — Shared QA dùng PostgreSQL trên Render và phải giữ dữ liệu sau restart/redeploy.",
            "ASM-03 — Migration schema phải cộng thêm và idempotent; cấm nạp lại seed diện rộng.",
            "ASM-04 — Credential S3, Brevo và database luôn riêng tư, không xuất hiện trong evidence.",
            "ASM-05 — OpenAI live đang tắt; PASS mock/fallback không phải nghiệm thu provider live.",
            "ASM-06 — Evidence Shared QA chỉ hợp lệ với Git SHA và Render deploy SHA đã cố định.",
        ],
    }[language]
    for row, value in enumerate(assumption_rows, 6):
        _merge_once(assumptions, f"B{row}:BG{row}")
        _write(assumptions, f"B{row}", value)
        assumptions.row_dimensions[row].height = 34
    _write(assumptions, "B42", "No additional supplement." if language == "en" else "Không có phụ lục bổ sung.")
    _merge_once(assumptions, "B42:BG42")
    assumptions.row_dimensions[42].height = 24
    _set_print(assumptions, "B1:BG42", "1:5")


def _requirement_rows(language):
    if language != "en":
        raise ValueError("Technical Specification v0.8 is an English-only SAP490 submission")
    _, source_rows = _markdown_table(
        IDTS109_DIR / "functional-requirements.md",
        "# Functional Requirements — Candidate",
    )
    requirement_ids = {
        "Authentication and session management": "SRS-FR-AUTH",
        "Profile and logout": "SRS-FR-AUTH",
        "Dashboard": "SRS-FR-MON",
        "PM monitoring": "SRS-FR-MON",
        "In-app notifications": "SRS-FR-NOTIFY",
        "Email outbox and delivery monitoring": "SRS-FR-NOTIFY",
        "Bug collaboration comments": "SRS-FR-COLLAB",
        "Bug evidence attachments": "SRS-FR-COLLAB",
        "AI assistance": "SRS-FR-AI",
        "AI review, apply and confirm": "SRS-FR-AI",
        "AI operational monitoring": "SRS-FR-AI",
    }
    rows = [tuple(row[:5] + [f"{row[5]} ({requirement_ids[row[5]]})"]) for row in source_rows]
    rows.extend([
        ("5", "The system shall create and update valid Bugs through governed draft and active writes.", "Tester / PM", "The actor is authorized and required catalog values are available.", "The committed Bug, history and notification state remain consistent.", "Bug create/update (SRS-FR-BUG)"),
        ("5.1", "The system shall assign an eligible active Developer whose responsibility matches the Bug classification.", "Tester / PM", "The Bug classification is valid and a matching active Developer is available.", "Assignee, status and next processor are updated only after backend validation.", "Developer assignment (SRS-FR-ASG)"),
        ("5.2", "The system shall enforce the eleven supported Bug lifecycle actions according to role, ownership and current status.", "Action-specific", "The requested action is permitted from the current Bug state and required inputs are present.", "The Bug, history and notification changes commit atomically or roll back together.", "Bug lifecycle (SRS-FR-LIFE)"),
    ])
    if language == "vi":
        translations = [
            "Xác thực người dùng và bảo vệ phiên bearer", "Tạo và cập nhật Bug hợp lệ qua ghi draft/active",
            "Phân công Developer đủ điều kiện", "Kiểm soát mười một action vòng đời chính xác",
            "Lưu bình luận và evidence tệp đính kèm", "Trả về workload và hàng đợi theo vai trò",
            "Lưu thông báo trong ứng dụng và trạng thái delivery outbox", "Cung cấp hỗ trợ AI có con người review",
        ]
        role_translations = {
            "Tester / Developer / PM": "Tester / Developer / PM",
            "Tester / PM": "Tester / PM",
            "Action-specific": "Tùy theo action",
            "Bug participants": "Người tham gia xử lý Bug",
            "System": "Hệ thống",
        }
        rows = [(r[0], translations[i], role_translations.get(r[2], r[2]), r[3], r[4], r[5]) for i, r in enumerate(rows)]
    return rows


def _fill_requirements(workbook, language):
    sheet = workbook["Functional Requirements"]
    _set_metadata(sheet, "Functional Requirements" if language == "en" else "Yêu cầu chức năng", language)
    _clear_values(sheet, 5, 70)
    headers = ["No.", "Business requirement", "Actor", "Precondition", "Outcome", "Related feature"]
    if language == "vi":
        headers = ["Mã yêu cầu", "Mục tiêu kỹ thuật", "Vai trò", "Service / thao tác", "Source", "Dữ liệu / evidence"]
    end = _write_template_table(
        sheet, 5,
        [("B", "F"), ("G", "Q"), ("R", "V"), ("W", "AB"), ("AC", "AP"), ("AQ", "BG")],
        headers, _requirement_rows(language), workbook["Message Definition"],
    )
    for row in range(6, end + 1):
        sheet.row_dimensions[row].height = 95
    _set_print(sheet, f"B1:BG{end}", "1:5")


def _fill_design(workbook, language):
    sheet = workbook["Technical Design"]
    _set_metadata(sheet, "Technical Design" if language == "en" else "Thiết kế kỹ thuật", language)
    descriptions = {
        "en": {
            5: "End-to-end CAP/Fiori workflow; architecture diagram is shown in the official graphic area. AuthSessions stores only the SHA-256 tokenHash; the raw bearer token is returned once.",
            7: "N/A — project planning is controlled in Jira; this specification documents the implemented technical baseline.",
            10: "Node.js package: idts-sap01; CAP modules are organized under app/, srv/ and db/.",
            12: "Production build truth: 48 deployable SAP HANA tables and 578 column declarations, including active domain tables, code lists, draft artifacts, DRAFT.DraftAdministrativeData, attachment artifacts, calculated-service helper artifacts and cds.outbox.Messages. The complete dictionary follows this template section.",
            16: "CDS types and code-list values define stable business codes; classic ABAP domains are not used.",
            32: "CDS elements define UUIDs, associations, compositions, timestamps, status codes and validation targets.",
            48: "CAP projections/read models expose Bugs, DeveloperWorkloads and role-aware monitoring data.",
            50: "N/A — JavaScript/CDS structures replace classic ABAP table types.",
            52: "db/schema.cds defines persistence; srv/service.cds defines service projections/actions/functions.",
            57: "app/bug-management-ui/annotations.cds and annotations/actions.cds provide Fiori metadata extensions.",
            60: "AuthService and BugService are declared in CDS and implemented by adjacent Node.js service handlers.",
            62: "OData V4 endpoints are /odata/v4/auth/ and /odata/v4/bug/. SAP BTP AppRouter protects the browser route through XSUAA and forwards authenticated requests to the CAP service.",
            64: "@UI, @Common, @Capabilities and action annotations drive List Report/Object Page behavior.",
            66: "N/A — classic ABAP Function Groups are not used; behavior is organized as CAP/Node.js modules.",
            68: "N/A — CAP event handlers and JavaScript helpers replace classic ABAP Function Modules.",
            74: "Login, profile, dashboard, Fiori Elements List Report/Object Page and review dialogs are documented in Screen Layout/Definition.",
            76: "JavaScript modules use explicit require/sap.ui.define dependencies; no ABAP include programs apply.",
            79: "Fiori controller extensions provide supported custom actions/sections without DOM or internal-control manipulation.",
            85: "BugService extends cds.ApplicationService; helper modules encapsulate permissions, transitions, history, content, monitoring, email and AI.",
            92: "Canonical safe messages are listed in Message Definition; no ABAP Message Class is used.",
            94: "Bug number generation is handled by backend logic; no SAP Number Range Object is configured.",
            96: "N/A — IDTS does not generate SAP Smart Forms.", 98: "N/A — no Smartform Style is used.",
            100: "Bug lifecycle states/actions are implemented as CAP bound actions with transaction-scoped history and notification side effects.",
            102: "Next processor, assignee, reason and action parameters form the workflow context; classic workflow container elements do not apply.",
            104: "Jira controls project tasks; runtime work queues are derived from Bug ownership/status and are not SAP Workflow tasks.",
            106: "SAP HANA Cloud/HDI stores business and attachment metadata. The configured S3 adapter stores attachment binary, Brevo sends email, Job Scheduler invokes the protected outbox processor, and Vercel AI Gateway routes advisory AI models.",
            108: "manifest.json routes Login/Dashboard/List Report/Object Page; actions return users to the affected Bug context.",
            115: "Official architecture diagram and sanitized Shared QA screenshots are embedded in the relevant template regions.",
            123: "Priority, Severity, Environment, SAP Module, Application Component and Defect Category are validated catalogs.",
            126: "Tester, Developer and PM roles are resolved server-side; UI visibility never replaces backend authorization.",
        },
        "vi": {
            5: "Luồng CAP/Fiori end-to-end; sơ đồ kiến trúc nằm trong vùng hình ảnh chính thức của template. AuthSessions chỉ lưu tokenHash SHA-256; raw bearer token chỉ được trả một lần.",
            7: "N/A — Jira quản lý kế hoạch dự án; đặc tả này mô tả baseline kỹ thuật đã triển khai.",
            10: "Node.js package: idts-sap01; module CAP được tổ chức trong app/, srv/ và db/.",
            12: "Entity CDS: Users, Bugs, Comments, Attachments, HistoryEvents, HistoryLogs, Notifications, NotificationDeliveries, AiSuggestions, DuplicateLinks và các code list.",
            16: "Type CDS và giá trị code list định nghĩa mã nghiệp vụ ổn định; không dùng ABAP domain cổ điển.",
            32: "Element CDS định nghĩa UUID, association, composition, timestamp, status code và validation target.",
            48: "Projection/read model CAP expose Bugs, DeveloperWorkloads và dữ liệu giám sát theo vai trò.",
            50: "N/A — cấu trúc JavaScript/CDS thay thế ABAP table type cổ điển.",
            52: "db/schema.cds định nghĩa persistence; srv/service.cds định nghĩa projection/action/function của service.",
            57: "app/bug-management-ui/annotations.cds và annotations/actions.cds cung cấp metadata extension cho Fiori.",
            60: "AuthService và BugService được khai báo trong CDS và triển khai bởi handler Node.js service nằm cạnh service definition.",
            62: "Endpoint OData V4: /odata/v4/auth/ và /odata/v4/bug/. Shared QA trên Render expose cùng contract.",
            64: "Annotation @UI, @Common, @Capabilities và action điều khiển List Report/Object Page.",
            66: "N/A — không dùng ABAP Function Group; behavior được tổ chức bằng module CAP/Node.js.",
            68: "N/A — CAP event handler và helper JavaScript thay thế ABAP Function Module.",
            74: "Login, profile, dashboard, Fiori Elements List Report/Object Page và review dialog được mô tả trong Screen Layout/Definition.",
            76: "Module JavaScript dùng dependency require/sap.ui.define rõ ràng; không áp dụng ABAP include program.",
            79: "Fiori controller extension cung cấp action/section tùy biến được hỗ trợ, không thao tác DOM hoặc internal control.",
            85: "BugService extends cds.ApplicationService; helper module tách permissions, transitions, history, content, monitoring, email và AI.",
            92: "Thông báo an toàn chuẩn nằm trong Message Definition; không dùng ABAP Message Class.",
            94: "Backend sinh Bug number; không cấu hình SAP Number Range Object.",
            96: "N/A — IDTS không tạo SAP Smart Form.", 98: "N/A — không dùng Smartform Style.",
            100: "Trạng thái/action vòng đời Bug được triển khai bằng CAP bound action với history và notification trong transaction.",
            102: "Next processor, assignee, reason và tham số action tạo workflow context; không áp dụng workflow container cổ điển.",
            104: "Jira quản lý task dự án; runtime queue được suy ra từ ownership/status của Bug, không phải SAP Workflow task.",
            106: "AWS S3 lưu binary tệp; Brevo gửi email; PostgreSQL trên Render lưu Shared QA; OpenAI live vẫn tắt.",
            108: "manifest.json định tuyến Login/Dashboard/List Report/Object Page; action đưa người dùng về Bug liên quan.",
            115: "Sơ đồ kiến trúc chính thức và screenshot Shared QA đã làm sạch được chèn vào vùng template phù hợp.",
            123: "Priority, Severity, Environment, SAP Module, Application Component và Defect Category là catalog được validate.",
            126: "Backend resolve vai trò Tester, Developer và PM; visibility UI không thay thế authorization backend.",
        },
    }[language]
    heading_rows = sorted(descriptions)
    for heading_row in heading_rows:
        next_heading = next((item for item in heading_rows if item > heading_row), 130)
        body_row = heading_row + 1
        if body_row >= next_heading:
            continue
        _merge_once(sheet, f"B{body_row}:AP{body_row}")
        _write(sheet, f"B{body_row}", descriptions[heading_row])
        sheet.row_dimensions[body_row].height = 48
    sheet._images = []
    diagram = XLImage(ROOT / "docs" / "diagrams" / "rendered" / "png" / "02-cap-fiori-architecture.png")
    diagram.width, diagram.height = 480, 340
    sheet.add_image(diagram, "AQ5")
    dictionary = _database_dictionary_rows()
    start = 132
    _write(sheet, f"B{start}", "1. Production HANA Data Dictionary — 48 tables / 578 columns")
    _merge_once(sheet, f"B{start}:BG{start}")
    sheet.row_dimensions[start].height = 28
    headers = ["Physical table", "Column", "Data type", "Key / null / default", "Relationship", "Business purpose", "CDS source", "Evidence"]
    rows = [
        (
            item["Physical HANA Table"], item["Column"], item["Data Type"],
            f"PK={item['Primary Key']}; nullable={item['Nullable']}; default={item['Default'] or '-'}",
            item["Relationship / Target"] or "-", item["Business Purpose"],
            item["CDS / Model Source"], item["Database Evidence"],
        )
        for item in dictionary
    ]
    end = _write_template_table(
        sheet, start + 1,
        [("B", "H"), ("I", "M"), ("N", "R"), ("S", "X"), ("Y", "AD"), ("AE", "AK"), ("AL", "AR"), ("AS", "BG")],
        headers, rows, workbook["Message Definition"],
    )
    for row in range(start + 2, end + 1):
        # Database dictionary cells are merged across template column groups, so
        # Excel/LibreOffice cannot auto-fit them. 120 pt is the verified minimum
        # that keeps the longest purpose/source/evidence text visible at 12 pt.
        sheet.row_dimensions[row].height = 120
    _set_print(sheet, f"B1:BG{end}", "1:4")


def _fill_standards(workbook, language):
    sheet = workbook["Development Standards"]
    _set_metadata(sheet, "Development Standards" if language == "en" else "Tiêu chuẩn phát triển", language)
    details = {
        "en": {
            5: "CAP entities/services use stable business names; JavaScript symbols use camelCase; constants use explicit codes.",
            35: "CDS defines UUID keys, associations/compositions, constraints and audit fields; the production CAP build emits HDI artifacts for SAP HANA Cloud.",
            38: "N/A — CAP/Node.js modules replace classic ABAP Function Groups.", 40: "N/A — CAP handlers/helpers replace classic ABAP Function Modules.",
            43: "Non-obvious entry points explain trigger, input, decision, side effect, next dependency and breakpoint.",
            47: "Follow repository lint/format conventions; do not reformat unrelated files.",
            50: "Focused programmatic, API, integration and browser tests require positive, negative, role and persistence evidence.",
            53: "Tester, Developer and PM roles are resolved server-side from the authenticated IDTS user.",
            55: "N/A — IDTS uses CAP read models and Fiori Elements, not ABAP RAP My Inbox.",
            61: "Use CAP query API and request transactions; validate before mutation and keep workflow/history/notification consistent.",
            63: "UI text must be human-facing and sanitized; never expose SQL, stack traces, credentials or developer-only copy.",
            66: "Configuration is profile/environment based; secrets remain private and never enter source or evidence.",
            69: "Use Fiori/UI5 controls and official diagram assets; do not copy mock HTML/CSS into runtime.",
            71: "Validate active Priority, Severity, Environment and classification catalogs in CAP.",
        },
        "vi": {
            5: "Entity/service CAP dùng tên nghiệp vụ ổn định; symbol JavaScript dùng camelCase; constant dùng code rõ ràng.",
            35: "CDS định nghĩa UUID, association/composition, constraint và audit field; PostgreSQL theo output deploy của CAP.",
            38: "N/A — module CAP/Node.js thay thế ABAP Function Group cổ điển.", 40: "N/A — CAP handler/helper thay thế ABAP Function Module cổ điển.",
            43: "Entry point không hiển nhiên phải giải thích trigger, input, quyết định, side effect, dependency tiếp theo và breakpoint.",
            47: "Theo lint/format của repository; không reformat file không liên quan.",
            50: "Test programmatic, API, integration và browser cần evidence positive, negative, role và persistence.",
            53: "Backend resolve vai trò Tester, Developer và PM từ người dùng IDTS đã xác thực.",
            55: "N/A — IDTS dùng CAP read model và Fiori Elements, không dùng ABAP RAP My Inbox.",
            61: "Dùng CAP query API và request transaction; validate trước mutation, giữ nhất quán workflow/history/notification.",
            63: "Text UI phải dễ hiểu và được làm sạch; không lộ SQL, stack trace, credential hoặc copy nội bộ.",
            66: "Cấu hình theo profile/environment; secret luôn riêng tư và không nằm trong source/evidence.",
            69: "Dùng Fiori/UI5 control và diagram chính thức; không copy HTML/CSS mock vào runtime.",
            71: "CAP validate catalog Priority, Severity, Environment và classification đang active.",
        },
    }[language]
    for heading_row, text in details.items():
        body_row = heading_row + 1
        _write(sheet, f"C{body_row}", text)
        sheet.row_dimensions[body_row].height = max(sheet.row_dimensions[body_row].height or 15, 42)
    _, standard_rows = _markdown_table(IDTS109_DIR / "development-standards.md", "# Development Standards — Candidate")
    start = 84
    _write(sheet, f"B{start}", "1. Detailed Development Standards")
    end = _write_template_table(
        sheet, start + 1,
        [("B", "F"), ("G", "L"), ("M", "AA"), ("AB", "AJ"), ("AK", "AQ"), ("AR", "BG")],
        ["Standard", "Area", "Rule", "Application", "Verification", "Evidence"],
        [tuple(row[:6]) for row in standard_rows], workbook["Message Definition"],
    )
    sheet.row_dimensions[start].height = 20
    for row in range(start + 2, end + 1):
        sheet.row_dimensions[row].height = 195
    _set_print(sheet, f"B1:BG{end}", "1:4")


def _fill_screen_layout(workbook, language):
    sheet = workbook["Screen Layout"]
    _set_metadata(sheet, "Screen Layout" if language == "en" else "Bố cục màn hình", language)
    _write(sheet, "B5", "Representative implemented screens" if language == "en" else "Các màn hình đã triển khai tiêu biểu", vertical="center")
    _write(sheet, "B6", "Layout 1 — Tester Bug Object Page" if language == "en" else "Bố cục 1 — Trang chi tiết Bug của Tester", vertical="center")
    _write(sheet, "B7", "Fiori Elements Object Page: summary, classification, assignment, collaboration, history and lifecycle actions." if language == "en" else "Fiori Elements Object Page: tóm tắt, phân loại, phân công, cộng tác, lịch sử và action vòng đời.")
    _merge_once(sheet, "B16:BG16")
    _write(sheet, "B16", "Layout 2 — Developer actions are shown below the Tester layout." if language == "en" else "Bố cục 2 — Action của Developer được minh họa bên dưới bố cục Tester.")
    sheet.row_dimensions[16].height = 24
    sheet._images = []
    tester = XLImage(ROOT / "docs" / "pm" / "evidence" / "idts-108" / "screenshots" / "48-pm-list-report.png")
    developer = XLImage(ROOT / "docs" / "pm" / "evidence" / "idts-108" / "screenshots" / "51-pm-bug-0024-object-page.png")
    tester.width, tester.height = 550, 310
    developer.width, developer.height = 550, 310
    sheet.add_image(tester, "B9")
    sheet.add_image(developer, "AI9")
    _, layout_rows = _markdown_table(IDTS108_README, "## 1. Screen Layout inventory")
    start = 18
    _merge_once(sheet, f"B{start}:BG{start}")
    _write(sheet, f"B{start}", "1. Screen Layout Inventory")
    sheet.row_dimensions[start].height = 24
    end = _write_template_table(
        sheet, start + 1,
        [("B", "D"), ("E", "J"), ("K", "Q"), ("R", "AB"), ("AC", "AG"), ("AH", "AM"), ("AN", "AT"), ("AU", "BG")],
        ["No.", "Screen / dialog", "Route / trigger", "Page type / areas", "Roles", "Navigation / result", "Source trace", "Evidence state"],
        [tuple(row[:8]) for row in layout_rows], workbook["Message Definition"],
    )
    for row in range(start + 2, end + 1):
        sheet.row_dimensions[row].height = 120
    _set_print(sheet, f"B1:BG{end}", "1:4")


def _screen_rows(language):
    rows = [
        ("Title", "Text", "I", "1", "String", 200, "Yes", "Bug title", "Required; safe inline message"),
        ("Priority", "VH", "I", "1", "Assoc", 36, "Yes", "Active values", "Backend validates active catalog"),
        ("Severity", "VH", "I", "1", "Assoc", 36, "Yes", "Active values", "Backend validates active catalog"),
        ("Environment", "VH", "I", "1", "Assoc", 36, "Yes", "Active values", "Backend validates active catalog"),
        ("Assignee", "VH", "I", "1", "Assoc", 36, "No", "Smart assign", "Tester/PM only; empty means Pending Assignment"),
        ("Bug Number", "Text", "O", "1", "String", 20, "No", "Backend generated", "Read-only identifier"),
        ("Status", "Status", "O", "1", "Assoc", 36, "No", "Derived", "Changed only by validated lifecycle actions"),
        ("Current Action Owner", "Text", "O", "1", "Assoc", 36, "No", "Derived", "Different from technical assignee when workflow changes"),
        ("Comment", "TextArea", "I", "N", "String", 2000, "No", "After save", "Authorized Bug participant"),
        ("Attachment", "Upload", "I", "N", "Binary", "10 MB", "No", "Upload", "Pending in client memory before Bug activation; metadata DB, binary S3"),
        ("History", "Timeline", "O", "N", "Comp", "-", "No", "Show More", "Read-only audit timeline"),
        ("Lifecycle action", "Button", "I", "1", "Action", "-", "Role", "Object Page", "Backend permission and transition validation"),
        ("AI review", "Dialog", "I", "1", "Action", "-", "Role", "Review", "Review-only unless explicit apply/confirm"),
        ("Notification", "List", "O", "N", "Comp", "-", "No", "User scope", "In-app state independent from email delivery"),
        ("Dashboard KPI", "Card", "O", "N", "ReadModel", "-", "No", "Role filter", "No write authority"),
        ("Email delivery", "Status", "O", "N", "Comp", "-", "No", "Worker", "PENDING/SENT/FAILED/SKIPPED"),
    ]
    if language == "vi":
        translations = {
            "Title": "Tiêu đề", "Priority": "Mức ưu tiên", "Severity": "Mức nghiêm trọng", "Environment": "Môi trường",
            "Assignee": "Developer được phân công", "Bug Number": "Mã Bug", "Status": "Trạng thái", "Current Action Owner": "Người xử lý hiện tại",
            "Comment": "Bình luận", "Attachment": "Tệp đính kèm", "History": "Lịch sử", "Lifecycle action": "Action vòng đời",
            "AI review": "Review AI", "Notification": "Thông báo", "Dashboard KPI": "KPI dashboard", "Email delivery": "Trạng thái gửi email",
        }
        defaults = {
            "Bug title": "Tiêu đề Bug", "Active value help": "Value help đang active",
            "Active values": "Giá trị active", "Smart assign": "Phân công thông minh",
            "Smart assignment help": "Value help phân công thông minh", "Backend generated": "Backend tự sinh",
            "Backend derived": "Backend suy ra", "After active save": "Sau khi lưu Bug active",
            "Derived": "Backend suy ra", "After save": "Sau khi lưu", "Upload": "Tải lên",
            "Upload control": "Control tải lên", "Show More": "Xem thêm", "Object Page action": "Action trên Object Page",
            "Object Page": "Trang chi tiết", "Review": "Xem xét", "Role filter": "Lọc theo vai trò", "Worker": "Tiến trình nền",
            "Review dialog": "Dialog review", "User scope": "Theo người dùng", "Role-filtered": "Lọc theo vai trò",
            "System worker": "Worker hệ thống",
        }
        remarks = {
            "Required; safe inline message": "Bắt buộc; thông báo inline an toàn",
            "Backend validates active catalog": "Backend kiểm tra catalog đang active",
            "Tester/PM only; empty means Pending Assignment": "Chỉ Tester/PM; để trống thì thành Pending Assignment",
            "Read-only identifier": "Mã định danh chỉ đọc", "Changed only by validated lifecycle actions": "Chỉ đổi qua action vòng đời đã validate",
            "Different from technical assignee when workflow changes": "Có thể khác assignee kỹ thuật khi workflow thay đổi",
            "Authorized Bug participant": "Người tham gia Bug đã được cấp quyền",
            "Pending in client memory before Bug activation; metadata DB, binary S3": "Tạm giữ trong bộ nhớ trình duyệt trước khi activate Bug; metadata ở DB, binary ở S3",
            "Read-only audit timeline": "Timeline audit chỉ đọc", "Backend permission and transition validation": "Backend kiểm tra quyền và transition",
            "Review-only unless explicit apply/confirm": "Chỉ review trừ khi apply/confirm rõ ràng",
            "In-app state independent from email delivery": "Trạng thái trong ứng dụng độc lập với việc gửi email",
            "No write authority": "Không cấp quyền ghi", "PENDING/SENT/FAILED/SKIPPED": "PENDING/SENT/FAILED/SKIPPED",
        }
        rows = [
            (translations[a], b, c, d, e, f, "Có" if g == "Yes" else ("Không" if g == "No" else "Theo vai trò"), defaults.get(h, h), remarks.get(i, i))
            for a, b, c, d, e, f, g, h, i in rows
        ]
    return rows


def _fill_screen_definition(workbook, language):
    sheet = workbook["Screen Definition"]
    _set_metadata(sheet, "Screen Definition" if language == "en" else "Định nghĩa màn hình", language)
    _write(sheet, "B5", "Bug List Report and Object Page" if language == "en" else "Danh sách Bug và trang chi tiết Bug", wrap=False)
    _merge_once(sheet, "B9:BG9")
    _write(sheet, "B9", "1. Bug List Report / Object Page" if language == "en" else "1. Danh sách Bug / trang chi tiết Bug")
    sheet.row_dimensions[9].height = 24
    if language == "vi":
        for coordinate, value in {"B10": "STT", "R10": "Thành phần màn hình", "AW10": "Ghi chú", "C11": "Tên", "I11": "Loại", "M11": "I/O", "O11": "Một/Nhiều", "R11": "Kiểu dữ liệu", "U11": "Độ dài", "X11": "Thập phân", "AA11": "Bắt buộc", "AE11": "Giá trị mặc định", "AI11": "Định dạng"}.items():
            _write(sheet, coordinate, value, vertical="center")
    _, field_rows = _markdown_table(IDTS108_README, "### 2.1 Shared Bug fields")
    _, action_rows = _markdown_table(IDTS108_README, "### 2.2 Actions and collaboration controls")
    _unmerge_rows(sheet, 12, 140)
    _clear_values(sheet, 12, 140)
    groups = [("B", "D"), ("E", "J"), ("K", "Q"), ("R", "X"), ("Y", "AB"), ("AC", "AH"), ("AI", "AN"), ("AO", "AT"), ("AU", "BG")]
    headers = ["No.", "Screen", "Field / action", "Binding / operation", "I/O", "Type / input", "Required / role", "Visibility", "Validation / failure behavior"]
    normalized_fields = [tuple(row[:9]) for row in field_rows]
    normalized_actions = [
        (row[0], row[1], row[2], row[3], "Action", row[4], row[5], "Role-controlled", row[6])
        for row in action_rows
    ]
    end = _write_template_table(sheet, 12, groups, headers, normalized_fields + normalized_actions, workbook["Message Definition"])
    _set_print(sheet, f"B1:BG{end}", "1:11")


def _fill_messages(workbook, language):
    sheet = workbook["Message Definition"]
    _set_metadata(sheet, "Message Definition" if language == "en" else "Định nghĩa thông báo", language)
    if language != "en":
        raise ValueError("Technical Specification v0.8 is an English-only SAP490 submission")
    _, message_rows = _markdown_table(IDTS109_DIR / "message-catalog.md", "## Catalog")
    _unmerge_rows(sheet, 5, 300)
    _clear_values(sheet, 5, 300)
    groups = [("B", "F"), ("G", "P"), ("Q", "Y"), ("Z", "AC"), ("AD", "AG"), ("AH", "AK"), ("AL", "AO"), ("AP", "AT"), ("AU", "AZ"), ("BA", "BG")]
    headers = ["Message ID", "User-facing text / safe summary", "Exact trigger / source", "HTTP / status", "Target", "Role / context", "Rollback behavior", "Sanitized logging", "Frontend handling", "Evidence"]
    end = _write_template_table(sheet, 5, groups, headers, [tuple(row[:10]) for row in message_rows], sheet)
    for row in range(7, end + 1):
        sheet.row_dimensions[row].height = 135
    _set_print(sheet, f"B1:BG{end}", "1:5")


def _flow_rows(language):
    rows = [
        ("FLOW-AUTH", "Submit sign-in", "POST /odata/v4/auth/login", "srv/auth.js::login", "Read Users; insert AuthSessions; return raw token once", "EVID-AUTH"),
        ("FLOW-DRAFT-NEW", "Choose Create", "NEW draft", "srv/service.js::prepareDraftNew", "Authorize create; initialize draft only", "EVID-BUG-WRITE"),
        ("FLOW-DRAFT-PATCH", "Edit draft field", "PATCH Bugs.drafts", "srv/bug-service/drafts.js::prepareDraftPatch", "Validate projected draft state; no active mutation", "EVID-BUG-WRITE"),
        ("FLOW-DRAFT-SAVE", "Save draft", "SAVE draft", "srv/bug-service/drafts.js::handleDraftSave", "Activate draft and route active CREATE/UPDATE", "EVID-BUG-WRITE"),
        ("FLOW-ACTIVE-CREATE", "Activate new Bug", "CREATE Bugs", "srv/bug-service/bug-write.js::prepareBugWrite", "Insert Bug; history; notification in request transaction", "EVID-BUG-WRITE"),
        ("FLOW-ACTIVE-UPDATE", "Save edited Bug", "UPDATE Bugs", "srv/bug-service/bug-write.js::prepareBugWrite", "Update Bug after permission/transition validation", "EVID-BUG-WRITE"),
        ("FLOW-ASSIGN", "Confirm Developer", "assignToDeveloper", "srv/bug-service/actions.js::assignToDeveloper", "Update assignee/status/next processor/history/notification", "EVID-ASSIGN"),
    ]
    for action, actors, before, after in LIFECYCLE_ACTIONS:
        rows.append((f"FLOW-{action.upper()}", f"{actors} invokes action", action, f"srv/bug-service/actions.js::{action}", f"Validate {before} → {after}; commit Bug/history/notification atomically", "EVID-IDTS-89"))
    rows.extend([
        ("FLOW-COMMENT-CREATE", "Add Comment", "CREATE Comments", "srv/bug-service/content.js::prepareCommentCreate", "Authorize participant; persist comment metadata", "EVID-COMMENT"),
        ("FLOW-ATTACH-DOWNLOAD", "Choose Download", "GET Attachments/content", "BugCollaboration.js::onDownloadAttachment", "Read authorized S3 binary; return safe filename/content", "EVID-ATTACH"),
        ("FLOW-ATTACH-DELETE", "Confirm Delete", "DELETE Attachments", "BugCollaboration.js::onDeleteAttachment", "Delete authorized metadata/S3 object; preserve Bug", "EVID-ATTACH"),
        ("FLOW-HISTORY-READ", "Open History / Show More", "GET HistoryEvents/HistoryLogs", "srv/bug-service/history.js", "Read paged immutable audit records; no mutation", "EVID-HISTORY"),
        ("FLOW-NOTIFY-OUTBOX", "Workflow commits notification", "Internal CAP transaction", "srv/email/outbox.js::writeNotificationRecord", "Insert Notifications and NotificationDeliveries", "EVID-EMAIL"),
        ("FLOW-EMAIL-WORKER", "Worker polls eligible rows", "Background worker", "srv/email/worker.js::startEmailWorker", "Claim with lock; send; update retry/SENT/FAILED/SKIPPED", "EVID-EMAIL"),
        ("FLOW-MONITOR", "Open dashboard/queue", "GET read model", "srv/bug-service/monitoring.js::readDeveloperWorkloads", "Return role-scoped KPI rows; no mutation", "EVID-MONITOR"),
    ])
    for function_id, action, source, objective_en, objective_vi in AI_FUNCTIONS:
        rows.append((f"FLOW-{function_id}", f"Invoke {action}", action, source, objective_en if language == "en" else objective_vi, "EVID-AI-FALLBACK"))
    if language == "vi":
        vi_base = {
            "FLOW-AUTH": ("Gửi biểu mẫu đăng nhập", "Đọc Users; thêm AuthSessions; chỉ trả raw token một lần"),
            "FLOW-DRAFT-NEW": ("Chọn Tạo", "Kiểm tra quyền tạo; chỉ khởi tạo draft"),
            "FLOW-DRAFT-PATCH": ("Sửa field trên draft", "Validate trạng thái draft dự kiến; chưa thay đổi Bug active"),
            "FLOW-DRAFT-SAVE": ("Lưu draft", "Activate draft rồi chuyển sang CREATE/UPDATE active"),
            "FLOW-ACTIVE-CREATE": ("Activate Bug mới", "Thêm Bug, history và notification trong request transaction"),
            "FLOW-ACTIVE-UPDATE": ("Lưu Bug đã sửa", "Cập nhật Bug sau khi kiểm tra permission/transition"),
            "FLOW-ASSIGN": ("Xác nhận Developer", "Cập nhật assignee/status/next processor/history/notification"),
            "FLOW-COMMENT-CREATE": ("Thêm bình luận", "Cấp quyền người tham gia; lưu metadata bình luận"),
            "FLOW-ATTACH-DOWNLOAD": ("Chọn tải xuống", "Đọc binary S3 đã cấp quyền; trả tên/nội dung an toàn"),
            "FLOW-ATTACH-DELETE": ("Xác nhận xóa", "Xóa metadata/S3 object đã cấp quyền; giữ nguyên Bug"),
            "FLOW-HISTORY-READ": ("Mở Lịch sử / Xem thêm", "Đọc audit record bất biến có phân trang; không mutation"),
            "FLOW-NOTIFY-OUTBOX": ("Workflow commit thông báo", "Thêm Notifications và NotificationDeliveries"),
            "FLOW-EMAIL-WORKER": ("Worker polling các dòng đủ điều kiện", "Claim bằng lock; gửi; cập nhật retry/SENT/FAILED/SKIPPED"),
            "FLOW-MONITOR": ("Mở dashboard/hàng đợi", "Trả KPI theo vai trò; không mutation"),
        }
        lifecycle_map = {f"FLOW-{action.upper()}": (actors, before, after) for action, actors, before, after in LIFECYCLE_ACTIONS}
        translated = []
        for flow, trigger, request, source, effect, evidence in rows:
            if flow in vi_base:
                trigger, effect = vi_base[flow]
            elif flow in lifecycle_map:
                actors, before, after = lifecycle_map[flow]
                trigger = f"{actors} gọi action"
                effect = f"Kiểm tra {before} → {after}; commit Bug/history/notification atomically"
            elif flow.startswith("FLOW-FN-AI-"):
                trigger = "Gọi action hỗ trợ AI"
            translated.append((flow, trigger, request, source, effect, evidence))
        rows = translated
    return rows


def _fill_implementation(workbook, language):
    sheet = workbook["Technical Implementation"]
    _write(sheet, "B3", "Technical Implementation" if language == "en" else "Triển khai kỹ thuật", wrap=False)
    _write(sheet, "B4", "IDTS-TECH", wrap=False)
    _write(sheet, "L4", "IDTS CAP/Fiori Technical Baseline" if language == "en" else "Nền tảng kỹ thuật IDTS CAP/Fiori", wrap=False)
    _write(sheet, "X3", "Created by:" if language == "en" else "Người tạo:", wrap=False)
    _write(sheet, "AC3", "DonHV", wrap=False)
    _write(sheet, "X4", "Modified date:" if language == "en" else "Ngày cập nhật:", wrap=False)
    _write(sheet, "AC4", DOCUMENT_DATE.isoformat(), wrap=False)
    _write(sheet, "X5", "Reviewed by:" if language == "en" else "Người review:", wrap=False)
    _write(sheet, "AC5", "Pending" if language == "en" else "Chờ duyệt", wrap=False)
    if language != "en":
        raise ValueError("Technical Specification v0.8 is an English-only SAP490 submission")
    _unmerge_rows(sheet, 7, 500)
    _clear_values(sheet, 7, 500, 2, 33)
    rows = []
    source_sections = _technical_implementation_sections()
    # Preserve only current trace IDs; retired browser queue/upload helpers must not
    # reappear in the mentor-facing Technical Specification.
    for section in source_sections:
        if section["trace"].startswith("TI-COLLAB-01"):
            section["trace"] += "; FLOW-COMMENT-CREATE"
    for section in source_sections + _supplemental_implementation_sections():
        rows.append((section["title"], "Technical trace ID", section["trace"]))
        for number, label, value in section["items"]:
            rows.append(("", f"{number}. {label}", value))
    end = _write_template_table(
        sheet, 7,
        [("B", "F"), ("G", "K"), ("L", "AG")],
        ["Function / action", "Implementation item", "Current technical implementation and evidence"],
        rows, workbook["Message Definition"],
    )
    for row in range(8, end + 1):
        if sheet[f"B{row}"].value:
            sheet.row_dimensions[row].height = 28
        else:
            sheet.row_dimensions[row].height = 42
    _set_print(sheet, f"B1:AG{end}", "1:7")


def _localize_template_headings(workbook):
    translations = {
        "Technical Specification": "Đặc tả kỹ thuật", "TECHNICAL SPECIFICATION": "ĐẶC TẢ KỸ THUẬT",
        "Introduction": "Giới thiệu", "Scope": "Phạm vi", "Assumptions": "Giả định",
        "Functional Requirements": "Yêu cầu chức năng", "Technical Design": "Thiết kế kỹ thuật",
        "Development Standards": "Tiêu chuẩn phát triển", "Screen Layout": "Bố cục màn hình",
        "Screen Definition": "Định nghĩa màn hình", "Message Definition": "Định nghĩa thông báo",
        "Technical Implementation": "Triển khai kỹ thuật", "Function ID": "Mã chức năng", "Function Name": "Tên chức năng",
        "Created by:": "Người tạo:", "Created date:": "Ngày tạo:", "Modified by:": "Người cập nhật:",
        "Modified date:": "Ngày cập nhật:", "Reviewed by:": "Người review:", "Reviewed date:": "Ngày review:",
        "Business Process": "Quy trình nghiệp vụ", "WBS & Timeline": "WBS và tiến độ",
        "Data Dictionary Objects ": "Đối tượng từ điển dữ liệu", "Data Dictionary Objects": "Đối tượng từ điển dữ liệu", "Package": "Package", "Tables": "Bảng dữ liệu",
        "Domain": "Miền dữ liệu", "Data Element": "Phần tử dữ liệu", "View": "View", "Table Type": "Kiểu bảng",
        "Data Definition": "Định nghĩa dữ liệu", "CDS Metadata Extension": "CDS Metadata Extension",
        "Service Definition": "Định nghĩa service", "Service Binding": "Service binding", "Annotation Model": "Mô hình annotation",
        "Function Group": "Function Group (không áp dụng)", "Function Module": "Function Module (không áp dụng)",
        "Screens": "Màn hình", "Includes": "Module phụ thuộc", "Enhancement Implementation": "Triển khai extension",
        "Class": "Lớp/service", "Message Class": "Danh mục thông báo", "Number Range Object": "Đối tượng sinh số",
        "Smart Forms": "Smart Forms (không áp dụng)", "Smartform Style": "Smartform Style (không áp dụng)",
        "Workflow Scenario": "Kịch bản workflow", "Container Element": "Dữ liệu workflow", "Task": "Task",
        "Provider": "Provider", "Navigation Target Object": "Đích điều hướng", "Graphic": "Đồ họa",
        "Catalogs": "Catalog", "Roles": "Vai trò", "Naming Convention": "Quy ước đặt tên", "Table Design": "Thiết kế bảng",
        "Code Comment": "Comment trong code", "Pretty Printer": "Định dạng code", "Test Program": "Chương trình kiểm thử",
        "Custom Role": "Vai trò tùy chỉnh", "RAP Model for My Inbox display": "RAP My Inbox (không áp dụng)",
        "DB Query": "Truy vấn database", "Texts in Code": "Text trong code", "System Program & Environment": "Chương trình và môi trường",
        "Graphics": "Đồ họa", "Pending": "Chờ duyệt", "Supplement": "Phụ lục",
    }
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if not isinstance(cell, MergedCell) and isinstance(cell.value, str):
                    stripped = cell.value.strip()
                    if stripped in translations:
                        cell.value = cell.value[: len(cell.value) - len(cell.value.lstrip())] + translations[stripped]


def generate_technical_specification(language):
    if language != "en":
        raise ValueError("Technical Specification v0.8 is an English-only SAP490 submission")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output = OUTPUT_DIR / f"Technical_Specification_IDTS_SAP01_{language}_v{VERSION}.xlsx"
    shutil.copy2(TEMPLATE, output)
    workbook = load_workbook(output)
    for name, item in list(workbook.defined_names.items()):
        if "#REF!" in str(item.attr_text):
            del workbook.defined_names[name]
    for sheet in workbook.worksheets:
        for name, item in list(sheet.defined_names.items()):
            if "#REF!" in str(item.attr_text):
                del sheet.defined_names[name]

    _fill_cover_and_history(workbook, language)
    _fill_intro_scope_assumptions(workbook, language)
    _fill_requirements(workbook, language)
    _fill_design(workbook, language)
    _fill_standards(workbook, language)
    _fill_screen_layout(workbook, language)
    _fill_screen_definition(workbook, language)
    _fill_messages(workbook, language)
    _fill_implementation(workbook, language)
    if language == "vi":
        _localize_template_headings(workbook)
    workbook.properties.title = f"IDTS SAP490 Technical Specification {language.upper()} v{VERSION}"
    workbook.properties.subject = "Official-template CAP/Fiori technical baseline"
    workbook.properties.creator = "IDTS SAP01 Team"
    workbook.save(output)
    return output
