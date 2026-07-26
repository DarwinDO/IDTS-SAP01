"""Generate the EN/VI Functional Specification from the official template."""

from __future__ import annotations

from copy import copy
from datetime import date
from pathlib import Path
import shutil

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment

from specification_catalog import FUNCTIONS, MESSAGES, PROCESS_STEPS, SCREENS, TECH_FLOWS


ROOT = Path(__file__).resolve().parents[2]
TEMPLATE = ROOT / "docs" / "sap490" / "templates" / "Deliverable_template" / "Functional_Specification.xlsx"
OUT = ROOT / "docs" / "sap490" / "generated"
DATE = date(2026, 7, 25)
VERSION = "0.7"


GROUPS = [
    ("AUTH", "Authentication and session", "SRS-FR-AUTH", "/odata/v4/auth/", "srv/auth.js", "qa:auth:programmatic"),
    ("BUG", "Bug creation and validation", "SRS-FR-BUG", "/odata/v4/bug/Bugs", "srv/bug-service/bug-write.js", "qa:idts41:programmatic"),
    ("ASSIGN", "Assignment and responsibility", "SRS-FR-ASG", "/odata/v4/bug/Bugs", "srv/bug-service/actions.js", "qa:idts67:programmatic"),
    ("LIFE", "Lifecycle actions", "SRS-FR-LIFE", "/odata/v4/bug/Bugs(...)/<action>", "srv/service.js", "qa:idts89:programmatic"),
    ("COLLAB", "Comments and attachments", "SRS-FR-COLLAB", "/odata/v4/bug/Comments; Attachments", "srv/bug-service/content.js", "qa:comments-attachments:programmatic"),
    ("MON", "Dashboard and monitoring", "SRS-FR-MON", "/odata/v4/bug/DeveloperWorkloads", "srv/bug-service/monitoring.js", "qa:pm-monitoring:programmatic"),
    ("NOTIFY", "In-app and email notification", "SRS-FR-NOTIFY", "/odata/v4/bug/NotificationDeliveries", "srv/email/worker.js", "qa:email-outbox:programmatic"),
    ("AI", "Human-reviewed AI assistance", "SRS-FR-AI", "/odata/v4/bug/<AI action>", "srv/ai/", "qa:idts72:acceptance"),
]

AI_SYMBOLS = [
    "acceptAiSuggestion",
    "rejectAiSuggestion",
    "ignoreAiSuggestion",
    "applyClassificationSuggestion",
    "confirmDuplicateSuggestion",
    "readAiOperationalMetrics",
    "operationStatus",
    "latencyMs",
]

LIFECYCLE_ACTIONS = [
    "assignToDeveloper",
    "moveToPendingAssignment",
    "markInReview",
    "requestMoreInformation",
    "resubmitToDeveloper",
    "rejectBug",
    "startProgress",
    "resolveBug",
    "sendToRetest",
    "closeBug",
    "reopenBug",
]


TEXT = {
    "en": {
        "title": "FUNCTIONAL SPECIFICATION",
        "name": "IDTS CAP/Fiori Functional Baseline",
        "history": "v0.6 corrects runtime traces, expands screen/message/processing records, and aligns acceptance wording with verified Shared QA evidence.",
        "n_a_smart": "N/A — Not applicable to SAP CAP/Fiori implementation. IDTS uses Fiori Elements/SAPUI5 pages and OData V4; it does not generate SAP Smart Forms.",
        "screen_sections": [
            ("Login", "login.html / login-page.js / ext/login/LoginController.js", "Custom authentication entry; safe error, no sign-up"),
            ("Profile menu", "ext/login/ProfileShell.js", "Signed-in identity, role and sign-out"),
            ("Dashboard", "dashboard.html / dashboard-page.js", "Role-aware KPI and queues"),
            ("Bug List Report", "manifest.json / annotations", "Search, filter, create and navigation"),
            ("Bug Object Page", "annotations/actions.cds", "Summary, classification, assignment, lifecycle and evidence"),
            ("Smart Assign", "SmartAssignDeveloper.js", "Filtered Developer selection and review-only explanation"),
            ("AI review dialogs", "ClassificationReview.js / DuplicateReview.js / HandoffSummaryReview.js / SmartAssignDeveloper.js", "Accept/reject/ignore; apply/confirm remain explicit actions"),
        ],
        "messages": [
            ("IDTS-MSG-400-REQ", "Required or invalid value.", "400", "prepareBugWrite / code-list and classification validators", "field target", "Tester/PM create or update", "request transaction rolls back", "Inline field message; sanitized validation log"),
            ("IDTS-MSG-401-AUTH", "Your session is missing or expired. Sign in again.", "401", "srv/auth/custom-auth.js", "request", "Anonymous/expired session", "no mutation", "Login redirect or safe error; no token detail"),
            ("IDTS-MSG-403-ROLE", "You are not authorized to perform this action.", "403", "enforceBugCreatePermission / enforceBugWritePermission / enforceActionPermission", "action or field", "Role outside permission", "no mutation", "Message dialog; safe actor/action log"),
            ("IDTS-MSG-409-AI", "This AI suggestion has already changed. Reload and review the latest state.", "409", "srv/ai/review.js / srv/ai/classification-apply.js", "AiSuggestions", "Authorized AI reviewer", "AI review/apply transaction rolls back", "Review dialog reloads; no generic lifecycle claim"),
            ("IDTS-MSG-ATTACH", "The attachment could not be processed safely.", "400/502", "BugCollaboration.js / srv/bug-service/content.js", "attachment", "Authorized Bug participant", "attachment write rolls back; Bug remains", "Upload control error; provider detail sanitized"),
            ("IDTS-MSG-EMAIL", "Email delivery failed; the bug update was kept.", "200 + delivery FAILED", "srv/email/worker.js", "NotificationDeliveries", "System worker", "workflow stays committed; delivery retry tracked", "No workflow error dialog; sanitized delivery log"),
            ("IDTS-MSG-AI", "AI assistance is unavailable. Continue with the normal workflow.", "200 safe fallback or sanitized provider error", "srv/ai/provider.js and focused AI action", "AI review dialog", "Authorized user", "no Bug workflow mutation", "Safe fallback/no-result state; no prompt/raw response"),
        ],
        "processing": {
            "AUTH": "login-page.js and LoginController.js call AuthService.login in srv/auth.js. The raw token is returned once to the browser; only its hash is stored in AuthSessions.tokenHash. ProfileShell.js later calls me/logout, while custom-auth.js resolves the bearer token for protected BugService requests.",
            "BUG": "Fiori editFlow.createDocument starts OData draft NEW. Field edits send PATCH to Bugs.drafts. SAVE activates through CREATE. prepareBugWrite validates required fields, active code lists, classification and server-derived bugNumber, reporter and status in one transaction.",
            "ASSIGN": "The value help reads AssignableDevelopers. assignToDeveloper or normal Bug write validates role, active Developer and DeveloperResponsibilities. No assignee produces Pending Assignment; a valid explicit assignee produces Assigned.",
            "LIFE": "The eleven bound actions are registered in srv/service.js. transitionBug or the specialized action validates role, reason, assignee and status transition, updates the Bug, derives nextProcessor, writes HistoryEvents/HistoryLogs and queues Notifications in the same transaction.",
            "COLLAB": "Comments are hidden until an active Bug exists. Before create SAVE, BugCollaboration.js holds selected files in pendingCreateAttachmentsByBugId (client memory). After Bug activation succeeds, it calls the attachment API; PostgreSQL stores metadata and S3 stores binary content. Delete removes the authorized record/object.",
            "MON": "Role-aware read models calculate workload, overdue, Pending Assignment, Rejected follow-up and Retest Required queues without granting write authority.",
            "NOTIFY": "The workflow transaction creates Notifications and NotificationDeliveries. The background worker claims PENDING/eligible FAILED rows, calls Brevo/SMTP privately and records SENT/FAILED/SKIPPED without rolling back the Bug workflow.",
            "AI": "Suggestion actions use allowlisted Bug data. Review actions acceptAiSuggestion, rejectAiSuggestion and ignoreAiSuggestion persist reviewer state. applyClassificationSuggestion and confirmDuplicateSuggestion are separate authorized mutations. readAiOperationalMetrics is PM-only; operationStatus and latencyMs contain sanitized telemetry. Live OpenAI remains disabled/not accepted.",
        },
    },
    "vi": {
        "title": "ĐẶC TẢ CHỨC NĂNG",
        "name": "Đường cơ sở chức năng IDTS CAP/Fiori",
        "history": "v0.6 sửa truy vết runtime, bổ sung bản ghi màn hình/thông báo/xử lý và đồng bộ nội dung nghiệm thu với bằng chứng Shared QA đã xác minh.",
        "n_a_smart": "N/A — Không áp dụng cho triển khai SAP CAP/Fiori. IDTS dùng trang Fiori Elements/SAPUI5 và OData V4; hệ thống không sinh SAP Smart Form.",
        "screen_sections": [
            ("Đăng nhập", "login.html / login-page.js / ext/login/LoginController.js", "Điểm vào xác thực tùy chỉnh; lỗi an toàn, không có đăng ký"),
            ("Menu hồ sơ", "ext/login/ProfileShell.js", "Danh tính, vai trò đang đăng nhập và đăng xuất"),
            ("Dashboard", "dashboard.html / dashboard-page.js", "KPI và hàng đợi theo vai trò"),
            ("Bug List Report", "manifest.json / annotations", "Tìm kiếm, lọc, tạo và điều hướng"),
            ("Bug Object Page", "annotations/actions.cds", "Tóm tắt, phân loại, phân công, vòng đời và bằng chứng"),
            ("Smart Assign", "SmartAssignDeveloper.js", "Lọc Developer và giải thích chỉ để tham khảo"),
            ("Hộp thoại đánh giá AI", "ClassificationReview.js / DuplicateReview.js / HandoffSummaryReview.js / SmartAssignDeveloper.js", "Chấp nhận/từ chối/bỏ qua; áp dụng/xác nhận là thao tác riêng có chủ ý"),
        ],
        "messages": [
            ("IDTS-MSG-400-REQ", "Thiếu dữ liệu hoặc giá trị không hợp lệ.", "400", "prepareBugWrite / bộ kiểm tra code list và phân loại", "trường dữ liệu", "Tester/PM tạo hoặc cập nhật", "hoàn tác request transaction", "Hiện lỗi tại đúng trường; log validation đã làm sạch"),
            ("IDTS-MSG-401-AUTH", "Phiên đăng nhập thiếu hoặc hết hạn. Vui lòng đăng nhập lại.", "401", "srv/auth/custom-auth.js", "request", "Ẩn danh/phiên hết hạn", "không ghi dữ liệu", "Chuyển về đăng nhập hoặc lỗi an toàn; không lộ token"),
            ("IDTS-MSG-403-ROLE", "Bạn không có quyền thực hiện thao tác này.", "403", "enforceBugCreatePermission / enforceBugWritePermission / enforceActionPermission", "thao tác hoặc trường", "Vai trò ngoài quyền", "không ghi dữ liệu", "Hộp thoại dễ hiểu; log actor/action an toàn"),
            ("IDTS-MSG-409-AI", "Gợi ý AI này đã thay đổi. Hãy tải lại và đánh giá trạng thái mới nhất.", "409", "srv/ai/review.js / srv/ai/classification-apply.js", "AiSuggestions", "Người đánh giá AI có quyền", "hoàn tác transaction đánh giá/áp dụng AI", "Tải lại hộp thoại đánh giá; không tuyên bố 409 chung cho lifecycle"),
            ("IDTS-MSG-ATTACH", "Không thể xử lý tệp đính kèm một cách an toàn.", "400/502", "BugCollaboration.js / srv/bug-service/content.js", "tệp đính kèm", "Người tham gia Bug có quyền", "hoàn tác ghi tệp; Bug vẫn giữ nguyên", "Lỗi tại upload control; chi tiết provider được làm sạch"),
            ("IDTS-MSG-EMAIL", "Gửi email thất bại; thay đổi của Bug vẫn được giữ.", "200 + delivery FAILED", "srv/email/worker.js", "NotificationDeliveries", "Worker hệ thống", "workflow đã commit; theo dõi retry delivery", "Không hiện lỗi workflow; log delivery đã làm sạch"),
            ("IDTS-MSG-AI", "AI đang không khả dụng. Hãy tiếp tục luồng bình thường.", "200 fallback an toàn hoặc lỗi provider đã làm sạch", "srv/ai/provider.js và AI action tương ứng", "hộp thoại đánh giá AI", "Người dùng có quyền", "không thay đổi workflow Bug", "Trạng thái fallback/không có kết quả; không lộ prompt/raw response"),
        ],
        "processing": {
            "AUTH": "login-page.js và LoginController.js gọi AuthService.login trong srv/auth.js. Raw token chỉ được trả một lần cho trình duyệt; database chỉ lưu bản băm trong AuthSessions.tokenHash. ProfileShell.js gọi me/logout, còn custom-auth.js resolve bearer token cho request BugService được bảo vệ.",
            "BUG": "Fiori editFlow.createDocument bắt đầu OData draft NEW. Sửa field gửi PATCH tới Bugs.drafts. SAVE kích hoạt draft qua CREATE. prepareBugWrite kiểm tra field bắt buộc, code list active, classification và tự sinh bugNumber, reporter, status trong một transaction.",
            "ASSIGN": "Value help đọc AssignableDevelopers. assignToDeveloper hoặc Bug write kiểm tra role, Developer active và DeveloperResponsibilities. Không có assignee thì Pending Assignment; assignee hợp lệ được chọn rõ ràng thì Assigned.",
            "LIFE": "Mười một bound action được đăng ký trong srv/service.js. transitionBug hoặc action chuyên biệt kiểm tra role, reason, assignee và status transition; sau đó update Bug, xác định nextProcessor, ghi HistoryEvents/HistoryLogs và tạo Notifications trong cùng transaction.",
            "COLLAB": "Comment bị ẩn cho đến khi có Bug active. Trước khi SAVE tạo Bug, BugCollaboration.js giữ file đã chọn trong pendingCreateAttachmentsByBugId (bộ nhớ phía client). Sau khi activate thành công, UI gọi attachment API; PostgreSQL lưu metadata và S3 lưu binary. Delete xóa record/object sau khi kiểm quyền.",
            "MON": "Read model theo role tính workload, overdue, Pending Assignment, Rejected follow-up và Retest Required nhưng không cấp quyền ghi.",
            "NOTIFY": "Workflow transaction tạo Notifications và NotificationDeliveries. Worker claim dòng PENDING/FAILED còn lượt, gọi Brevo/SMTP bằng cấu hình private rồi ghi SENT/FAILED/SKIPPED mà không rollback Bug.",
            "AI": "Suggestion action chỉ dùng dữ liệu Bug allowlist. acceptAiSuggestion, rejectAiSuggestion và ignoreAiSuggestion lưu trạng thái review. applyClassificationSuggestion và confirmDuplicateSuggestion là mutation riêng có kiểm quyền. readAiOperationalMetrics chỉ cho PM; operationStatus và latencyMs là telemetry đã làm sạch. OpenAI live vẫn disabled/not accepted.",
        },
    },
}


def writable(ws, coordinate):
    cell = ws[coordinate]
    if not isinstance(cell, MergedCell):
        return cell
    for merged in ws.merged_cells.ranges:
        if coordinate in merged:
            return ws.cell(merged.min_row, merged.min_col)
    raise ValueError(f"No merge anchor for {ws.title}!{coordinate}")


def write(ws, coordinate, value):
    cell = writable(ws, coordinate)
    cell.value = value


def write_wrapped(ws, coordinate, value, *, vertical="top"):
    cell = writable(ws, coordinate)
    cell.value = value
    alignment = copy(cell.alignment)
    cell.alignment = Alignment(
        horizontal=alignment.horizontal,
        vertical=vertical,
        text_rotation=alignment.text_rotation,
        wrap_text=True,
        shrink_to_fit=False,
        indent=alignment.indent,
    )


def make_text_visible(ws, coordinates):
    """Preserve template typography but force printable text to black."""
    for coordinate in coordinates:
        cell = writable(ws, coordinate)
        font = copy(cell.font)
        font.color = "FF000000"
        cell.font = font


def merge_row(ws, row, start_column, end_column):
    target = f"{start_column}{row}:{end_column}{row}"
    if target not in {str(item) for item in ws.merged_cells.ranges}:
        ws.merge_cells(target)


def clear_rows(ws, start, end):
    for row in ws.iter_rows(min_row=start, max_row=end):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None


def unmerge_region(ws, min_row, max_row, min_col, max_col):
    """Remove template merges only inside a generated table body."""
    for merged in list(ws.merged_cells.ranges):
        if not (
            merged.max_row < min_row or merged.min_row > max_row
            or merged.max_col < min_col or merged.min_col > max_col
        ):
            ws.unmerge_cells(str(merged))


def copy_style(source, target):
    target._style = copy(source._style)
    target.number_format = source.number_format
    target.protection = copy(source.protection)


def write_formal_table(ws, start_row, groups, headers, rows, *, header_source, data_source, row_height=42):
    """Write one record per row using logical column groups from the official grid."""
    end_row = start_row + len(rows)
    min_col = min(ws[f"{start}1"].column for start, _ in groups)
    max_col = max(ws[f"{end}1"].column for _, end in groups)
    unmerge_region(ws, start_row, end_row, min_col, max_col)
    clear_rows(ws, start_row, end_row)

    header_template = ws[header_source]
    data_template = ws[data_source]
    for row_offset, values in enumerate([headers, *rows]):
        row = start_row + row_offset
        is_header = row_offset == 0
        template = header_template if is_header else data_template
        for (start, end), value in zip(groups, values):
            start_col = ws[f"{start}1"].column
            end_col = ws[f"{end}1"].column
            for column in range(start_col, end_col + 1):
                cell = ws.cell(row, column)
                copy_style(template, cell)
                cell.alignment = Alignment(
                    horizontal="center" if is_header else "left",
                    vertical="center" if is_header else "top",
                    wrap_text=True,
                    shrink_to_fit=False,
                )
            if start != end:
                ws.merge_cells(f"{start}{row}:{end}{row}")
            ws[f"{start}{row}"] = value
        ws.row_dimensions[row].height = 30 if is_header else row_height
    return end_row


def set_print_region(ws, area, *, title_rows=None):
    ws.print_area = area
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def remove_broken_defined_names(workbook):
    broken = [
        name
        for name, item in workbook.defined_names.items()
        if "#REF!" in str(item.attr_text)
    ]
    for name in broken:
        del workbook.defined_names[name]


def metadata(ws):
    for row in (2, 3, 4):
        for column in ("AS", "AT", "AU", "AV", "AW", "AX", "BB", "BC", "BD", "BE"):
            try:
                write(ws, f"{column}{row}", None)
            except ValueError:
                pass
    if ws.title == "Function Overview":
        values = {
            "I3": "IDTS-FS",
            "AB3": "IDTS CAP/Fiori",
            "AT2": " DonHV",
            "BC2": DATE.isoformat(),
            "AT3": " DonHV",
            "BE3": DATE.isoformat(),
            "AT4": " Mentor / Supervisor",
            "BE4": "Pending",
        }
    elif ws.title == "Screen Layout":
        values = {
            "AS2": " DonHV",
            "BB2": DATE.isoformat(),
            "AS3": " Mentor / Supervisor",
            "BD3": "Pending",
        }
    else:
        date_column = "BC" if ws.title == "Message Definition" else "BB"
        values = {
            "I3": "IDTS-FS",
            "W3": "IDTS CAP/Fiori",
            "AS2": " DonHV",
            f"{date_column}2": DATE.isoformat(),
            "AS3": " DonHV",
            f"{date_column}3": DATE.isoformat(),
            "AS4": " Mentor / Supervisor",
            f"{date_column}4": "Pending",
        }
    for coordinate, value in values.items():
        try:
            write(ws, coordinate, value)
        except ValueError:
            continue


def localize_visible_functional_labels(workbook):
    translations = {
        "Functional Specification": "Đặc tả chức năng",
        "FUNCTIONAL SPECIFICATION": "ĐẶC TẢ CHỨC NĂNG",
        "Function Overview": "Tổng quan chức năng",
        "Process Flow": "Luồng quy trình",
        "Screen Layout": "Bố cục màn hình",
        "Screen Definition": "Định nghĩa màn hình",
        "Smart Form Structure": "Cấu trúc Smart Form",
        "Message Definition": "Định nghĩa thông báo",
        "Processing Description": "Mô tả xử lý",
        "Function ID": "Mã chức năng",
        "Function Name": "Tên chức năng",
        "Created Date": "Ngày tạo",
        "Last Update Date": "Ngày cập nhật cuối",
        "Creator": "Người tạo",
        "Reviewer": "Người review",
        "Approver": "Người phê duyệt",
        "Version": "Phiên bản",
        "Description": "Mô tả",
        "Modified date": "Ngày sửa",
        "Modified by": "Người sửa",
        "Pending": "Chờ duyệt",
    }
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell, MergedCell) and isinstance(cell.value, str):
                    normalized = cell.value.strip()
                    if normalized in translations:
                        leading = cell.value[: len(cell.value) - len(cell.value.lstrip())]
                        cell.value = leading + translations[normalized]


def build(language):
    labels = TEXT[language]
    output = OUT / f"Functional_Specification_IDTS_SAP01_{language}_v{VERSION}.xlsx"
    OUT.mkdir(parents=True, exist_ok=True)
    shutil.copy2(TEMPLATE, output)
    workbook = load_workbook(output)
    remove_broken_defined_names(workbook)

    cover = workbook["Cover"]
    for coordinate, value in {
        "B8": labels["title"],
        "N11": "IDTS",
        "Z11": "Issue and Defect Tracking System in SAP",
        "N12": DATE.isoformat(),
        "Z12": DATE.isoformat(),
        "AE17": "DonHV",
        "Z17": "Pending",
        "U17": "Pending",
    }.items():
        write(cover, coordinate, value)
    set_print_region(cover, "B1:BH24")

    history = workbook["Histories"]
    clear_rows(history, 3, 10)
    history_rows = {
        "en": [
            ("0.1", "Initial functional baseline", "Cover and initial process scope", date(2026, 6, 21)),
            ("0.2", "Sprint workflow alignment", "Function overview and lifecycle flow", date(2026, 7, 2)),
            ("0.3", "Shared QA and integration", "Attachment, email and monitoring behavior", date(2026, 7, 24)),
            ("0.4", "AI advisory baseline", "Human-review AI functions and boundaries", date(2026, 7, 24)),
            ("0.5", "Official-template remediation", "All 9 sheets; SAP sample residue removed", date(2026, 7, 25)),
            ("0.6", labels["history"], "Runtime trace, completeness and formality", DATE),
            ("0.7", "Formal table and traceability remediation", "Function, process, screen, message and processing tables", DATE),
        ],
        "vi": [
            ("0.1", "Nền đặc tả chức năng ban đầu", "Trang bìa và phạm vi quy trình ban đầu", date(2026, 6, 21)),
            ("0.2", "Đồng bộ quy trình Sprint", "Tổng quan chức năng và luồng vòng đời", date(2026, 7, 2)),
            ("0.3", "Shared QA và tích hợp", "Hành vi tệp đính kèm, email và giám sát", date(2026, 7, 24)),
            ("0.4", "Nền AI tư vấn", "Chức năng AI có con người đánh giá và ranh giới", date(2026, 7, 24)),
            ("0.5", "Khắc phục theo template chính thức", "Đủ 9 sheet; loại nội dung mẫu SAP", date(2026, 7, 25)),
            ("0.6", labels["history"], "Truy vết runtime, độ đầy đủ và tính trang trọng", DATE),
            ("0.7", "Chuẩn hóa bảng và truy vết", "Bảng chức năng, quy trình, màn hình, thông báo và xử lý", DATE),
        ],
    }[language]
    for index, (version, summary, affected, changed_date) in enumerate(history_rows, 3):
        for coordinate, value in {
            f"B{index}": index - 2,
            f"C{index}": version,
            f"D{index}": summary,
            f"E{index}": affected,
            f"F{index}": changed_date.isoformat(),
            f"G{index}": "DonHV",
        }.items():
            write(history, coordinate, value)
        history.row_dimensions[index].height = 34
    set_print_region(history, "B1:G10", title_rows="1:2")

    overview = workbook["Function Overview"]
    metadata(overview)
    write(overview, "I3", "IDTS-FS")
    write(overview, "AB3", labels["name"])
    write(overview, "AB6", "SRS-FR-AUTH; SRS-FR-BUG; SRS-FR-ASG; SRS-FR-LIFE; SRS-FR-COLLAB; SRS-FR-MON; SRS-FR-NOTIFY; SRS-FR-AI")
    function_headers = {
        "en": ["Function ID", "Function name", "Business objective", "Actor", "Trigger", "Expected result", "Requirement"],
        "vi": ["Mã chức năng", "Tên chức năng", "Mục tiêu nghiệp vụ", "Vai trò", "Điểm kích hoạt", "Kết quả mong đợi", "Yêu cầu"],
    }[language]
    function_rows = [
        [item["id"], item["name"][language], item["objective"][language], item["actors"],
         item["trigger"][language], item["result"][language], item["requirement"]]
        for item in FUNCTIONS
    ]
    end_row = write_formal_table(
        overview, 15,
        [("B", "F"), ("G", "N"), ("O", "Z"), ("AA", "AE"), ("AF", "AJ"), ("AK", "AS"), ("AT", "BH")],
        function_headers, function_rows, header_source="B14", data_source="B15", row_height=54,
    )
    write_wrapped(
        overview, f"B{end_row + 2}",
        ("Baseline and evidence are frozen in the IDTS-103 evidence report; every function maps to an exact source and focused test."
         if language == "en" else
         "Baseline và bằng chứng được cố định trong báo cáo IDTS-103; mỗi chức năng được map tới source và test tập trung chính xác."),
    )
    merge_row(overview, end_row + 2, "B", "BH")
    overview.row_dimensions[end_row + 2].height = 36
    set_print_region(overview, f"B1:BH{end_row + 2}", title_rows="1:14")

    process = workbook["Process Flow"]
    metadata(process)
    process._images = []
    process_headers = {
        "en": ["Step", "Actor", "Action", "Status before", "Status after", "Next processor", "Side effect"],
        "vi": ["Bước", "Vai trò", "Hành động", "Trạng thái trước", "Trạng thái sau", "Người xử lý tiếp", "Side effect"],
    }[language]
    process_rows = []
    for step, actor, action, before, after, next_processor, side_effect in PROCESS_STEPS:
        if language == "vi":
            action = {
                "Sign in": "Đăng nhập", "Create Bug draft and save": "Tạo draft Bug và lưu",
                "Review and work on assigned Bug": "Review và xử lý Bug được giao",
                "Request more information": "Yêu cầu thêm thông tin",
                "Resubmit corrected information": "Gửi lại thông tin đã sửa",
                "Resolve Bug": "Resolve Bug", "Retest and close or reopen": "Retest rồi đóng hoặc mở lại",
            }.get(action, action)
            next_processor = {
                "Authenticated user/session": "Người dùng/phiên đã xác thực", "Developer or Tester/PM": "Developer hoặc Tester/PM",
                "Developer": "Developer", "Reporter Tester": "Tester báo lỗi", "Tester / PM": "Tester / PM", "None / Developer": "Không có / Developer",
            }.get(next_processor, next_processor)
            side_effect = {
                "Auth session token issued once": "Token phiên xác thực được cấp một lần",
                "Bug, history and notification commit": "Commit Bug, lịch sử và thông báo",
                "Exact action history": "Lịch sử action chính xác", "Notification to reporter": "Thông báo cho reporter",
                "History and notification": "Lịch sử và thông báo", "Resolution history and notification": "Lịch sử resolve và thông báo",
                "Final or reopened history": "Lịch sử đóng hoặc mở lại",
            }.get(side_effect, side_effect)
        process_rows.append([step, actor, action, before, after, next_processor, side_effect])
    process_end = write_formal_table(
        process, 6,
        [("B", "E"), ("F", "J"), ("K", "R"), ("S", "W"), ("X", "AB"), ("AC", "AH"), ("AI", "AP")],
        process_headers, process_rows, header_source="B5", data_source="B6", row_height=46,
    )
    flow_image = XLImage(ROOT / "docs" / "diagrams" / "rendered" / "png" / "04-end-to-end-defect-flow.png")
    flow_image.width = 560
    flow_image.height = 315
    process.add_image(flow_image, "AR6")
    write(process, "B44", "Supplement / Bổ sung: source baseline 8009b2a6a72d73db28f190b3a0bcbb65b1ff4740")
    process_supplement = [
        "Role flow: Tester/PM creates and assigns; the assigned Developer reviews and resolves; Tester/PM retests, closes or reopens.",
        "Control flow: every write is revalidated by CAP authorization, code-list, assignee and lifecycle rules before commit.",
        "Atomic side effects: Bug state, next processor, exact-action history and in-app notification commit in one request transaction.",
        "External boundaries: PostgreSQL stores business metadata, S3 stores attachment binary, and the email worker runs after workflow commit.",
    ]
    if language == "vi":
        process_supplement = [
            "Luồng vai trò: Tester/PM tạo và phân công; Developer được giao việc review/xử lý; Tester/PM kiểm thử lại, đóng hoặc mở lại.",
            "Luồng kiểm soát: mọi thao tác ghi đều được CAP kiểm tra lại quyền, code list, assignee và vòng đời trước khi commit.",
            "Side effect nguyên tử: trạng thái Bug, người xử lý tiếp theo, exact-action history và thông báo trong ứng dụng commit trong cùng request transaction.",
            "Ranh giới ngoài: PostgreSQL lưu metadata nghiệp vụ, S3 lưu binary attachment, email worker chạy sau workflow commit.",
        ]
    for row, text in enumerate(process_supplement, 45):
        merge_row(process, row, "B", "BG")
        write_wrapped(process, f"B{row}", text)
        process.row_dimensions[row].height = 34
    set_print_region(process, f"B1:BH{max(process_end, 48)}", title_rows="1:5")

    layout = workbook["Screen Layout"]
    metadata(layout)
    clear_rows(layout, 28, 78)
    screen_headers = {
        "en": ["Screen ID", "Screen name", "Page type", "Role", "Main controls/areas", "Navigation", "Technical binding"],
        "vi": ["Mã màn hình", "Tên màn hình", "Loại trang", "Vai trò", "Control/khu vực chính", "Điều hướng", "Binding kỹ thuật"],
    }[language]
    screen_rows = []
    for screen_id, name_en, name_vi, page_type, entry, controller, binding, role, main_areas, navigation in SCREENS:
        if language == "vi":
            page_type = {
                "Custom SAPUI5 page": "Trang SAPUI5 tùy chỉnh", "SAPUI5 popover": "Popover SAPUI5",
                "Fiori Elements List Report": "Fiori Elements List Report", "Fiori Elements Object Page": "Fiori Elements Object Page",
                "Object Page section": "Section Object Page", "Application section/popover": "Section/popover ứng dụng",
                "Dialog/value help": "Dialog/value help", "Dialog": "Dialog",
            }.get(page_type, page_type)
            main_areas = {
                "Email, password, safe message": "Email, password, thông báo an toàn", "Name, email, role, Sign Out": "Tên, email, role, Đăng xuất",
                "KPI cards, queues, workload": "KPI card, hàng đợi, workload", "Filters, table, Create": "Bộ lọc, bảng, Tạo",
                "Summary, classification, assignment, lifecycle": "Tóm tắt, phân loại, phân công, vòng đời",
                "Thread and Add Comment": "Luồng bình luận và Thêm bình luận", "Upload, download, delete": "Tải lên, tải xuống, xóa",
                "Timeline and Show More": "Timeline và Xem thêm", "Read state and delivery status": "Trạng thái đọc và delivery",
                "Search, workload, responsibility, explanation": "Tìm kiếm, workload, responsibility, giải thích",
                "Accept, reject, ignore, apply": "Chấp nhận, từ chối, bỏ qua, áp dụng", "Candidates, review, confirm": "Ứng viên, review, xác nhận",
                "Grounded summary and review": "Tóm tắt có căn cứ và review",
            }.get(main_areas, main_areas)
            navigation = {
                "Dashboard / protected app": "Dashboard / ứng dụng được bảo vệ", "Login after logout": "Đăng nhập sau khi logout",
                "Bug List Report": "Bug List Report", "Bug Object Page": "Bug Object Page", "Focused sections/dialogs": "Section/dialog chuyên trách",
                "Same Object Page": "Cùng Object Page", "Related Bug": "Bug liên quan", "Assignment section": "Section phân công",
                "Classification section": "Section phân loại", "Bug summary": "Tóm tắt Bug", "History section": "Section lịch sử",
            }.get(navigation, navigation)
        screen_rows.append([screen_id, name_en if language == "en" else name_vi, page_type, role, main_areas, navigation, f"{entry}; {controller}; {binding}"])
    layout_end = write_formal_table(
        layout, 27,
        [("B", "E"), ("F", "M"), ("N", "R"), ("S", "Z"), ("AA", "AH"), ("AI", "AP"), ("AQ", "BH")],
        screen_headers, screen_rows, header_source="C27", data_source="C28", row_height=54,
    )
    set_print_region(layout, f"B1:BH{layout_end}", title_rows="1:27")

    definition = workbook["Screen Definition"]
    metadata(definition)
    clear_rows(definition, 12, 74)
    write(definition, "B9", "1. IDTS Fiori screens")
    write(definition, "B12", "1.1 Main fields and actions")
    fields = [
        ("Title", "Edm.String", "I/O", "Single", "Bugs.title", "255", "Yes", "Text", "No", "Required; backend validated"),
        ("Description", "Edm.String", "I/O", "Single", "Bugs.description", "-", "Yes", "Text area", "No", "Required during activation"),
        ("Steps to Reproduce", "Edm.String", "I/O", "Single", "Bugs.stepsToReproduce", "-", "Yes", "Text area", "No", "Required during activation"),
        ("Actual Result", "Edm.String", "I/O", "Single", "Bugs.actualResult", "-", "Yes", "Text area", "No", "Required during activation"),
        ("Expected Result", "Edm.String", "I/O", "Single", "Bugs.expectedResult", "-", "Yes", "Text area", "No", "Required during activation"),
        ("Priority", "Association", "I/O", "Single", "Bugs.priority", "-", "Yes", "Fixed value help", "Yes", "Active code only"),
        ("Severity", "Association", "I/O", "Single", "Bugs.severity", "-", "Yes", "Fixed value help", "Yes", "Active code only"),
        ("Environment", "Association", "I/O", "Single", "Bugs.environment", "-", "Yes", "Fixed value help", "Yes", "Active code only"),
        ("Application Component", "Association", "I/O", "Single", "Bugs.applicationComponent", "-", "Yes", "Value help", "Yes", "Active classification"),
        ("Defect Category", "Association", "I/O", "Single", "Bugs.defectCategory", "-", "Yes", "Value help", "Yes", "Must match Component Category"),
        ("SAP Module", "Association", "I/O", "Single", "Bugs.sapModule", "-", "No", "Value help", "Yes", "Optional assignment context"),
        ("Assignee", "Association", "I/O", "Single", "Bugs.assignee", "-", "No", "Smart value help", "Yes", "Tester/PM; responsibility validated"),
        ("Status", "Association", "O", "Single", "Bugs.status", "-", "Yes", "Semantic status", "No", "Backend-owned"),
        ("Current Action Owner", "Association", "O", "Single", "Bugs.nextProcessorUser", "-", "No", "Display", "No", "Distinct from technical assignee"),
        ("Comment", "Composition", "I/O", "Multi", "Bugs.comments", "-", "No", "Thread/list", "No", "Hidden during create; available after save"),
        ("Attachment", "Composition", "I/O", "Multi", "Bugs.attachments", "-", "No", "Upload set", "No", "Pending client memory → API → PostgreSQL/S3"),
        ("History", "Composition", "O", "Multi", "Bugs.historyEvents", "-", "No", "Timeline/show more", "No", "Exact action, actor and changed fields"),
        ("Notification", "Composition", "O", "Multi", "Notifications", "-", "No", "List/status", "No", "In-app plus email delivery state"),
        ("Lifecycle Actions", "Bound actions", "I/O", "Multi", "BugService.<action>", "-", "No", "Action buttons", "No", "Visibility advisory; backend authorizes"),
        ("AI Review", "Bound actions", "I/O", "Multi", "AiSuggestions", "-", "No", "Review dialog", "No", "Review does not mutate Bug workflow"),
    ]
    field_rows = [13, 14, 15, 17, 20, 22, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38]
    for index, (record, row) in enumerate(zip(fields, field_rows), 1):
        name, kind, io, cardinality, binding, length, required, fmt, help_flag, remark = record
        if language == "vi":
            name = {
                "Title": "Tiêu đề", "Description": "Mô tả", "Steps to Reproduce": "Các bước tái hiện",
                "Actual Result": "Kết quả thực tế", "Expected Result": "Kết quả mong đợi", "Priority": "Độ ưu tiên",
                "Severity": "Mức độ nghiêm trọng", "Environment": "Môi trường", "Application Component": "Application Component",
                "Defect Category": "Defect Category", "SAP Module": "SAP Module", "Assignee": "Người được phân công",
                "Status": "Trạng thái", "Current Action Owner": "Người xử lý hiện tại", "Comment": "Bình luận",
                "Attachment": "Tệp đính kèm", "History": "Lịch sử", "Notification": "Thông báo",
                "Lifecycle Actions": "Action vòng đời", "AI Review": "Review AI",
            }[name]
            required = "Có" if required == "Yes" else "Không"
            help_flag = "Có" if help_flag == "Yes" else "Không"
            remark = {
                "Required; backend validated": "Bắt buộc; backend kiểm tra",
                "Required during activation": "Bắt buộc khi activate",
                "Active code only": "Chỉ nhận code active",
                "Active classification": "Chỉ nhận phân loại active",
                "Must match Component Category": "Phải khớp Component Category",
                "Optional assignment context": "Ngữ cảnh phân công tùy chọn",
                "Tester/PM; responsibility validated": "Tester/PM; kiểm tra responsibility",
                "Backend-owned": "Backend quản lý",
                "Distinct from technical assignee": "Khác với technical assignee",
                "Hidden during create; available after save": "Ẩn khi create; có sau khi save",
                "Pending client memory → API → PostgreSQL/S3": "Bộ nhớ tạm client → API → PostgreSQL/S3",
                "Exact action, actor and changed fields": "Action, actor và field thay đổi chính xác",
                "In-app plus email delivery state": "Trạng thái trong ứng dụng và email delivery",
                "Visibility advisory; backend authorizes": "Visibility chỉ hỗ trợ UX; backend phân quyền",
                "Review does not mutate Bug workflow": "Review không mutation workflow Bug",
            }[remark]
        values = {
            f"B{row}": index,
            f"C{row}": name,
            f"I{row}": kind,
            f"M{row}": io,
            f"O{row}": cardinality,
            f"R{row}": binding,
            f"U{row}": length,
            f"AA{row}": required,
            f"AI{row}": fmt,
            f"AQ{row}": help_flag,
            f"AW{row}": remark,
        }
        for coordinate, value in values.items():
            write(definition, coordinate, value)
        definition.row_dimensions[row].height = max(definition.row_dimensions[row].height or 19.5, 34)
        make_text_visible(definition, list(values))
    make_text_visible(
        definition,
        ["B9", "B10", "R10", "AW10", "C11", "I11", "M11", "O11", "R11", "U11",
         "X11", "AA11", "AE11", "AI11", "AM11", "AQ11", "B12"],
    )
    set_print_region(definition, "B1:BH38", title_rows="1:12")

    smart = workbook["Smart Form Structure"]
    smart._images = []
    merge_row(smart, 2, "B", "AG")
    merge_row(smart, 4, "B", "AG")
    write_wrapped(smart, "B2", "Smart Form Structure — N/A", vertical="center")
    write_wrapped(smart, "B4", labels["n_a_smart"])
    smart.row_dimensions[2].height = 28
    smart.row_dimensions[4].height = 52
    clear_rows(smart, 44, 57)
    write(smart, "B44", 1)
    write(smart, "E44", "Applicability")
    write(smart, "M44", "N/A")
    write(smart, "R44", "Entire sheet")
    write(smart, "Z44", labels["n_a_smart"])
    smart.row_dimensions[44].height = 54
    set_print_region(smart, "B1:AG44", title_rows="1:4")

    messages = workbook["Message Definition"]
    metadata(messages)
    clear_rows(messages, 6, 40)
    message_headers = {
        "en": ["Message ID", "User message", "Business trigger", "Screen/target", "Role", "Timing", "Expected UI behavior"],
        "vi": ["Mã thông báo", "Nội dung người dùng", "Điều kiện nghiệp vụ", "Màn hình/target", "Vai trò", "Thời điểm", "Hành vi UI mong đợi"],
    }[language]
    message_rows = [
        [item["id"], item["message"][language], item["trigger"][language], f"{item['screen']} / {item['target']}",
         item["role"], item["timing"][language], item["ui"][language]]
        for item in MESSAGES
    ]
    message_end = write_formal_table(
        messages, 5,
        [("B", "E"), ("F", "N"), ("O", "Z"), ("AA", "AF"), ("AG", "AM"), ("AN", "AV"), ("AW", "BH")],
        message_headers, message_rows, header_source="B5", data_source="B6", row_height=58,
    )
    set_print_region(messages, f"B1:BH{message_end}", title_rows="1:5")

    processing = workbook["Processing Description"]
    metadata(processing)
    clear_rows(processing, 6, 80)
    processing_headers = {
        "en": ["Flow ID", "UI trigger", "OData operation", "Business validation", "Result", "Side effect", "Evidence"],
        "vi": ["Mã luồng", "Điểm kích hoạt UI", "Thao tác OData", "Kiểm tra nghiệp vụ", "Kết quả", "Side effect", "Bằng chứng"],
    }[language]
    processing_rows = []
    for flow_id, trigger, http, _service, handler, _transaction, effect, response, evidence in TECH_FLOWS:
        if language == "vi":
            trigger = {
                "FLOW-AUTH": "Gửi biểu mẫu đăng nhập", "FLOW-DRAFT-CREATE": "Tạo và chỉnh draft", "FLOW-ACTIVE-EDIT": "Chỉnh Bug active và lưu",
                "FLOW-ASSIGN": "Xác nhận Developer", "FLOW-LIFECYCLE": "Gọi action vòng đời", "FLOW-COLLAB": "Thêm bình luận hoặc tệp",
                "FLOW-MON": "Mở dashboard/bộ lọc", "FLOW-EMAIL": "Xử lý dòng outbox", "FLOW-AI": "Yêu cầu/review/apply hỗ trợ AI",
            }[flow_id]
            effect = {
                "FLOW-AUTH": "Đọc Users; thêm AuthSessions", "FLOW-DRAFT-CREATE": "Thêm Bugs; lịch sử/thông báo",
                "FLOW-ACTIVE-EDIT": "Cập nhật Bugs; lịch sử/thông báo", "FLOW-ASSIGN": "Bug/next processor/lịch sử/thông báo",
                "FLOW-LIFECYCLE": "Bug/trạng thái/lịch sử/thông báo", "FLOW-COLLAB": "Bộ nhớ tạm client; metadata PostgreSQL; binary S3",
                "FLOW-MON": "Không mutation", "FLOW-EMAIL": "Trạng thái/retry NotificationDeliveries",
                "FLOW-AI": "AiSuggestions.operationStatus/latencyMs; tùy chọn mutation Bug/DuplicateLink rõ ràng",
            }[flow_id]
            response = {
                "FLOW-AUTH": "Token chỉ trả một lần hoặc 401 an toàn", "FLOW-DRAFT-CREATE": "Bug active hoặc lỗi validation an toàn",
                "FLOW-ACTIVE-EDIT": "Bug active đã cập nhật", "FLOW-ASSIGN": "Assigned hoặc 400/403 an toàn",
                "FLOW-LIFECYCLE": "Transition chính xác được cho phép", "FLOW-COLLAB": "Evidence được lưu hoặc lỗi an toàn",
                "FLOW-MON": "KPI theo vai trò", "FLOW-EMAIL": "SENT/FAILED/SKIPPED", "FLOW-AI": "Kết quả review hoặc fallback an toàn",
            }[flow_id]
        processing_rows.append([flow_id, trigger, http, handler, response, effect, evidence])
    processing_end = write_formal_table(
        processing, 6,
        [("B", "F"), ("G", "O"), ("P", "W"), ("X", "AE"), ("AF", "AM"), ("AN", "AV"), ("AW", "BH")],
        processing_headers, processing_rows, header_source="B4", data_source="B6", row_height=60,
    )
    set_print_region(processing, f"B1:BH{processing_end}", title_rows="1:6")

    if language == "vi":
        localize_visible_functional_labels(workbook)
    workbook.properties.title = f"IDTS SAP490 Functional Specification {language.upper()} v{VERSION}"
    workbook.properties.subject = "Official-template functional baseline for CAP/Fiori"
    workbook.properties.creator = "IDTS SAP01 Team"
    workbook.save(output)
    return output


def main():
    for language in ("en", "vi"):
        print(build(language).relative_to(ROOT))


if __name__ == "__main__":
    main()
