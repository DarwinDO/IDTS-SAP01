"""Generate Technical Specification and Configuration Note from official templates."""

from __future__ import annotations

from copy import copy
from datetime import date
from pathlib import Path
import shutil

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment

from specification_catalog import MESSAGES, SCREENS, TECH_FLOWS, TECH_REQUIREMENTS


ROOT = Path(__file__).resolve().parents[2]
TEMPLATES = ROOT / "docs" / "sap490" / "templates" / "Deliverable_template"
OUT = ROOT / "docs" / "sap490" / "generated"
DATE = date(2026, 7, 25)
TECH_VERSION = "0.6"
CONFIG_VERSION = "0.5"

TECH_TEXT = {
    "en": {
        "title": "TECHNICAL SPECIFICATION",
        "name": "IDTS CAP/Fiori Technical Baseline",
        "history": "v0.5 corrects exact files/symbols, completes assignment/monitoring trace, and formalizes transaction, provider, migration, and failure paths.",
        "intro": [
            "IDTS is a SAP CAP Node.js application exposed through OData V4 and consumed by SAP Fiori Elements/SAPUI5.",
            "CDS defines the domain and service contract. Node.js handlers enforce authentication, authorization, validation, transactions, audit/history, notifications, attachments and AI review boundaries.",
        ],
        "scope": [
            "In scope: AuthService, BugService, draft/active writes, eleven lifecycle actions, comments, attachments, history, notifications, monitoring and human-reviewed AI.",
            "Out of scope: ABAP/RAP implementation, SAP Transport Requests, autonomous AI decisions, final production topology and live OpenAI acceptance.",
        ],
        "assumptions": [
            "Local development uses SQLite. Shared QA uses Render and PostgreSQL. SAP HANA remains a future deployment option, not the accepted current runtime.",
            "Attachment metadata is stored in PostgreSQL and binary content in AWS S3. Email uses an outbox with Brevo/SMTP private configuration.",
            "OpenAI live remains DISABLED / NOT ACCEPTED. Mock, fallback and no-mutation tests do not prove a live provider.",
        ],
        "standards": [
            ("Naming Convention", "CAP entities/services use stable business names; JavaScript symbols use camelCase; constants use explicit codes."),
            ("Table Design", "CDS entities define UUID keys, associations/compositions, constraints and audit fields. PostgreSQL naming follows CAP deployment output."),
            ("Function Group", "N/A — classic ABAP Function Groups are not used; modules are organized under srv/bug-service, srv/auth and srv/ai."),
            ("Function Module", "N/A — classic ABAP Function Modules are not used; CAP event handlers and JavaScript helpers provide the behavior."),
            ("Code Comment", "Non-obvious entry points explain trigger, input, decision, side effect, next dependency and breakpoint in Vietnamese; mirrors remain bilingual."),
            ("Pretty Printer", "Use repository lint/format conventions; do not reformat unrelated files."),
            ("Test Program", "Focused programmatic, API, integration and browser tests are required with positive, negative, role and persistence evidence."),
            ("Custom Role", "Tester, Developer and PM are business roles resolved server-side from the authenticated IDTS user."),
            ("RAP Model for My Inbox display", "N/A — IDTS uses CAP OData read models and Fiori Elements rather than ABAP RAP My Inbox."),
            ("DB Query", "Use CAP query API and request transaction; validate before mutation and keep workflow/history/notification consistency."),
            ("Texts in Code", "UI text is human-facing and sanitized; internal SQL, stack traces, credentials and developer-only copy are forbidden."),
            ("System Program & Environment", "Configuration is profile/environment based; secrets remain private and are never stored in source or evidence."),
            ("Graphics", "Fiori/UI5 controls and official diagram assets are used; no copied mock HTML/CSS in runtime."),
            ("Catalogs", "Priority, severity, environment and classification values are active code-list records validated by CAP."),
        ],
    },
    "vi": {
        "title": "ĐẶC TẢ KỸ THUẬT",
        "name": "Đường cơ sở kỹ thuật IDTS CAP/Fiori",
        "history": "v0.5 sửa đúng file/symbol, hoàn thiện truy vết phân công/giám sát và mô tả trang trọng transaction, provider, migration cùng failure path.",
        "intro": [
            "IDTS là ứng dụng SAP CAP Node.js, expose qua OData V4 và được SAP Fiori Elements/SAPUI5 sử dụng.",
            "CDS định nghĩa domain và service contract. Handler Node.js kiểm soát authentication, authorization, validation, transaction, audit/history, notification, attachment và ranh giới AI review.",
        ],
        "scope": [
            "Trong phạm vi: AuthService, BugService, draft/active write, mười một lifecycle action, comment, attachment, history, notification, monitoring và AI có human review.",
            "Ngoài phạm vi: triển khai ABAP/RAP, SAP Transport Request, AI tự quyết định, topology production cuối cùng và nghiệm thu OpenAI live.",
        ],
        "assumptions": [
            "Local dùng SQLite. Shared QA dùng Render và PostgreSQL. SAP HANA là hướng tương lai, không phải runtime hiện đã nghiệm thu.",
            "Metadata attachment lưu PostgreSQL, binary lưu AWS S3. Email dùng outbox với cấu hình Brevo/SMTP private.",
            "OpenAI live vẫn DISABLED / NOT ACCEPTED. Test mock, fallback và no-mutation không chứng minh provider live.",
        ],
        "standards": [
            ("Quy ước đặt tên", "Entity/service CAP dùng tên nghiệp vụ ổn định; symbol JavaScript dùng camelCase; constant dùng code rõ ràng."),
            ("Thiết kế bảng", "Entity CDS định nghĩa UUID, association/composition, constraint và audit field. PostgreSQL theo output deploy của CAP."),
            ("Function Group", "N/A — không dùng ABAP Function Group; module được tổ chức trong srv/bug-service, srv/auth và srv/ai."),
            ("Function Module", "N/A — không dùng ABAP Function Module; CAP event handler và helper JavaScript đảm nhiệm hành vi."),
            ("Comment code", "Entry point khó phải giải thích trigger, input, quyết định, side effect, dependency tiếp theo và breakpoint bằng tiếng Việt; mirror song ngữ."),
            ("Định dạng code", "Theo lint/format của repository; không reformat file không liên quan."),
            ("Chương trình test", "Bắt buộc có test programmatic/API/integration/browser với positive, negative, role và persistence evidence."),
            ("Role tùy chỉnh", "Tester, Developer và PM là business role được backend resolve từ IDTS user đã đăng nhập."),
            ("RAP My Inbox", "N/A — IDTS dùng CAP OData read model và Fiori Elements, không dùng ABAP RAP My Inbox."),
            ("DB Query", "Dùng CAP query API và request transaction; validate trước mutation, giữ nhất quán workflow/history/notification."),
            ("Text trong code", "Text UI phải dễ hiểu và sanitize; cấm lộ SQL, stack trace, credential hoặc copy nội bộ."),
            ("Môi trường", "Cấu hình theo profile/environment; secret luôn private và không nằm trong source/evidence."),
            ("Đồ họa", "Dùng Fiori/UI5 control và diagram chính thức; không copy HTML/CSS mock vào runtime."),
            ("Catalog", "Priority, severity, environment và classification là code-list active được CAP validate."),
        ],
    },
}

CONFIG_ITEMS = {
    "en": [
        ("Authentication", "All | srv/auth.js; srv/auth/custom-auth.js; /odata/v4/auth/", "Required; session duration from private CAP config", "DonHV", "Login/me/logout + anonymous 401 + sanitized error evidence", "Revoke sessions or roll back runtime deploy"),
        ("BugService", "All | srv/service.cds; srv/service.js; /odata/v4/bug/", "Required; no private default", "DonHV", "Metadata/compile, role and lifecycle regression", "Roll back runtime deploy"),
        ("CAP profiles", "Local/QA | package.json and private environment", "SQLite local; PostgreSQL QA; HANA direction undecided", "DonHV", "Resolve profile and database binding without printing URL", "Restore prior private profile/binding"),
        ("PostgreSQL", "Shared QA | Render database binding", "Required for shared QA; no credential in workbook", "DonHV", "Persistence/readback after restart; 21 PASSED + 6 UAT PREPARED truth", "Backup/readback then restore or migrate"),
        ("SQLite", "Local | db.sqlite generated locally", "Default local-only development store", "Each developer", "cds deploy local + focused tests", "Regenerate local database from approved seed"),
        ("AWS S3", "Shared QA | private attachment binding", "Required for live attachment binary; no key stored here", "DonHV", "Upload/download/SHA-256/reload/delete evidence", "Disable binding and use approved local fallback only outside QA"),
        ("Email provider", "Shared QA | private Brevo/SMTP configuration", "Optional delivery channel; workflow remains independent", "DonHV", "NotificationDeliveries PENDING→SENT and inbox evidence", "Disable email; preserve in-app notification/outbox rows"),
        ("Render", "Shared QA | service settings/deploy control", "Manual deploy; pre-deploy remains safe; no broad seed reload", "DonHV", "Frozen commit/deploy ID, health 200, protected 401, authenticated smoke", "Roll back to prior live deploy"),
        ("AI", "All | private AI config and srv/ai provider seam", "Disabled by default; live OpenAI NOT ACCEPTED", "DonHV", "25/25 disabled-provider/fallback/no-mutation evidence", "Keep disabled; normal workflow remains available"),
        ("Security", "All | private environment and repository gates", "Mandatory; no secret/private value in source or evidence", "All members", "Secret scan, safe error and role-boundary tests", "Revoke exposed value and rotate credential"),
        ("Migration", "Shared QA | additive migration helper", "Required only for schema change; idempotent", "DonHV", "Dry-run, backup, run twice, schema/data readback", "Transaction rollback or restore backup"),
        ("Health check", "Shared QA | /odata/v4/auth/$metadata and protected OData", "Required before acceptance", "DonHV/NhanT", "Metadata 200; anonymous 401; authenticated 200; no 5xx", "Stop acceptance and roll back deploy"),
        ("Developer dataset", "Shared QA | Users and DeveloperResponsibilities", "14 users: 12 Developers; 30 responsibilities", "DonHV", "Count/readback plus smart-assignment smoke", "Restore database backup; do not broad-seed reload"),
    ],
    "vi": [
        ("Xác thực", "Mọi môi trường | srv/auth.js; srv/auth/custom-auth.js; /odata/v4/auth/", "Bắt buộc; thời hạn phiên lấy từ cấu hình CAP riêng tư", "DonHV", "Bằng chứng đăng nhập/me/đăng xuất + ẩn danh 401 + lỗi đã làm sạch", "Thu hồi phiên hoặc quay lại bản deploy trước"),
        ("BugService", "Mọi môi trường | srv/service.cds; srv/service.js; /odata/v4/bug/", "Bắt buộc; không có giá trị riêng tư mặc định", "DonHV", "Metadata/compile, quyền theo vai trò và hồi quy vòng đời", "Quay lại bản deploy trước"),
        ("CAP profile", "Local/QA | package.json và biến môi trường riêng tư", "SQLite local; PostgreSQL QA; hướng HANA chưa quyết định", "DonHV", "Xác nhận profile và liên kết database nhưng không in URL", "Khôi phục profile/liên kết riêng tư trước đó"),
        ("PostgreSQL", "Shared QA | liên kết database Render", "Bắt buộc cho Shared QA; không ghi credential trong workbook", "DonHV", "Lưu bền/readback sau restart; sự thật 21 PASSED + 6 UAT PREPARED", "Backup/readback rồi khôi phục hoặc migrate"),
        ("SQLite", "Local | db.sqlite được tạo cục bộ", "Kho dữ liệu mặc định chỉ cho phát triển local", "Mỗi developer", "cds deploy local + test tập trung", "Tạo lại database local từ seed đã duyệt"),
        ("AWS S3", "Shared QA | liên kết tệp đính kèm riêng tư", "Bắt buộc cho binary tệp live; không lưu key tại đây", "DonHV", "Bằng chứng upload/download/SHA-256/reload/delete", "Tắt liên kết và chỉ dùng fallback local được duyệt ngoài QA"),
        ("Nhà cung cấp email", "Shared QA | cấu hình Brevo/SMTP riêng tư", "Kênh giao tùy chọn; workflow độc lập", "DonHV", "NotificationDeliveries PENDING→SENT và bằng chứng inbox", "Tắt email; giữ thông báo trong ứng dụng và outbox"),
        ("Render", "Shared QA | cấu hình service/kiểm soát deploy", "Deploy thủ công; pre-deploy an toàn; không nạp lại seed rộng", "DonHV", "Commit/deploy ID cố định, health 200, route bảo vệ 401, smoke có đăng nhập", "Quay lại deploy live trước"),
        ("AI", "Mọi môi trường | cấu hình AI riêng tư và provider seam srv/ai", "Mặc định tắt; OpenAI live CHƯA NGHIỆM THU", "DonHV", "Bằng chứng 25/25 cho provider tắt/fallback/không mutation", "Giữ trạng thái tắt; workflow thường vẫn hoạt động"),
        ("Bảo mật", "Mọi môi trường | biến riêng tư và gate repository", "Bắt buộc; không để secret/giá trị riêng tư trong source/evidence", "Tất cả thành viên", "Secret scan, lỗi an toàn và test ranh giới vai trò", "Thu hồi giá trị lộ và xoay credential"),
        ("Migration", "Shared QA | helper migration cộng thêm", "Chỉ bắt buộc khi đổi schema; idempotent", "DonHV", "Dry-run, backup, chạy hai lần, readback schema/data", "Rollback transaction hoặc khôi phục backup"),
        ("Health check", "Shared QA | /odata/v4/auth/$metadata và OData được bảo vệ", "Bắt buộc trước nghiệm thu", "DonHV/NhanT", "Metadata 200; ẩn danh 401; có đăng nhập 200; không 5xx", "Dừng nghiệm thu và quay lại deploy trước"),
        ("Bộ dữ liệu developer", "Shared QA | Users và DeveloperResponsibilities", "14 user: 12 Developer; 30 responsibility", "DonHV", "Đếm/readback và smoke smart assignment", "Khôi phục backup; không nạp lại seed rộng"),
    ],
}


def writable(ws, coordinate):
    cell = ws[coordinate]
    if not isinstance(cell, MergedCell):
        return cell
    for merged in ws.merged_cells.ranges:
        if coordinate in merged:
            return ws.cell(merged.min_row, merged.min_col)
    raise ValueError(f"No merge anchor for {ws.title}!{coordinate}")


def write(ws, coordinate, value):
    cell = writable(ws, coordinate)
    cell.value = value


def write_wrapped(ws, coordinate, value, *, vertical="top"):
    cell = writable(ws, coordinate)
    cell.value = value
    alignment = copy(cell.alignment)
    cell.alignment = Alignment(
        horizontal=alignment.horizontal,
        vertical=vertical,
        text_rotation=alignment.text_rotation,
        wrap_text=True,
        shrink_to_fit=False,
        indent=alignment.indent,
    )


def merge_row(ws, row, start_column, end_column):
    target = f"{start_column}{row}:{end_column}{row}"
    if target not in {str(item) for item in ws.merged_cells.ranges}:
        ws.merge_cells(target)


def clear_rows(ws, start, end):
    for row in ws.iter_rows(min_row=start, max_row=end):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None


def unmerge_region(ws, min_row, max_row, min_col, max_col):
    for merged in list(ws.merged_cells.ranges):
        if not (
            merged.max_row < min_row or merged.min_row > max_row
            or merged.max_col < min_col or merged.min_col > max_col
        ):
            ws.unmerge_cells(str(merged))


def copy_style(source, target):
    target._style = copy(source._style)
    target.number_format = source.number_format
    target.protection = copy(source.protection)


def write_formal_table(ws, start_row, groups, headers, rows, *, header_source, data_source, row_height=44):
    end_row = start_row + len(rows)
    min_col = min(ws[f"{start}1"].column for start, _ in groups)
    max_col = max(ws[f"{end}1"].column for _, end in groups)
    unmerge_region(ws, start_row, end_row, min_col, max_col)
    clear_rows(ws, start_row, end_row)
    header_template = ws[header_source]
    data_template = ws[data_source]
    for row_offset, values in enumerate([headers, *rows]):
        row = start_row + row_offset
        is_header = row_offset == 0
        template = header_template if is_header else data_template
        for (start, end), value in zip(groups, values):
            start_col = ws[f"{start}1"].column
            end_col = ws[f"{end}1"].column
            for column in range(start_col, end_col + 1):
                cell = ws.cell(row, column)
                copy_style(template, cell)
                cell.alignment = Alignment(
                    horizontal="center" if is_header else "left",
                    vertical="center" if is_header else "top",
                    wrap_text=True,
                    shrink_to_fit=False,
                )
            if start != end:
                ws.merge_cells(f"{start}{row}:{end}{row}")
            ws[f"{start}{row}"] = value
        ws.row_dimensions[row].height = 30 if is_header else row_height
    return end_row


def set_print_region(ws, area, *, title_rows=None):
    ws.print_area = area
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def remove_broken_defined_names(workbook):
    broken = [
        name
        for name, item in workbook.defined_names.items()
        if "#REF!" in str(item.attr_text)
    ]
    for name in broken:
        del workbook.defined_names[name]


def metadata(ws, function_id="IDTS-TECH", function_name="IDTS CAP/Fiori"):
    for row in (2, 3, 4):
        for column in ("AS", "AT", "AU", "AV", "AW", "AX", "BB", "BC", "BD"):
            try:
                write(ws, f"{column}{row}", None)
            except ValueError:
                pass
    values = {
        "F3": function_id,
        "U3": function_name,
        "G3": None,
        "V3": None,
        "AS2": " DonHV",
        "BB2": DATE.isoformat(),
        "AS3": " DonHV",
        "BD3": DATE.isoformat(),
        "AS4": " Mentor / Supervisor",
        "BD4": "Pending",
    }
    for coordinate, value in values.items():
        try:
            write(ws, coordinate, value)
        except ValueError:
            pass


def localize_visible_technical_labels(workbook):
    translations = {
        "Technical Specification": "Đặc tả kỹ thuật",
        "TECHNICAL SPECIFICATION": "ĐẶC TẢ KỸ THUẬT",
        "Introduction": "Giới thiệu",
        "Scope": "Phạm vi",
        "Assumptions": "Giả định",
        "Functional Requirements": "Yêu cầu chức năng",
        "Technical Design": "Thiết kế kỹ thuật",
        "Development Standards": "Tiêu chuẩn phát triển",
        "Screen Layout": "Bố cục màn hình",
        "Screen Definition": "Định nghĩa màn hình",
        "Message Definition": "Định nghĩa thông báo",
        "Technical Implementation": "Triển khai kỹ thuật",
        "Function ID": "Mã chức năng",
        "Function Name": "Tên chức năng",
        "Created Date": "Ngày tạo",
        "Last Update Date": "Ngày cập nhật cuối",
        "Creator": "Người tạo",
        "Reviewer": "Người review",
        "Approver": "Người phê duyệt",
        "Version": "Phiên bản",
        "Description": "Mô tả",
        "Modified date": "Ngày sửa",
        "Modified by": "Người sửa",
        "Created by:": "Người tạo:",
        "Created date:": "Ngày tạo:",
        "Updated by:": "Người cập nhật:",
        "Updated date:": "Ngày cập nhật:",
        "Function Type": "Loại chức năng",
        "Business Process": "Quy trình nghiệp vụ",
        "Data Dictionary Objects": "Đối tượng từ điển dữ liệu",
        "Function Group": "Nhóm hàm (Function Group)",
        "Function Module": "Module hàm (Function Module)",
        "Pending": "Chờ duyệt",
    }
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell, MergedCell) and isinstance(cell.value, str):
                    normalized = cell.value.strip()
                    if normalized in translations:
                        leading = cell.value[: len(cell.value) - len(cell.value.lstrip())]
                        cell.value = leading + translations[normalized]


def technical(language):
    labels = TECH_TEXT[language]
    output = OUT / f"Technical_Specification_IDTS_SAP01_{language}_v{TECH_VERSION}.xlsx"
    shutil.copy2(TEMPLATES / "Technical_Specification.xlsx", output)
    workbook = load_workbook(output)
    remove_broken_defined_names(workbook)

    cover = workbook["Cover"]
    for coordinate, value in {
        "B8": labels["title"],
        "N11": "IDTS",
        "Z11": "Issue and Defect Tracking System in SAP",
        "N12": "IDTS-TECH",
        "N13": labels["name"],
        "N14": DATE.isoformat(),
        "Z14": DATE.isoformat(),
        "AE19": "DonHV",
        "Z19": "Pending",
        "U19": "Pending",
    }.items():
        write(cover, coordinate, value)
    set_print_region(cover, "B1:BG24")

    history = workbook["Histories"]
    clear_rows(history, 3, 10)
    history_rows = {
        "en": [
            ("0.1", "Initial technical baseline", "Architecture, service and persistence scope", date(2026, 6, 21)),
            ("0.2", "Shared QA integrations", "PostgreSQL, S3, Render and email outbox", date(2026, 7, 2)),
            ("0.3", "Security and AI boundary", "Authentication, safe error and advisory AI", date(2026, 7, 24)),
            ("0.4", "Official-template remediation", "All 12 sheets and exact action inventory", date(2026, 7, 25)),
            ("0.5", labels["history"], "Trace, completeness and formal layout", DATE),
            ("0.6", "Formal table and bilingual remediation", "Structured catalogs, shared messages and Vietnamese parity", DATE),
        ],
        "vi": [
            ("0.1", "Nền kỹ thuật ban đầu", "Phạm vi kiến trúc, service và persistence", date(2026, 6, 21)),
            ("0.2", "Tích hợp Shared QA", "PostgreSQL, S3, Render và email outbox", date(2026, 7, 2)),
            ("0.3", "Ranh giới bảo mật và AI", "Xác thực, lỗi an toàn và AI tư vấn", date(2026, 7, 24)),
            ("0.4", "Khắc phục theo template chính thức", "Đủ 12 sheet và danh mục action chính xác", date(2026, 7, 25)),
            ("0.5", labels["history"], "Truy vết, độ đầy đủ và bố cục trang trọng", DATE),
            ("0.6", "Chuẩn hóa bảng và song ngữ", "Catalog có cấu trúc, message dùng chung và parity tiếng Việt", DATE),
        ],
    }[language]
    for index, (version, summary, affected, changed_date) in enumerate(history_rows, 3):
        for coordinate, value in {
            f"B{index}": index - 2,
            f"C{index}": version,
            f"D{index}": summary,
            f"E{index}": affected,
            f"F{index}": changed_date.isoformat(),
            f"G{index}": "DonHV",
        }.items():
            write(history, coordinate, value)
        history.row_dimensions[index].height = 34
    set_print_region(history, "B1:G10", title_rows="1:2")

    for name in (
        "Introduction",
        "Scope",
        "Assumptions",
        "Functional Requirements",
        "Technical Design",
        "Development Standards",
        "Screen Layout",
        "Screen Definition",
        "Message Definition",
    ):
        metadata(workbook[name])

    intro = workbook["Introduction"]
    clear_rows(intro, 15, 28)
    intro_headers = (
        ["Item", "Value", "Technical meaning"] if language == "en"
        else ["Hạng mục", "Giá trị", "Ý nghĩa kỹ thuật"]
    )
    intro_rows = [
        ("Architecture", "SAP CAP Node.js + OData V4", labels["intro"][0]),
        ("User interface", "SAP Fiori Elements / SAPUI5", labels["intro"][1]),
        ("Purpose", "Merged runtime baseline", "Describe implementation and control boundaries without changing runtime."),
    ]
    if language == "vi":
        intro_rows = [
            ("Kiến trúc", "SAP CAP Node.js + OData V4", labels["intro"][0]),
            ("Giao diện", "SAP Fiori Elements / SAPUI5", labels["intro"][1]),
            ("Mục đích", "Baseline runtime đã merge", "Mô tả cách triển khai và ranh giới kiểm soát mà không thay đổi runtime."),
        ]
    write_formal_table(intro, 15, [("C", "N"), ("O", "AA"), ("AB", "BG")], intro_headers, intro_rows, header_source="C15", data_source="C15")
    set_print_region(intro, "B1:BG18", title_rows="1:15")

    scope = workbook["Scope"]
    clear_rows(scope, 6, 35)
    scope_headers = (
        ["Scope ID", "Classification", "Content", "Reason / limitation"] if language == "en"
        else ["Mã phạm vi", "Phân loại", "Nội dung", "Lý do / giới hạn"]
    )
    scope_rows = [
        ("SCP-01", "In scope", "AuthService, BugService, draft/active writes and eleven lifecycle actions", "Core merged CAP runtime"),
        ("SCP-02", "In scope", "Comments, attachments, history, notifications and monitoring", "Supported collaboration and operational visibility"),
        ("SCP-03", "In scope", "Human-reviewed AI suggestions, decisions, explicit apply/confirm and metrics", "Advisory AI with authorization and no autonomous workflow mutation"),
        ("SCP-04", "Out of scope", "ABAP/RAP and SAP Transport Requests", "IDTS is implemented with CAP Node.js and Git/PR deployment control"),
        ("SCP-05", "Out of scope", "Final production topology and live OpenAI acceptance", "Shared QA is for mentor/demo; live OpenAI remains disabled"),
    ]
    if language == "vi":
        scope_rows = [
            ("SCP-01", "Trong phạm vi", "AuthService, BugService, ghi draft/active và mười một action vòng đời", "Runtime CAP lõi đã merge"),
            ("SCP-02", "Trong phạm vi", "Bình luận, tệp đính kèm, lịch sử, thông báo và giám sát", "Hỗ trợ cộng tác và khả năng quan sát vận hành"),
            ("SCP-03", "Trong phạm vi", "Suggestion AI có review, quyết định, apply/confirm rõ ràng và metrics", "AI tư vấn có phân quyền và không tự động mutation workflow"),
            ("SCP-04", "Ngoài phạm vi", "ABAP/RAP và SAP Transport Request", "IDTS dùng CAP Node.js và kiểm soát deploy bằng Git/PR"),
            ("SCP-05", "Ngoài phạm vi", "Topology production cuối và nghiệm thu OpenAI live", "Shared QA phục vụ mentor/demo; OpenAI live vẫn tắt"),
        ]
    write_formal_table(scope, 6, [("B", "H"), ("I", "O"), ("P", "AL"), ("AM", "BG")], scope_headers, scope_rows, header_source="B6", data_source="B6")
    set_print_region(scope, "B1:BG11", title_rows="1:6")

    assumptions = workbook["Assumptions"]
    clear_rows(assumptions, 6, 45)
    assumption_headers = (
        ["ID", "Assumption", "Environment", "Impact", "Verification"] if language == "en"
        else ["Mã", "Giả định", "Môi trường", "Ảnh hưởng", "Cách xác minh"]
    )
    assumption_rows = [
        ("ASM-01", "SQLite is used only for local development", "Local", "Local data is not Shared QA evidence", "Resolve CAP profile and run focused tests"),
        ("ASM-02", "Shared QA uses PostgreSQL on Render", "Shared QA", "Acceptance data must persist after restart", "Read back record after restart/redeploy"),
        ("ASM-03", "Schema migration is additive and idempotent", "Shared QA", "Broad seed reload is forbidden", "Dry-run, run twice and compare schema/data"),
        ("ASM-04", "S3 and Brevo credentials remain private", "Shared QA", "Evidence must not expose provider secrets", "Secret scan and sanitized integration evidence"),
        ("ASM-05", "Live OpenAI is disabled / not accepted", "All", "Mock/fallback PASS is not provider-live PASS", "Check configuration and label evidence accurately"),
        ("ASM-06", "Render deployment is manual and commit-frozen", "Shared QA", "Evidence is valid only for the recorded SHA", "Compare Git SHA with Render deploy SHA"),
    ]
    if language == "vi":
        assumption_rows = [
            ("ASM-01", "SQLite chỉ dùng cho phát triển local", "Local", "Dữ liệu local không phải evidence Shared QA", "Xác định CAP profile và chạy test tập trung"),
            ("ASM-02", "Shared QA dùng PostgreSQL trên Render", "Shared QA", "Dữ liệu nghiệm thu phải còn sau restart", "Đọc lại record sau restart/redeploy"),
            ("ASM-03", "Migration schema phải cộng thêm và idempotent", "Shared QA", "Cấm nạp lại seed diện rộng", "Dry-run, chạy hai lần và so sánh schema/data"),
            ("ASM-04", "Credential S3 và Brevo luôn riêng tư", "Shared QA", "Evidence không được lộ secret provider", "Secret scan và evidence tích hợp đã làm sạch"),
            ("ASM-05", "OpenAI live đang tắt / chưa nghiệm thu", "Mọi môi trường", "PASS mock/fallback không phải PASS provider live", "Kiểm tra cấu hình và ghi nhãn evidence chính xác"),
            ("ASM-06", "Render deploy thủ công theo commit cố định", "Shared QA", "Evidence chỉ hợp lệ cho SHA đã ghi", "So sánh Git SHA với Render deploy SHA"),
        ]
    write_formal_table(assumptions, 6, [("B", "F"), ("G", "U"), ("V", "AC"), ("AD", "AP"), ("AQ", "BG")], assumption_headers, assumption_rows, header_source="B6", data_source="B6")
    set_print_region(assumptions, "B1:BG12", title_rows="1:6")

    requirements = workbook["Functional Requirements"]
    clear_rows(requirements, 5, 70)
    requirement_headers = (
        ["Requirement ID", "Goal", "Role", "Service / API", "Processing rule", "Source component", "Data / side effect", "Evidence"]
        if language == "en" else
        ["Mã yêu cầu", "Mục tiêu", "Vai trò", "Service / API", "Quy tắc xử lý", "Thành phần nguồn", "Dữ liệu / side effect", "Evidence"]
    )
    requirement_rows = []
    for req_id, goal, role, api, rule, source, data, evidence in TECH_REQUIREMENTS:
        if language == "vi":
            translations = {
                "SRS-FR-AUTH": ("Xác thực người dùng và kiểm soát phiên bearer", "Kiểm tra password hash; lưu tokenHash; resolve user cho request được bảo vệ"),
                "SRS-FR-BUG": ("Tạo/cập nhật Bug hợp lệ qua draft và active write", "Kiểm tra field bắt buộc, catalog active và phân loại dẫn xuất trước commit"),
                "SRS-FR-ASG": ("Phân công Developer đủ điều kiện", "Kiểm tra role, Developer active và DeveloperResponsibilities"),
                "SRS-FR-LIFE": ("Kiểm soát mười một action vòng đời chính xác", "Kiểm tra role, status, reason, assignee và xác định next processor"),
                "SRS-FR-COLLAB": ("Lưu bình luận và evidence tệp đính kèm", "Phân quyền cộng tác trên Bug active và tách metadata khỏi binary storage"),
                "SRS-FR-MON": ("Trả về workload và hàng đợi theo vai trò", "Tạo KPI/filter chỉ đọc từ ownership và ngày của Bug hiện tại"),
                "SRS-FR-NOTIFY": ("Lưu thông báo và trạng thái delivery outbox", "Commit notification cùng workflow; gửi ngoài transaction với retry/lock"),
                "SRS-FR-AI": ("Cung cấp hỗ trợ AI có human review", "Dùng dữ liệu allowlist, lưu review state và yêu cầu apply/confirm rõ ràng"),
            }
            goal, rule = translations[req_id]
        requirement_rows.append((req_id, goal, role, api, rule, source, data, evidence))
    write_formal_table(
        requirements, 5,
        [("B", "F"), ("G", "Q"), ("R", "V"), ("W", "AB"), ("AC", "AN"), ("AO", "AV"), ("AW", "BC"), ("BD", "BG")],
        requirement_headers, requirement_rows, header_source="B5", data_source="B5", row_height=58,
    )
    set_print_region(requirements, "B1:BG13", title_rows="1:5")

    design = workbook["Technical Design"]
    clear_rows(design, 6, 129)
    design_tables = [
        (
            "Component Map" if language == "en" else "Bản đồ thành phần",
            ["Component", "Responsibility", "Public boundary", "Source"] if language == "en" else ["Thành phần", "Trách nhiệm", "Ranh giới công khai", "Source"],
            [
                ("AuthService", "Authentication and bearer session lifecycle" if language == "en" else "Xác thực và vòng đời phiên bearer", "/odata/v4/auth/", "srv/auth.js; srv/auth/custom-auth.js"),
                ("BugService", "Bug workflow, collaboration, monitoring and AI actions" if language == "en" else "Workflow Bug, cộng tác, giám sát và action AI", "/odata/v4/bug/", "srv/service.cds; srv/service.js"),
                ("Fiori application", "Role-aware user experience" if language == "en" else "Trải nghiệm người dùng theo vai trò", "UI routes and OData bindings" if language == "en" else "Route UI và OData binding", "app/bug-management-ui"),
            ],
        ),
        (
            "Entity Relationship" if language == "en" else "Quan hệ entity",
            ["Parent", "Relationship", "Child / target", "Purpose"] if language == "en" else ["Entity cha", "Quan hệ", "Entity con / đích", "Mục đích"],
            [
                ("Bugs", "composition", "Comments; Attachments; HistoryEvents; Notifications", "Lifecycle-owned collaboration and audit" if language == "en" else "Cộng tác và audit thuộc vòng đời Bug"),
                ("Users", "association", "AuthSessions; Bugs reporter/assignee/nextProcessor", "Identity and ownership" if language == "en" else "Danh tính và ownership"),
                ("Notifications", "composition", "NotificationDeliveries", "Outbox delivery tracking" if language == "en" else "Theo dõi delivery outbox"),
                ("AiSuggestions", "association", "Bugs; DuplicateLinks", "Human-reviewed AI outcomes" if language == "en" else "Kết quả AI có human review"),
            ],
        ),
        (
            "Module Dependency" if language == "en" else "Phụ thuộc module",
            ["Caller", "Callee", "Contract", "Reason"] if language == "en" else ["Bên gọi", "Bên được gọi", "Contract", "Lý do"],
            [
                ("srv/service.js", "srv/bug-service/*", "Registered CAP handlers" if language == "en" else "CAP handler đã đăng ký", "Keep service bootstrap small" if language == "en" else "Giữ service bootstrap gọn"),
                ("Bug actions", "history/notification helpers", "Transaction-scoped side effects" if language == "en" else "Side effect trong transaction", "Atomic workflow audit" if language == "en" else "Audit workflow atomic"),
                ("Fiori controllers", "OData V4 model", "Metadata/actions/entities", "No direct database access" if language == "en" else "Không truy cập database trực tiếp"),
            ],
        ),
        (
            "Transaction Boundary" if language == "en" else "Ranh giới transaction",
            ["Operation", "In transaction", "Outside transaction", "Failure behavior"] if language == "en" else ["Thao tác", "Trong transaction", "Ngoài transaction", "Hành vi khi lỗi"],
            [
                ("Bug lifecycle", "Bug, next processor, history, notification", "Email provider send" if language == "en" else "Gửi qua email provider", "Rollback business data when audit write fails" if language == "en" else "Rollback dữ liệu nghiệp vụ khi ghi audit lỗi"),
                ("Attachment", "PostgreSQL metadata", "S3 binary provider", "Sanitize provider error; keep Bug" if language == "en" else "Làm sạch lỗi provider; giữ Bug"),
                ("AI review/apply", "AiSuggestions and explicit authorized mutation" if language == "en" else "AiSuggestions và mutation rõ ràng đã cấp quyền", "Optional provider request" if language == "en" else "Request provider tùy chọn", "No autonomous workflow mutation" if language == "en" else "Không tự động mutation workflow"),
            ],
        ),
        (
            "External Integration" if language == "en" else "Tích hợp bên ngoài",
            ["Provider", "Purpose", "Private configuration", "Verification"] if language == "en" else ["Provider", "Mục đích", "Cấu hình riêng tư", "Cách xác minh"],
            [
                ("AWS S3", "Attachment binary storage" if language == "en" else "Lưu binary tệp đính kèm", "Environment binding" if language == "en" else "Binding môi trường", "Upload/download/hash/reload/delete"),
                ("Brevo", "Email delivery" if language == "en" else "Gửi email", "Environment binding" if language == "en" else "Binding môi trường", "PENDING → SENT and inbox evidence" if language == "en" else "PENDING → SENT và evidence inbox"),
                ("OpenAI", "Optional AI provider" if language == "en" else "Provider AI tùy chọn", "Disabled" if language == "en" else "Đã tắt", "NOT ACCEPTED; fallback/no-mutation only" if language == "en" else "CHƯA NGHIỆM THU; chỉ fallback/no-mutation"),
                ("Render PostgreSQL", "Shared QA persistence" if language == "en" else "Persistence Shared QA", "Database binding" if language == "en" else "Binding database", "Readback after restart/redeploy" if language == "en" else "Readback sau restart/redeploy"),
            ],
        ),
    ]
    design_groups = [("B", "J"), ("K", "V"), ("W", "AM"), ("AN", "BG")]
    design_row = 6
    for title, headers, rows in design_tables:
        merge_row(design, design_row, "B", "BG")
        write_wrapped(design, f"B{design_row}", title, vertical="center")
        design.row_dimensions[design_row].height = 26
        design_row = write_formal_table(design, design_row + 1, design_groups, headers, rows, header_source="B5", data_source="B6", row_height=46) + 2
    design._images = []
    architecture = XLImage(ROOT / "docs" / "diagrams" / "rendered" / "png" / "02-cap-fiori-architecture.png")
    architecture.width = 720
    architecture.height = 520
    design.add_image(architecture, f"B{design_row}")
    set_print_region(design, f"B1:BG{design_row + 26}", title_rows="1:5")

    standards = workbook["Development Standards"]
    clear_rows(standards, 5, 90)
    standard_headers = (
        ["Standard ID", "Area", "Rule", "Applies to", "Verification"] if language == "en"
        else ["Mã quy tắc", "Lĩnh vực", "Quy tắc", "Phạm vi áp dụng", "Cách xác minh"]
    )
    standard_rows = [
        (f"STD-{index:02d}", heading, detail, "app/; srv/; db/; scripts/", "Review + lint + focused test" if language == "en" else "Review + lint + test tập trung")
        for index, (heading, detail) in enumerate(labels["standards"], 1)
    ]
    write_formal_table(
        standards, 5,
        [("B", "F"), ("G", "N"), ("O", "AM"), ("AN", "AT"), ("AU", "BG")],
        standard_headers, standard_rows, header_source="B5", data_source="B6", row_height=48,
    )
    set_print_region(standards, f"B1:BG{5 + len(standard_rows)}", title_rows="1:5")

    layout = workbook["Screen Layout"]
    clear_rows(layout, 5, 35)
    layout_headers = (
        ["Screen ID", "Route / entry", "Page / view", "Controller / extension", "OData binding", "Major controls", "Role"]
        if language == "en" else
        ["Mã màn hình", "Route / entry", "Trang / view", "Controller / extension", "OData binding", "Control chính", "Vai trò"]
    )
    layout_rows = []
    for screen_id, name_en, name_vi, page_type, route, controller, binding, role, controls, navigation in SCREENS:
        if language == "vi":
            controls = {
                "Email, password, safe message": "Email, password, thông báo an toàn", "Name, email, role, Sign Out": "Tên, email, role, Đăng xuất",
                "KPI cards, queues, workload": "KPI card, hàng đợi, workload", "Filters, table, Create": "Bộ lọc, bảng, Tạo",
                "Summary, classification, assignment, lifecycle": "Tóm tắt, phân loại, phân công, vòng đời",
                "Thread and Add Comment": "Luồng bình luận và Thêm bình luận", "Upload, download, delete": "Tải lên, tải xuống, xóa",
                "Timeline and Show More": "Timeline và Xem thêm", "Read state and delivery status": "Trạng thái đọc và delivery",
                "Search, workload, responsibility, explanation": "Tìm kiếm, workload, responsibility, giải thích",
                "Accept, reject, ignore, apply": "Chấp nhận, từ chối, bỏ qua, áp dụng", "Candidates, review, confirm": "Ứng viên, review, xác nhận",
                "Grounded summary and review": "Tóm tắt có căn cứ và review",
            }.get(controls, controls)
        layout_rows.append((screen_id, route, name_en if language == "en" else name_vi, controller, binding, controls, role))
    write_formal_table(
        layout, 5,
        [("B", "F"), ("G", "M"), ("N", "U"), ("V", "AF"), ("AG", "AN"), ("AO", "AX"), ("AY", "BG")],
        layout_headers, layout_rows, header_source="B5", data_source="B6", row_height=50,
    )
    set_print_region(layout, f"B1:BG{5 + len(layout_rows)}", title_rows="1:5")

    definition = workbook["Screen Definition"]
    clear_rows(definition, 9, 41)
    definition_headers = (
        ["Screen ID", "UI element", "Annotation / manifest binding", "Handler", "OData operation", "Authorization", "Failure behavior"]
        if language == "en" else
        ["Mã màn hình", "Thành phần UI", "Annotation / manifest binding", "Handler", "Thao tác OData", "Phân quyền", "Hành vi khi lỗi"]
    )
    definition_rows = []
    for screen_id, name_en, name_vi, page_type, route, controller, binding, role, controls, navigation in SCREENS:
        failure = "Safe message; no unauthorized mutation" if language == "en" else "Thông báo an toàn; không mutation trái quyền"
        definition_rows.append((screen_id, name_en if language == "en" else name_vi, route, controller, binding, role, failure))
    write_formal_table(
        definition, 9,
        [("B", "F"), ("G", "N"), ("O", "W"), ("X", "AG"), ("AH", "AO"), ("AP", "AW"), ("AX", "BG")],
        definition_headers, definition_rows, header_source="B10", data_source="B11", row_height=50,
    )
    set_print_region(definition, f"B1:BG{9 + len(definition_rows)}", title_rows="1:9")

    messages = workbook["Message Definition"]
    clear_rows(messages, 5, 31)
    message_headers = (
        ["Message ID", "HTTP / status", "Exact source", "Target", "Rollback", "Sanitized logging", "Frontend handling / evidence"]
        if language == "en" else
        ["Mã thông báo", "HTTP / trạng thái", "Source chính xác", "Target", "Rollback", "Log đã làm sạch", "Xử lý frontend / evidence"]
    )
    message_rows = [
        (item["id"], item["http"], item["source"], item["target"], item["rollback"][language], item["log"][language], f"{item['ui'][language]}; {item['evidence']}")
        for item in MESSAGES
    ]
    write_formal_table(
        messages, 5,
        [("B", "G"), ("H", "L"), ("M", "V"), ("W", "AB"), ("AC", "AL"), ("AM", "AV"), ("AW", "BG")],
        message_headers, message_rows, header_source="B5", data_source="B6", row_height=56,
    )
    set_print_region(messages, f"B1:BG{5 + len(message_rows)}", title_rows="1:5")

    implementation = workbook["Technical Implementation"]
    clear_rows(implementation, 5, 35)
    write(implementation, "B3", "Technical Implementation" if language == "en" else "Triển khai kỹ thuật")
    write(implementation, "B4", "IDTS-TECH")
    write(implementation, "L4", labels["name"])
    implementation_headers = (
        ["Flow ID", "UI trigger", "HTTP / OData", "Service", "Handler / helper", "Transaction", "Storage / provider effect", "Response / evidence"]
        if language == "en" else
        ["Mã luồng", "Trigger UI", "HTTP / OData", "Service", "Handler / helper", "Transaction", "Ảnh hưởng storage / provider", "Response / evidence"]
    )
    implementation_rows = []
    for flow_id, trigger, request, service, handler, transaction, effect, response, evidence in TECH_FLOWS:
        if language == "vi":
            trigger = {
                "FLOW-AUTH": "Gửi biểu mẫu đăng nhập", "FLOW-DRAFT-CREATE": "Tạo và chỉnh draft", "FLOW-ACTIVE-EDIT": "Chỉnh Bug active và lưu",
                "FLOW-ASSIGN": "Xác nhận Developer", "FLOW-LIFECYCLE": "Gọi action vòng đời", "FLOW-COLLAB": "Thêm bình luận hoặc tệp",
                "FLOW-MON": "Mở dashboard/bộ lọc", "FLOW-EMAIL": "Xử lý dòng outbox", "FLOW-AI": "Yêu cầu/review/apply hỗ trợ AI",
            }[flow_id]
            response = {
                "FLOW-AUTH": "Token chỉ trả một lần hoặc 401 an toàn", "FLOW-DRAFT-CREATE": "Bug active hoặc lỗi validation an toàn",
                "FLOW-ACTIVE-EDIT": "Bug active đã cập nhật", "FLOW-ASSIGN": "Assigned hoặc 400/403 an toàn",
                "FLOW-LIFECYCLE": "Transition chính xác được cho phép", "FLOW-COLLAB": "Evidence được lưu hoặc lỗi an toàn",
                "FLOW-MON": "KPI theo vai trò", "FLOW-EMAIL": "SENT/FAILED/SKIPPED", "FLOW-AI": "Kết quả review hoặc fallback an toàn",
            }[flow_id]
            transaction = {
                "FLOW-AUTH": "Transaction của request", "FLOW-DRAFT-CREATE": "Transaction request CAP", "FLOW-ACTIVE-EDIT": "Transaction request CAP",
                "FLOW-ASSIGN": "Transaction request CAP", "FLOW-LIFECYCLE": "Transaction request CAP",
                "FLOW-COLLAB": "Ranh giới request CAP + provider S3", "FLOW-MON": "Request chỉ đọc",
                "FLOW-EMAIL": "Transaction worker sau workflow commit", "FLOW-AI": "Transaction của action",
            }[flow_id]
            effect = {
                "FLOW-AUTH": "Đọc Users; thêm AuthSessions", "FLOW-DRAFT-CREATE": "Thêm Bugs; lịch sử/thông báo",
                "FLOW-ACTIVE-EDIT": "Cập nhật Bugs; lịch sử/thông báo", "FLOW-ASSIGN": "Bug/next processor/lịch sử/thông báo",
                "FLOW-LIFECYCLE": "Bug/trạng thái/lịch sử/thông báo", "FLOW-COLLAB": "Bộ nhớ tạm client; metadata PostgreSQL; binary S3",
                "FLOW-MON": "Không mutation", "FLOW-EMAIL": "Trạng thái/retry NotificationDeliveries",
                "FLOW-AI": "AiSuggestions.operationStatus/latencyMs; tùy chọn mutation Bug/DuplicateLink rõ ràng",
            }[flow_id]
        implementation_rows.append((flow_id, trigger, request, service, handler, transaction, effect, f"{response}; {evidence}"))
    implementation_end = write_formal_table(
        implementation, 5,
        [("B", "D"), ("E", "G"), ("H", "J"), ("K", "M"), ("N", "P"), ("Q", "S"), ("T", "V"), ("W", "Y")],
        implementation_headers, implementation_rows, header_source="B5", data_source="B6", row_height=66,
    )
    for row in range(6, implementation_end + 1):
        for column in ("B", "E", "H", "K", "N", "Q", "T", "W"):
            cell = implementation[f"{column}{row}"]
            font = copy(cell.font)
            font.name = "Times New Roman"
            font.sz = 12
            cell.font = font
    set_print_region(implementation, f"B1:Y{implementation_end}", title_rows="1:5")

    if language == "vi":
        localize_visible_technical_labels(workbook)
    workbook.properties.title = f"IDTS SAP490 Technical Specification {language.upper()} v{TECH_VERSION}"
    workbook.properties.subject = "Official-template technical baseline for CAP/Fiori"
    workbook.properties.creator = "IDTS SAP01 Team"
    workbook.save(output)
    return output


def configuration(language):
    output = OUT / f"Configuration_Note_IDTS_SAP01_{language}_v{CONFIG_VERSION}.xlsx"
    shutil.copy2(TEMPLATES / "Configuration_Note.xlsx", output)
    workbook = load_workbook(output)
    remove_broken_defined_names(workbook)

    cover = workbook["Cover"]
    write(cover, "I19", "IDTS-SAP01")
    write(cover, "I20", CONFIG_VERSION)
    write(cover, "I21", DATE.isoformat())

    changes = workbook["Record of change"]
    clear_rows(changes, 4, 8)
    history_rows = {
        "en": [
            (1, "2026-06-21", "0.1", "Initial configuration baseline", "Document local CAP/SQLite setup", "Mentor / Supervisor", "Pending"),
            (2, "2026-07-02", "0.2", "Shared QA configuration", "Add Render, PostgreSQL, S3 and Brevo controls", "Mentor / Supervisor", "Pending"),
            (3, "2026-07-24", "0.3", "Security and AI controls", "Clarify private configuration and disabled live OpenAI", "Mentor / Supervisor", "Pending"),
            (4, "2026-07-25", "0.4", "Official-template remediation", "Align five official sheets with deployed CAP/Fiori baseline", "Mentor / Supervisor", "Pending"),
            (5, DATE.isoformat(), CONFIG_VERSION, "Runtime trace and formal-layout correction", "Add exact location, verification, rollback and evidence controls", "Mentor / Supervisor", "Pending"),
        ],
        "vi": [
            (1, "2026-06-21", "0.1", "Nền cấu hình ban đầu", "Ghi nhận CAP/SQLite local", "Người hướng dẫn", "Chờ đánh giá"),
            (2, "2026-07-02", "0.2", "Cấu hình Shared QA", "Bổ sung kiểm soát Render, PostgreSQL, S3 và Brevo", "Người hướng dẫn", "Chờ đánh giá"),
            (3, "2026-07-24", "0.3", "Kiểm soát bảo mật và AI", "Làm rõ cấu hình riêng tư và OpenAI live đang tắt", "Người hướng dẫn", "Chờ đánh giá"),
            (4, "2026-07-25", "0.4", "Khắc phục theo template chính thức", "Đồng bộ đủ năm sheet với nền CAP/Fiori đã triển khai", "Người hướng dẫn", "Chờ đánh giá"),
            (5, DATE.isoformat(), CONFIG_VERSION, "Sửa truy vết runtime và bố cục trang trọng", "Bổ sung vị trí, cách kiểm tra, rollback và bằng chứng chính xác", "Người hướng dẫn", "Chờ đánh giá"),
        ],
    }[language]
    for row, record in enumerate(history_rows, 4):
        for column, value in enumerate(record, 1):
            coordinate = changes.cell(row, column).coordinate
            write_wrapped(changes, coordinate, value, vertical="center")
        changes.row_dimensions[row].height = 44

    checklist = workbook["Checklist"]
    clear_rows(checklist, 4, 40)
    for row, (item, location, required_default, owner, verify, rollback) in enumerate(CONFIG_ITEMS[language], 4):
        values = ("IDTS", row - 3, item, location, required_default, f"{owner} | Verify: {verify}", f"Rollback/evidence: {rollback}")
        for coordinate, value in zip(
            (f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}"),
            values,
        ):
            write_wrapped(checklist, coordinate, value, vertical="top")
        checklist.row_dimensions[row].height = 64
    common_note = (
        "Security note: this workbook documents keys and locations only. It never stores credentials, tokens, private endpoints, database URLs or personal data."
        if language == "en"
        else "Lưu ý bảo mật: workbook này chỉ mô tả tên cấu hình và vị trí. Không lưu credential, token, endpoint riêng tư, URL database hoặc dữ liệu cá nhân."
    )
    write_wrapped(checklist, "D20", common_note, vertical="top")
    checklist.row_dimensions[20].height = 44

    notes = {
        "en": {
            "4": "N/A — Not applicable to SAP CAP/Fiori implementation. IDTS does not use classic SAP customizing transactions or T-codes. Equivalent runtime configuration uses CAP profiles and private environment variables.",
            "5": "N/A — This is not an SAP Transport Request. IDTS uses Git branches, reviewed pull requests, additive database migration and controlled Render deployment.",
        },
        "vi": {
            "4": "N/A — Không áp dụng cho triển khai SAP CAP/Fiori. IDTS không dùng transaction customizing hoặc T-code SAP cổ điển. Cấu hình tương đương dùng CAP profile và biến môi trường private.",
            "5": "N/A — Đây không phải SAP Transport Request. IDTS dùng Git branch, pull request được review, migration database cộng thêm và deploy Render có kiểm soát.",
        },
    }[language]
    for sheet_name in ("4", "5"):
        sheet = workbook[sheet_name]
        clear_rows(sheet, 4, 20)
        write(sheet, "A4", "IDTS CAP/Fiori configuration control" if language == "en" else "Kiểm soát cấu hình IDTS CAP/Fiori")
        write(
            sheet,
            "C5",
            "No credentials, tokens, private endpoint or personal data are stored in this artifact."
            if language == "en"
            else "Tài liệu này không lưu credential, token, endpoint riêng tư hoặc dữ liệu cá nhân.",
        )
        write(sheet, "C6", notes[sheet_name])
        sheet.row_dimensions[6].height = 72

    workbook.properties.title = f"IDTS SAP490 Configuration Note {language.upper()} v{CONFIG_VERSION}"
    workbook.properties.subject = "Official-template secret-free CAP/Fiori configuration baseline"
    workbook.properties.creator = "IDTS SAP01 Team"
    workbook.save(output)
    return output


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    for language in ("en", "vi"):
        print(technical(language).relative_to(ROOT))
        print(configuration(language).relative_to(ROOT))


if __name__ == "__main__":
    main()
