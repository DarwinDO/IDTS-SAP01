"""Regression checks that current SAP490 workbooks preserve official templates."""

from __future__ import annotations

import unittest
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_DIR = ROOT / "docs" / "sap490" / "templates" / "Deliverable_template"
REPORT_TEMPLATE = (
    ROOT / "docs" / "sap490" / "templates" / "2_SAP490_Test Report Template (1).xlsx"
)
OUTPUT_DIR = ROOT / "docs" / "sap490" / "generated"

ARTIFACTS = {
    "scenario": {
        "template": TEMPLATE_DIR / "Test_Scenario.xlsx",
        "outputs": [
            OUTPUT_DIR / "Test_Scenario_IDTS_SAP01_en_v0.3.xlsx",
        ],
    },
    "unit": {
        "template": TEMPLATE_DIR / "Unit_Test.xlsx",
        "outputs": [
            OUTPUT_DIR / "Unit_Test_IDTS_SAP01_en_v0.4.xlsx",
        ],
    },
    "functional": {
        "template": TEMPLATE_DIR / "Functional_Test.xlsx",
        "outputs": [
            OUTPUT_DIR / "Functional_Test_IDTS_SAP01_en_v0.3.xlsx",
        ],
    },
    "report": {
        "template": REPORT_TEMPLATE,
        "outputs": [
            OUTPUT_DIR / "Test_Report_IDTS_SAP01_en_v0.4.xlsx",
        ],
    },
    "uat": {
        "template": TEMPLATE_DIR / "UAT.xlsx",
        "outputs": [
            OUTPUT_DIR / "UAT_IDTS_SAP01_en_prepared_v0.2.xlsx",
        ],
    },
    "defect": {
        "template": TEMPLATE_DIR / "Test_And_Fix_Bug.xlsx",
        "outputs": [
            OUTPUT_DIR / "Test_And_Fix_Bug_IDTS_SAP01_en_v0.5.xlsx",
        ],
        # The original task explicitly requires removing unused Issue 2 / Issue 4 remnants.
        "allowed_sheets": ["Fix and bugs"],
    },
}

FIT_TO_ONE_PAGE_WIDE = {
    ("scenario", "Test Cases"),
    ("unit", "UT"),
    ("functional", "Test Cases"),
    ("functional", "Test Result"),
    ("functional", "Test Data Description"),
    ("uat", "Test Scenario"),
    ("uat", "Test Cases"),
}


def merged_ranges(workbook):
    return {
        sheet.title: {str(cell_range) for cell_range in sheet.merged_cells.ranges}
        for sheet in workbook.worksheets
    }


def structural_merges(kind, sheet):
    limits = {
        ("scenario", "Test Scenario"): 2,
        ("scenario", "Test Cases"): 7,
        ("unit", "UT"): 7,
        ("functional", "Test Cases"): 7,
        ("functional", "Test Result"): 6,
        ("functional", "Test Data Description"): 4,
        ("report", "Cover"): 17,
        ("report", "Test Cases"): 8,
        ("report", "Test Statistics"): 10,
        ("report", "Feature 1"): 10,
        ("report", "Feature 2"): 10,
        ("uat", "Test Scenario"): 3,
        ("uat", "Test Cases"): 7,
        ("uat", "Test Result"): 4,
        ("defect", "Fix and bugs"): 1,
    }
    limit = limits.get((kind, sheet.title))
    if limit is None:
        return {str(cell_range) for cell_range in sheet.merged_cells.ranges}
    return {
        str(cell_range)
        for cell_range in sheet.merged_cells.ranges
        if cell_range.min_row <= limit
    }


class TemplateFidelityTest(unittest.TestCase):
    def test_current_workbooks_retain_template_sheet_contract_and_merges(self):
        failures = []
        for kind, contract in ARTIFACTS.items():
            template = load_workbook(contract["template"], data_only=False)
            expected_sheets = contract.get("allowed_sheets", template.sheetnames)
            for output_path in contract["outputs"]:
                current = load_workbook(output_path, data_only=False)
                if current.sheetnames != expected_sheets:
                    failures.append(
                        f"{output_path.name}: sheets {current.sheetnames!r} != "
                        f"{expected_sheets!r}"
                    )
                    continue

                for sheet_name in expected_sheets:
                    expected_merges = structural_merges(kind, template[sheet_name])
                    actual_merges = {
                        str(cell_range)
                        for cell_range in current[sheet_name].merged_cells.ranges
                    }
                    if not expected_merges.issubset(actual_merges):
                        missing_merges = sorted(expected_merges - actual_merges)
                        failures.append(
                            f"{output_path.name}/{sheet_name}: "
                            f"missing template merges {missing_merges[:5]!r}"
                        )

                expected_states = {
                    ws.title: ws.sheet_state
                    for ws in template.worksheets
                    if ws.title in expected_sheets
                }
                actual_states = {
                    ws.title: ws.sheet_state for ws in current.worksheets
                }
                if actual_states != expected_states:
                    failures.append(
                        f"{output_path.name}: sheet states {actual_states!r} != "
                        f"{expected_states!r}"
                    )

                expected_images = {
                    ws.title: len(ws._images)
                    for ws in template.worksheets
                    if ws.title in expected_sheets
                }
                actual_images = {
                    ws.title: len(ws._images) for ws in current.worksheets
                }
                if actual_images != expected_images:
                    failures.append(
                        f"{output_path.name}: images {actual_images!r} != "
                        f"{expected_images!r}"
                    )

                for sheet in current.worksheets:
                    setup = sheet.page_setup
                    fit_to_page = sheet.sheet_properties.pageSetUpPr.fitToPage
                    if fit_to_page and setup.fitToWidth is not None and setup.scale is not None:
                        failures.append(
                            f"{output_path.name}/{sheet.title}: scale={setup.scale} "
                            "overrides fit-to-width print layout"
                        )
                    if (
                        (kind, sheet.title) in FIT_TO_ONE_PAGE_WIDE
                        and setup.fitToWidth != 1
                    ):
                        failures.append(
                            f"{output_path.name}/{sheet.title}: fitToWidth="
                            f"{setup.fitToWidth}, expected 1 to prevent split fields"
                        )

        self.assertFalse(failures, "\n" + "\n".join(failures))


if __name__ == "__main__":
    unittest.main(verbosity=2)
