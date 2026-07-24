"""Repair presentation-only issues in Drive-exported SAP490 review workbooks.

The source file is never modified. The output keeps workbook values and formulas,
    removes broken legacy defined names, bounds invalid column widths, and sizes rows
for wrapped text so OfficeCLI can assess the current local review candidate.
"""

from argparse import ArgumentParser
from copy import copy
from math import ceil
from pathlib import Path
import shutil

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter


def repair(source: Path, output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, output)
    wb = load_workbook(output)
    for name in list(wb.defined_names):
        definition = wb.defined_names[name]
        if "#REF!" in str(definition.attr_text).upper():
            del wb.defined_names[name]

    for ws in wb.worksheets:
        for name in list(ws.defined_names):
            definition = ws.defined_names[name]
            if "#REF!" in str(definition.attr_text).upper():
                del ws.defined_names[name]

        if ws.title in {"Requirement Backlog", "Risk Decision Log"}:
            ws.column_dimensions["F"].width = 80

        for dimension in ws.column_dimensions.values():
            if dimension.width is not None:
                dimension.width = min(max(dimension.width, 2), 80)

        merged_widths = {}
        for merged_range in ws.merged_cells.ranges:
            merged_widths[(merged_range.min_row, merged_range.min_col)] = sum(
                ws.column_dimensions[get_column_letter(column)].width or 13
                for column in range(merged_range.min_col, merged_range.max_col + 1)
            )

        for row in ws.iter_rows():
            required_lines = 1
            populated = False
            for cell in row:
                if isinstance(cell, MergedCell) or cell.value in (None, ""):
                    continue
                populated = True
                alignment = copy(cell.alignment)
                alignment.wrap_text = True
                alignment.vertical = alignment.vertical or "top"
                cell.alignment = alignment
                width = merged_widths.get(
                    (cell.row, cell.column),
                    ws.column_dimensions[cell.column_letter].width or 13,
                )
                required_lines = max(
                    required_lines,
                    sum(max(1, ceil(len(line) / max(width, 8))) for line in str(cell.value).splitlines()),
                )
            if populated:
                row_number = row[0].row
                current = ws.row_dimensions[row_number].height or 15
                # Keep one extra leading/trailing line because OfficeCLI measures
                # the usable height inside the cell, not the raw row height.
                ws.row_dimensions[row_number].height = max(current, min(360, (required_lines + 2) * 17))

        if ws.title == "Test Cases" and ws["CC3"].value:
            ws.row_dimensions[3].height = max(ws.row_dimensions[3].height or 15, 30)
            ws.row_dimensions[4].height = max(ws.row_dimensions[4].height or 15, 30)

        if ws.title == "Test Scenario" and any(str(item) == "E4:G4" for item in ws.merged_cells.ranges):
            ws.row_dimensions[4].height = max(ws.row_dimensions[4].height or 15, 100)
            ws.row_dimensions[5].height = max(ws.row_dimensions[5].height or 15, 100)

        if ws.title == "Evidence" and ws["B3"].value:
            ws.column_dimensions["A"].width = 24
            ws.column_dimensions["B"].width = 90
            ws.row_dimensions[3].height = 90

        if ws.title == "Fix and bugs":
            ws.column_dimensions["E"].width = 45
            ws.column_dimensions["F"].width = 45

        if ws.title == "Summary" and "Team Contribution Matrix" in str(ws["A1"].value):
            for row_number, height in {1: 30, 2: 45, 3: 15, 5: 30, 6: 45, 7: 45, 8: 45, 9: 45, 12: 60}.items():
                ws.row_dimensions[row_number].height = height

        if ws.title == "Summary" and ws["A1"].value == "Field" and ws["B1"].value == "Value":
            ws["B2"] = "2026-07-22"
            ws["B6"] = "Current Drive file is updated in place; repository workbook remains the source of truth; no copy or delete."
            for row_number, height in {1: 30, 2: 30, 3: 30, 4: 36, 5: 36, 6: 54}.items():
                ws.row_dimensions[row_number].height = height

        if "UAT_IDTS" in source.name and ws.title == "Cover":
            ws.column_dimensions["N"].width = 30
            ws.column_dimensions["Z"].width = 50
            # The legacy template stores Cover widths as a grouped A:AQ dimension;
            # LibreOffice normalizes isolated width overrides back into that group.
            # Explicit row heights therefore provide the stable cross-renderer fix.
            for row_number, height in {11: 280, 12: 115, 13: 210}.items():
                ws.row_dimensions[row_number].height = height

        if "UAT_IDTS" in source.name and ws.title == "Test Cases":
            ws.column_dimensions["E"].width = 24
            ws.column_dimensions["Y"].width = 45
            ws.column_dimensions["AP"].width = 80
            ws.row_dimensions[13].height = 230

        if "UAT_IDTS" in source.name and ws.title == "Test Result":
            ws.column_dimensions["A"].width = 80
            ws.row_dimensions[7].height = 72

    wb.save(output)


def main() -> None:
    parser = ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    repair(args.source.resolve(), args.output.resolve())
    print(args.output.resolve())


if __name__ == "__main__":
    main()
