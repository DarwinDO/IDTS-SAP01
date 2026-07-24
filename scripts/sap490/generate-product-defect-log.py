"""Generate SAP490 Test and Fix Bug workbooks from confirmed product defects only."""

from copy import copy
from pathlib import Path
import shutil

from openpyxl import load_workbook
from openpyxl.styles import Alignment


ROOT = Path(__file__).resolve().parents[2]
TEMPLATE = ROOT / "docs" / "sap490" / "templates" / "Deliverable_template" / "Test_And_Fix_Bug.xlsx"
OUT = ROOT / "docs" / "sap490" / "generated"
VERSION = "0.5"

DEFECTS = [
    ("IDTS-13", "Object Page stale after lifecycle action", "Fiori side effects targeted properties instead of the bound entity, so action changes did not fully refresh the Object Page.", "The active bug, status, ownership, and related sections refresh after a successful lifecycle action.", "Changed action side effects to `TargetEntities: [$self]`; CAP compile and browser UAT passed."),
    ("IDTS-19", "History Timeline fails when text exists", "A boolean XML expression used `${...}` inside the binding expression and caused a UI5 FormatException when timeline text existed.", "History renders readable timeline rows for events with text.", "Replaced the invalid expression pattern and retested through browser UAT."),
    ("IDTS-32", "Invalid priority and severity values accepted", "Free-text invalid catalog codes could be activated and bypass master-data validity.", "Only active catalog values can be persisted by authorized users.", "Added CAP target/assert and active-code validation; focused regression passed."),
    ("IDTS-35", "Blank app after successful login", "The OData metadata request ran before the XHR bearer-token interceptor was installed.", "After login, metadata loads with bearer authentication and the bug list renders.", "Moved guard/interceptor before UI5 bootstrap and fixed the async argument closure; UI5 build and regression passed."),
    ("IDTS-41", "Developer can start unauthorized root draft", "Create permission was not enforced at root draft `NEW`.", "Only Tester and PM can create a bug draft.", "Added backend authorization before draft creation; programmatic validation passed."),
    ("IDTS-49", "Draft activation rejects missing reporter", "Reporter was derived only during active CREATE, after draft SAVE validation had already run.", "An authorized user can activate a draft without client-supplied reporter data.", "Initialized reporter during draft NEW/SAVE before required-field validation; focused regression passed."),
    ("IDTS-52", "UI5 login page renders blank", "Unstable UI5 enum access attempted to read undefined `Error` values.", "Login controls render and display safe validation feedback.", "Used stable UI5 string enum values; browser smoke passed."),
    ("IDTS-53", "Profile overlay intercepts Sign Out", "The fixed profile host covered the sign-out control.", "Users can open the profile popover and sign out reliably.", "Adjusted pointer-event handling for host and button; browser smoke passed."),
    ("IDTS-55", "Comments and attachments Object Page interaction defects", "Custom fragments lost binding context, unsupported relative refresh left comments stale, and attachment child context was treated as a bug root.", "Comments enable/post/reload immediately and attachment removal operates on the correct bug context.", "Added public context propagation, supported refresh flow, and strict root-context detection; interaction smoke passed."),
    ("IDTS-56", "Smart Assign treats hidden component category as missing", "The UI relied on an internal field not always selected on the Object Page.", "Smart Assign derives/uses valid classification context and keeps manual assignment authority.", "Corrected context handling and added focused UI/browser tests."),
    ("IDTS-58", "Sprint 4 profile/comment/attachment UX defects", "Dashboard profile overlapped Refresh, comment feed repeated author identity, and attachment delete stayed enabled during a draft.", "Profile, comment identity, and draft-state actions remain readable and safe.", "Moved profile action, normalized comment display, and disabled delete while a draft exists; browser smoke passed."),
    ("IDTS-78", "Standalone AI sections appear outside business context", "AI actions remained manifest custom sections with visible titles after placement polish.", "AI review actions appear inline near their business context, without standalone Object Page sections.", "Removed standalone sections, retained review-only action rows, and moved handoff summary into History; deployed QA verification passed."),
]

# Keep the Vietnamese file genuinely review-ready instead of merely relabelling
# English content.  The defect IDs are the cross-language traceability key.
VI_DEFECTS = [
    ("IDTS-13", "Object Page không làm mới sau thao tác vòng đời", "Side effect của Fiori chỉ nhắm vào thuộc tính thay vì entity đang bind, nên Object Page không được làm mới đầy đủ sau khi thay đổi trạng thái.", "Sau thao tác vòng đời thành công, bug đang mở, trạng thái, người phụ trách và các phần liên quan được làm mới.", "Đổi side effect của action thành `TargetEntities: [$self]`; CAP compile và UAT trên trình duyệt đã đạt."),
    ("IDTS-19", "History Timeline lỗi khi có nội dung văn bản", "Biểu thức XML boolean dùng `${...}` bên trong binding expression và gây UI5 FormatException khi timeline có text.", "History hiển thị các dòng timeline dễ đọc cho event có nội dung.", "Thay mẫu biểu thức không hợp lệ và kiểm thử lại bằng UAT trên trình duyệt."),
    ("IDTS-32", "Chấp nhận giá trị priority và severity không hợp lệ", "Mã catalog nhập tự do có thể được activate và vượt qua kiểm tra tính hợp lệ của master data.", "Chỉ giá trị catalog đang active mới được người dùng có quyền lưu.", "Bổ sung CAP target/assert và kiểm tra active-code; regression trọng tâm đã đạt."),
    ("IDTS-35", "Ứng dụng trắng trang sau khi đăng nhập thành công", "Yêu cầu OData metadata chạy trước khi XHR interceptor gắn bearer token được cài đặt.", "Sau đăng nhập, metadata tải kèm bearer authentication và danh sách bug hiển thị.", "Chuyển guard/interceptor lên trước UI5 bootstrap và sửa async closure; UI5 build và regression đã đạt."),
    ("IDTS-41", "Developer có thể tạo root draft không được cấp quyền", "Quyền tạo chưa được kiểm tra tại root draft `NEW`.", "Chỉ Tester và PM được tạo bug draft.", "Bổ sung authorization phía backend trước khi tạo draft; kiểm tra programmatic đã đạt."),
    ("IDTS-49", "Kích hoạt draft bị từ chối do thiếu reporter", "Reporter chỉ được suy ra ở CREATE active, sau khi validation khi SAVE draft đã chạy.", "Người dùng được cấp quyền có thể activate draft mà không phải gửi reporter từ client.", "Khởi tạo reporter ở draft NEW/SAVE trước kiểm tra required-field; regression trọng tâm đã đạt."),
    ("IDTS-52", "Trang đăng nhập UI5 hiển thị trắng", "Truy cập enum UI5 không ổn định cố đọc giá trị `Error` chưa xác định.", "Các control đăng nhập hiển thị và phản hồi validation an toàn.", "Dùng giá trị chuỗi enum UI5 ổn định; browser smoke đã đạt."),
    ("IDTS-53", "Lớp phủ profile chặn nút Sign Out", "Profile host fixed che phủ control sign-out.", "Người dùng mở profile popover và sign out ổn định.", "Điều chỉnh pointer-event cho host và button; browser smoke đã đạt."),
    ("IDTS-55", "Lỗi tương tác Comments và Attachments trên Object Page", "Custom fragment mất binding context, relative refresh không được hỗ trợ làm comment cũ, và child context của attachment bị coi là bug root.", "Comment được bật/đăng/tải lại ngay và thao tác xoá attachment dùng đúng bug context.", "Bổ sung public context propagation, refresh flow được hỗ trợ và nhận diện root-context chặt chẽ; interaction smoke đã đạt."),
    ("IDTS-56", "Smart Assign coi component category ẩn là thiếu", "UI phụ thuộc vào trường nội bộ không luôn được chọn trên Object Page.", "Smart Assign suy ra/dùng classification context hợp lệ và vẫn tôn trọng quyền gán thủ công.", "Sửa cách xử lý context và bổ sung UI/browser test trọng tâm."),
    ("IDTS-58", "Lỗi UX profile/comment/attachment của Sprint 4", "Profile Dashboard chồng lên Refresh, comment feed lặp thông tin author và xoá attachment vẫn bật khi có draft.", "Profile, thông tin comment và action theo trạng thái draft luôn rõ ràng, an toàn.", "Di chuyển profile action, chuẩn hoá hiển thị comment và khoá delete khi có draft; browser smoke đã đạt."),
    ("IDTS-78", "Phần AI độc lập nằm ngoài business context", "Các AI action vẫn là manifest custom section có tiêu đề hiển thị sau đợt tinh chỉnh vị trí.", "AI review action xuất hiện inline gần business context và không còn Object Page section độc lập.", "Loại bỏ standalone section, giữ action row review-only và chuyển handoff summary vào History; deployed QA verification đã đạt."),
]


def copy_row_style(ws, source, target):
    for column in range(1, 7):
        ws.cell(target, column)._style = copy(ws.cell(source, column)._style)
        ws.cell(target, column).alignment = copy(ws.cell(source, column).alignment)


def row_height(values):
    """Size rows from the template's visible column widths, avoiding clipped text."""
    widths = (41.25, 43.88, 35.75, 13.0, 13.0)
    line_count = max(len(str(value)) / width for value, width in zip(values, widths))
    return max(90, min(210, round((line_count + 1) * 15)))


def fill(language):
    OUT.mkdir(parents=True, exist_ok=True)
    output = OUT / f"Test_And_Fix_Bug_IDTS_SAP01_{language}_v{VERSION}.xlsx"
    shutil.copy2(TEMPLATE, output)
    wb = load_workbook(output)
    wb.defined_names.clear()
    ws = wb["Fix and bugs"]
    headers = ("No", "Bug", "Details", "Expected result", "Fix") if language == "en" else ("STT", "Lỗi sản phẩm", "Chi tiết", "Kết quả mong đợi", "Cách sửa")
    for cell, value in zip(("A1", "B1", "C1", "D1", "E1"), headers):
        ws[cell] = value
    defects = DEFECTS if language == "en" else VI_DEFECTS
    # The official template deliberately merges rows 3-4 and 6-7.  Use their
    # writable anchor rows, then extend only the unmerged continuation rows.
    target_rows = [2, 3, 5, 6, *range(8, 8 + len(defects) - 4)]
    for index, (defect, row) in enumerate(zip(defects, target_rows), 1):
        if row >= 8:
            copy_row_style(ws, 2, row)
        issue, title, details, expected, fix = defect
        ws.cell(row, 1).value = index
        ws.cell(row, 2).value = f"{issue} - {title}"
        ws.cell(row, 3).value = details
        ws.cell(row, 4).value = expected
        ws.cell(row, 5).value = fix
        source = "Member status + QA evidence" if language == "en" else "Member status + bằng chứng QA"
        ws.cell(row, 6).value = f"{source} ({issue})"
        for column in range(2, 7):
            ws.cell(row, column).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = row_height((title, details, expected, fix, f"{source} ({issue})"))
    unused_notes = {
        "en": (
            "Intentionally unused — confirmed IDTS product defects are consolidated in 'Fix and bugs'; no second defect category is used.",
            "N/A — no additional product-defect dataset is maintained for this review version.",
        ),
        "vi": (
            "Cố ý không sử dụng — các lỗi sản phẩm IDTS đã xác nhận được tổng hợp trong sheet 'Fix and bugs'; không dùng nhóm lỗi thứ hai.",
            "N/A — phiên bản review này không duy trì thêm bộ dữ liệu lỗi sản phẩm khác.",
        ),
    }
    for sheet_name, note in zip(("Issue 2", "Issue 4"), unused_notes[language]):
        note_ws = wb[sheet_name]
        note_ws["A1"] = note
        note_ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
        note_ws.column_dimensions["A"].width = 80
        note_ws.row_dimensions[1].height = 48
    wb.properties.title = f"IDTS SAP490 Test and Fix Bug {language.upper()} v{VERSION}"
    wb.properties.subject = "Confirmed product defects only"
    wb.save(output)
    print(output.relative_to(ROOT))


if __name__ == "__main__":
    fill("en")
    fill("vi")
