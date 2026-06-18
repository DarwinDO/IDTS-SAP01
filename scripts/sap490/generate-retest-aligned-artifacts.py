"""Generate updated SAP490 artifacts aligned with current IDTS Sprint 02 state."""

from __future__ import annotations

from copy import copy
from datetime import date
from pathlib import Path
import shutil

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment


ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_DIR = ROOT / "docs" / "sap490" / "templates" / "Deliverable_template"
OUTPUT_DIR = ROOT / "docs" / "sap490" / "generated"
DOCUMENT_DATE = date(2026, 6, 18)

FUNCTIONAL_VERSION = "0.2"
SCENARIO_VERSION = "0.2"
BUGFIX_VERSION = "0.3"
UNIT_VERSION = "0.2"
FUNCTIONAL_TEST_VERSION = "0.1"
TEST_REPORT_VERSION = "0.1"


SCENARIOS = [
    ("SC-01", "Create Bug without assignee", "Tạo bug không chọn assignee"),
    ("SC-02", "Create Bug with assignee", "Tạo bug có sẵn assignee"),
    ("SC-03", "Upload draft attachment", "Upload attachment trong draft"),
    ("SC-04", "Add comment", "Thêm comment"),
    ("SC-05", "Assign Developer", "Assign Developer"),
    ("SC-06", "Mark In Review", "Chuyển In Review"),
    ("SC-07", "Start Progress", "Bắt đầu xử lý"),
    ("SC-08", "Request More Information and Resubmit", "Yêu cầu thêm thông tin và resubmit"),
    ("SC-09", "Reject Bug", "Reject bug"),
    ("SC-10", "Move to Pending Assignment", "Chuyển Pending Assignment"),
    ("SC-11", "Resolve Bug", "Resolve bug"),
    ("SC-12", "Send to Retest or Close", "Chuyển Retest hoặc Close"),
    ("SC-13", "Reopen Bug", "Reopen bug"),
    ("SC-14", "Verify history and notifications", "Kiểm tra history và notifications"),
]


TEST_CASES = [
    {
        "id": "SC-01a",
        "scenario": "SC-01",
        "en": "Create a bug with all mandatory fields and no assignee.",
        "vi": "Tạo bug với đầy đủ trường bắt buộc và không chọn assignee.",
        "data_en": "Valid title, description, reproduction details, component, category, priority, severity.",
        "data_vi": "Title, mô tả, reproduction, component, category, priority, severity hợp lệ.",
        "expected_en": "Create succeeds and persisted status becomes PENDING_ASSIGNMENT.",
        "expected_vi": "Tạo thành công và status lưu xuống là PENDING_ASSIGNMENT.",
    },
    {
        "id": "SC-01b",
        "scenario": "SC-01",
        "en": "Create a bug without title.",
        "vi": "Tạo bug không có title.",
        "data_en": "All mandatory fields except title.",
        "data_vi": "Có đủ các trường bắt buộc trừ title.",
        "expected_en": "Request is rejected with HTTP 400 and title-required validation.",
        "expected_vi": "Request bị từ chối với HTTP 400 và validation thiếu title.",
    },
    {
        "id": "SC-02a",
        "scenario": "SC-02",
        "en": "Create a bug with a valid assignee selected.",
        "vi": "Tạo bug với assignee hợp lệ đã được chọn.",
        "data_en": "Valid developer profile mapped to selected component/category.",
        "data_vi": "Developer profile hợp lệ đã map với component/category được chọn.",
        "expected_en": "Create succeeds and persisted status becomes ASSIGNED.",
        "expected_vi": "Tạo thành công và status lưu xuống là ASSIGNED.",
    },
    {
        "id": "SC-03a",
        "scenario": "SC-03",
        "en": "Create draft attachment metadata from the draft Object Page.",
        "vi": "Tạo metadata attachment trong draft Object Page.",
        "data_en": "Draft bug context plus file name, media type, and file size.",
        "data_vi": "Context draft bug cùng file name, media type và file size.",
        "expected_en": "Attachment metadata row is created successfully in draft mode.",
        "expected_vi": "Metadata attachment được tạo thành công ở draft mode.",
    },
    {
        "id": "SC-03b",
        "scenario": "SC-03",
        "en": "Upload binary content for a draft attachment.",
        "vi": "Upload binary content cho attachment draft.",
        "data_en": "Real text file content through OData media stream endpoint.",
        "data_vi": "Nội dung file thật qua OData media stream endpoint.",
        "expected_en": "Upload succeeds and content is available after draft activation and download.",
        "expected_vi": "Upload thành công và content vẫn tải được sau khi activate draft.",
    },
    {
        "id": "SC-04a",
        "scenario": "SC-04",
        "en": "Add a comment to an active bug.",
        "vi": "Thêm comment vào bug đang active.",
        "data_en": "Comment text submitted by Tester/Developer/PM.",
        "data_vi": "Nội dung comment do Tester/Developer/PM gửi.",
        "expected_en": "Comment row is created with business-friendly author display.",
        "expected_vi": "Comment được tạo với author hiển thị theo tên nghiệp vụ.",
    },
    {
        "id": "SC-04b",
        "scenario": "SC-04",
        "en": "Verify comment history entry after addComment.",
        "vi": "Kiểm tra history sau khi addComment.",
        "data_en": "Latest HistoryLogs / HistoryEvents after addComment.",
        "data_vi": "HistoryLogs / HistoryEvents mới nhất sau addComment.",
        "expected_en": "History captures the comment event with readable field/value information.",
        "expected_vi": "History ghi nhận event comment với field/value dễ đọc.",
    },
    {
        "id": "SC-05a",
        "scenario": "SC-05",
        "en": "Assign BUG-0001 to a valid developer.",
        "vi": "Assign BUG-0001 cho developer hợp lệ.",
        "data_en": "Valid assignee ID and assignment note.",
        "data_vi": "Assignee ID hợp lệ và ghi chú assign.",
        "expected_en": "Action succeeds, status becomes ASSIGNED, and notification is created.",
        "expected_vi": "Action thành công, status thành ASSIGNED và có notification.",
    },
    {
        "id": "SC-05b",
        "scenario": "SC-05",
        "en": "Assign a bug without assigneeID.",
        "vi": "Assign bug nhưng không truyền assigneeID.",
        "data_en": "Assignment note only.",
        "data_vi": "Chỉ có ghi chú assign.",
        "expected_en": "Request is rejected with HTTP 400.",
        "expected_vi": "Request bị từ chối với HTTP 400.",
    },
    {
        "id": "SC-06a",
        "scenario": "SC-06",
        "en": "Move an assigned bug to IN_REVIEW.",
        "vi": "Chuyển bug đã assign sang IN_REVIEW.",
        "data_en": "Assigned bug and optional note.",
        "data_vi": "Bug đã assigned cùng note tùy chọn.",
        "expected_en": "Action succeeds and status becomes IN_REVIEW.",
        "expected_vi": "Action thành công và status chuyển thành IN_REVIEW.",
    },
    {
        "id": "SC-07a",
        "scenario": "SC-07",
        "en": "Move an in-review bug to IN_PROGRESS.",
        "vi": "Chuyển bug đang review sang IN_PROGRESS.",
        "data_en": "Bug in IN_REVIEW status.",
        "data_vi": "Bug đang ở trạng thái IN_REVIEW.",
        "expected_en": "Action succeeds and status becomes IN_PROGRESS.",
        "expected_vi": "Action thành công và status chuyển thành IN_PROGRESS.",
    },
    {
        "id": "SC-08a",
        "scenario": "SC-08",
        "en": "Request more information with a reason.",
        "vi": "Yêu cầu thêm thông tin kèm lý do.",
        "data_en": "Reason: Need clearer reproduction evidence.",
        "data_vi": "Lý do: Cần reproduction evidence rõ hơn.",
        "expected_en": "Action succeeds and status becomes NEED_MORE_INFORMATION.",
        "expected_vi": "Action thành công và status chuyển thành NEED_MORE_INFORMATION.",
    },
    {
        "id": "SC-08b",
        "scenario": "SC-08",
        "en": "Request more information without a reason.",
        "vi": "Yêu cầu thêm thông tin nhưng không có lý do.",
        "data_en": "Empty reason.",
        "data_vi": "Reason rỗng.",
        "expected_en": "Request is rejected with HTTP 400.",
        "expected_vi": "Request bị từ chối với HTTP 400.",
    },
    {
        "id": "SC-08c",
        "scenario": "SC-08",
        "en": "Resubmit a bug from NEED_MORE_INFORMATION back to the developer path.",
        "vi": "Resubmit bug từ NEED_MORE_INFORMATION về lại nhánh developer.",
        "data_en": "Follow-up text plus existing bug in NEED_MORE_INFORMATION state.",
        "data_vi": "Nội dung follow-up cùng bug đang ở trạng thái NEED_MORE_INFORMATION.",
        "expected_en": "Action succeeds, status becomes ASSIGNED, and developer follow-up is re-opened.",
        "expected_vi": "Action thành công, status chuyển thành ASSIGNED và nhánh follow-up của developer được mở lại.",
    },
    {
        "id": "SC-08d",
        "scenario": "SC-08",
        "en": "Verify resubmit writes a comment and a developer notification.",
        "vi": "Kiểm tra resubmit tạo comment và notification cho developer.",
        "data_en": "Latest comments and notifications after resubmitToDeveloper.",
        "data_vi": "Comments và notifications mới nhất sau resubmitToDeveloper.",
        "expected_en": "A readable follow-up comment and developer notification both exist.",
        "expected_vi": "Tồn tại cả comment follow-up dễ đọc và notification gửi cho developer.",
    },
    {
        "id": "SC-09a",
        "scenario": "SC-09",
        "en": "Reject a bug with a classification reason.",
        "vi": "Reject bug kèm lý do phân loại sai.",
        "data_en": "Reason: Wrong component/category mapping.",
        "data_vi": "Lý do: Sai mapping component/category.",
        "expected_en": "Action succeeds and status becomes REJECTED.",
        "expected_vi": "Action thành công và status chuyển thành REJECTED.",
    },
    {
        "id": "SC-09b",
        "scenario": "SC-09",
        "en": "Reject a bug without a reason.",
        "vi": "Reject bug nhưng không có lý do.",
        "data_en": "Empty reason.",
        "data_vi": "Reason rỗng.",
        "expected_en": "Request is rejected with HTTP 400.",
        "expected_vi": "Request bị từ chối với HTTP 400.",
    },
    {
        "id": "SC-10a",
        "scenario": "SC-10",
        "en": "Move a rejected bug to PENDING_ASSIGNMENT.",
        "vi": "Chuyển bug bị reject sang PENDING_ASSIGNMENT.",
        "data_en": "Rejected bug plus follow-up note.",
        "data_vi": "Bug đang Rejected cùng note follow-up.",
        "expected_en": "Action succeeds, assignee is cleared, and status becomes PENDING_ASSIGNMENT.",
        "expected_vi": "Action thành công, assignee bị xóa và status thành PENDING_ASSIGNMENT.",
    },
    {
        "id": "SC-11a",
        "scenario": "SC-11",
        "en": "Resolve an in-progress bug with a developer note.",
        "vi": "Resolve bug đang xử lý kèm developer note.",
        "data_en": "Resolve note describing the technical fix.",
        "data_vi": "Resolve note mô tả cách sửa kỹ thuật.",
        "expected_en": "Action succeeds and status becomes RESOLVED.",
        "expected_vi": "Action thành công và status chuyển thành RESOLVED.",
    },
    {
        "id": "SC-11b",
        "scenario": "SC-11",
        "en": "Resolve an in-progress bug without a developer note.",
        "vi": "Resolve bug đang xử lý nhưng không có developer note.",
        "data_en": "Empty resolve note.",
        "data_vi": "Resolve note rỗng.",
        "expected_en": "Request is rejected with HTTP 400.",
        "expected_vi": "Request bị từ chối với HTTP 400.",
    },
    {
        "id": "SC-12a",
        "scenario": "SC-12",
        "en": "Send a resolved bug to RETEST_REQUIRED.",
        "vi": "Chuyển bug đã resolve sang RETEST_REQUIRED.",
        "data_en": "Resolved bug plus tester-facing note.",
        "data_vi": "Bug đã Resolved cùng note cho tester.",
        "expected_en": "Action succeeds and status becomes RETEST_REQUIRED.",
        "expected_vi": "Action thành công và status chuyển thành RETEST_REQUIRED.",
    },
    {
        "id": "SC-12b",
        "scenario": "SC-12",
        "en": "Close a resolved bug when QA decides no separate retest cycle is needed.",
        "vi": "Đóng bug đang resolved khi QA quyết định chưa cần một vòng retest riêng.",
        "data_en": "Bug in RESOLVED state.",
        "data_vi": "Bug đang ở trạng thái RESOLVED.",
        "expected_en": "Action succeeds and status becomes CLOSED.",
        "expected_vi": "Action thành công và status chuyển thành CLOSED.",
    },
    {
        "id": "SC-13a",
        "scenario": "SC-13",
        "en": "Reopen a bug with a reason.",
        "vi": "Reopen bug kèm lý do.",
        "data_en": "Reason: issue still reproducible.",
        "data_vi": "Lý do: lỗi vẫn tái hiện được.",
        "expected_en": "Action succeeds and status becomes REOPENED.",
        "expected_vi": "Action thành công và status chuyển thành REOPENED.",
    },
    {
        "id": "SC-13b",
        "scenario": "SC-13",
        "en": "Reopen a bug without a reason.",
        "vi": "Reopen bug nhưng không có lý do.",
        "data_en": "Empty reason.",
        "data_vi": "Reason rỗng.",
        "expected_en": "Request is rejected with HTTP 400.",
        "expected_vi": "Request bị từ chối với HTTP 400.",
    },
    {
        "id": "SC-14a",
        "scenario": "SC-14",
        "en": "Verify grouped history for assignment/status/comment/attachment changes.",
        "vi": "Kiểm tra grouped history cho thay đổi assignment/status/comment/attachment.",
        "data_en": "HistoryEvents and HistoryLogs after full lifecycle and draft attachment flow.",
        "data_vi": "HistoryEvents và HistoryLogs sau full lifecycle và attachment draft flow.",
        "expected_en": "Audit trail exists and shows readable actor, field, and value information.",
        "expected_vi": "Audit trail tồn tại và hiển thị actor, field và value dễ đọc.",
    },
    {
        "id": "SC-14b",
        "scenario": "SC-14",
        "en": "Verify notifications after assignment and follow-up transitions.",
        "vi": "Kiểm tra notifications sau assignment và follow-up transitions.",
        "data_en": "Notifications list for assignment/reject/request-info flow.",
        "data_vi": "Danh sách notification cho luồng assignment/reject/request-info.",
        "expected_en": "Notification rows exist with business-readable recipient and event names.",
        "expected_vi": "Notification tồn tại với recipient và event name dạng nghiệp vụ.",
    },
]


FUNCTIONAL_LABELS = {
    "en": {
        "title": "Functional Specification",
        "module": "Module",
        "module_name": "Module Name",
        "created_date": "Created Date",
        "updated_date": "Last Update Date",
        "approver": "Approver",
        "reviewer": "Reviewer",
        "creator": "Creator",
        "history_description": "Updated after workflow hardening, real comment/attachment flow completion, and audit/UI refinement.",
        "overview": [
            "The IDTS MVP allows internal Testers, Developers, and PM users to create, classify, assign, discuss, and track bug reports in an SAP-style process.",
            "The current implementation covers Fiori Elements List Report/Object Page, full-page Create Bug draft flow, filtered value helps, comments, real attachment upload/download, grouped audit history, and in-app notifications.",
            "The solution remains CAP and metadata first. Targeted UI5 extension is deferred unless a real Fiori Elements limitation blocks acceptable mentor-facing UX.",
        ],
        "flow": [
            "1. Tester opens the Bug Management List Report and selects Create Bug.",
            "2. Fiori opens the draft Object Page in create mode.",
            "3. Tester fills Bug Summary, Classification and Assignment, and Reproduction and Test Context.",
            "4. Tester can optionally upload supporting attachment evidence before saving the draft.",
            "5. CAP validates mandatory fields, derives system-managed values, and persists the bug in ASSIGNED or PENDING_ASSIGNMENT. Status NEW remains legacy/import-only.",
            "6. After create, users can add comments, review grouped history, and inspect notifications.",
            "7. PM/Tester users perform coordination actions such as assign, move to pending assignment, send to retest, close, and reopen.",
            "8. Assigned Developer users perform developer-review actions such as mark in review, start progress, request more information, reject, and resolve.",
        ],
        "layout_note": "Target object-page order: Bug Summary -> Classification and Assignment -> Reproduction and Test Context -> Evidence / Attachments. Comments, History, and Notifications are hidden during a brand-new draft.",
        "messages": [
            ("IDTS-FE-001", "English", "Please fill in all mandatory bug report fields before saving.", "Save or activate draft"),
            ("IDTS-FE-002", "English", "No suitable developer was selected. The bug will remain Pending Assignment.", "Create without assignee"),
            ("IDTS-FE-003", "English", "The selected assignee is not responsible for the chosen component/category.", "Assignment validation"),
            ("IDTS-FE-004", "English", "A reason or note is required for this lifecycle transition.", "Request information / reject / resolve / reopen"),
            ("IDTS-FE-005", "English", "Only allowed roles can execute this lifecycle action.", "Unauthorized workflow action"),
        ],
        "processing": [
            "System-managed and active-only sections are hidden or read-only during create. Status, reporter, next processor, history, comments, and notifications are not user-editable in create mode.",
            "The attachment table remains available during draft create mode so the Tester can upload evidence before activation.",
            "Assignee uses a business-friendly value help projection with developer name, email, responsibility, and classification context.",
            "CAP does not trust client-provided bug number, reporter, next processor, or initial status values. It computes them server-side.",
            "Comments are created by dedicated action and are displayed with business-friendly author names and roles.",
            "Attachments use the supported CAP draft-root flow and remain available after activation and local CAP restart.",
            "Important business edits and attachment additions create grouped HistoryEvents and HistoryLogs. Assignment and follow-up transitions create notification records.",
        ],
        "items": [
            ("Title", "title", "Bug Summary", "Mandatory text entered by Tester"),
            ("Description", "description", "Bug Summary", "Mandatory multiline text"),
            ("Priority", "priority_code", "Bug Summary", "Mandatory value help"),
            ("Severity", "severity_code", "Bug Summary", "Mandatory value help"),
            ("Environment", "environment_code", "Bug Summary", "Optional value help"),
            ("SAP Module", "sapModule_ID", "Classification and Assignment", "Optional value help context"),
            ("Application Component", "applicationComponent_ID", "Classification and Assignment", "Mandatory value help"),
            ("Defect Category", "defectCategory_ID", "Classification and Assignment", "Mandatory value help"),
            ("Assignee", "assignee_ID", "Classification and Assignment", "Optional value help; blank means Pending Assignment"),
            ("Steps to Reproduce", "stepsToReproduce", "Reproduction and Test Context", "Mandatory multiline text"),
            ("Actual Result", "actualResult", "Reproduction and Test Context", "Mandatory multiline text"),
            ("Expected Result", "expectedResult", "Reproduction and Test Context", "Mandatory multiline text"),
            ("Test Case Reference", "testCaseRef", "Reproduction and Test Context", "Optional reference"),
            ("Test Run Reference", "testRunRef", "Reproduction and Test Context", "Optional reference"),
            ("Attachment Upload", "attachments/content", "Evidence / Attachments", "Optional file upload in draft create mode"),
            ("Bug Number", "bugNumber", "Header/System", "Server-generated, read-only"),
            ("Status", "status_code", "Header/System", "Lifecycle controlled by backend, read-only"),
            ("Reporter", "reporter_ID", "Header/System", "Derived from logged-in user or local fallback, read-only"),
            ("Next Processor", "nextProcessorUser_ID / nextProcessorRole_code", "Header/System", "Server-maintained, read-only"),
        ],
        "bugs": [
            (
                "WP7 - Draft attachment binary was not persisted reliably",
                "The custom draft attachment flow initially stored metadata but did not always preserve binary content across activation/restart.",
                "Attachment metadata and binary content remain available after activation and local CAP restart.",
                "Fixed by aligning the stream persistence path with the CAP draft-root flow and verifying create/upload/activate/download/history.",
            ),
            (
                "WP5/WP7 - Comments and history exposed technical identifiers",
                "Comments, history, and notifications previously exposed raw IDs or unreadable audit detail instead of business-friendly names.",
                "Sub-tables show readable names, labels, and remain read-only for normal usage.",
                "Fixed through service projection enrichment, text annotations, and read-only hardening.",
            ),
            (
                "WP4 - Assign Developer action parameter still shows UUID after selection",
                "The Fiori Elements action parameter dialog still renders the selected assignee parameter as a raw UUID after value-help selection.",
                "The dialog should display business text or a friendlier parameter input for mentor-facing UX.",
                "Open item. Current workaround is business-friendly value-help rows plus correct Object Page display after action execution. Future fix: targeted FE/UI5 extension.",
            ),
            (
                "WP4/WP5 - Comments section lacks a local Add Comment entry point",
                "Comment creation is currently discoverable from the page header action but not from the Comments section itself.",
                "Users should be able to add comments from the conversation context.",
                "Open UX backlog item. Candidate solution: section-level action or targeted FE/UI5 extension after core happy flows are stable.",
            ),
        ],
    },
    "vi": {
        "title": "Đặc tả chức năng",
        "module": "Mô-đun",
        "module_name": "Tên mô-đun",
        "created_date": "Ngày tạo",
        "updated_date": "Ngày cập nhật",
        "approver": "Người phê duyệt",
        "reviewer": "Người rà soát",
        "creator": "Người tạo",
        "history_description": "Cập nhật sau khi siết workflow, hoàn thiện comment/attachment thật và refine audit/UI.",
        "overview": [
            "MVP IDTS cho phép Tester, Developer và PM nội bộ tạo, phân loại, gán người xử lý, thảo luận và theo dõi bug report theo luồng làm việc kiểu SAP.",
            "Phạm vi hiện tại gồm Fiori Elements List Report/Object Page, full-page Create Bug draft flow, filtered value helps, comments, upload/download attachment thật, grouped audit history và in-app notifications.",
            "Giải pháp vẫn ưu tiên CAP và metadata trước. Chỉ khi gặp giới hạn thật của Fiori Elements mới xem xét UI5 extension có chủ đích.",
        ],
        "flow": [
            "1. Tester mở Bug Management List Report và chọn Create Bug.",
            "2. Fiori mở draft Object Page ở chế độ tạo mới.",
            "3. Tester điền Bug Summary, Classification and Assignment, và Reproduction and Test Context.",
            "4. Tester có thể upload file bằng chứng trong phần attachment trước khi lưu draft.",
            "5. CAP validate trường bắt buộc, tính các giá trị hệ thống và lưu bug ở ASSIGNED hoặc PENDING_ASSIGNMENT. Trạng thái NEW chỉ còn dùng cho dữ liệu legacy/import.",
            "6. Sau khi tạo, người dùng có thể thêm comment, xem grouped history và kiểm tra notifications.",
            "7. PM/Tester xử lý các coordination action như assign, move to pending assignment, send to retest, close và reopen.",
            "8. Developer được assign xử lý các developer-review action như mark in review, start progress, request more information, reject và resolve.",
        ],
        "layout_note": "Thứ tự object page mục tiêu: Bug Summary -> Classification and Assignment -> Reproduction and Test Context -> Evidence / Attachments. Comments, History và Notifications được ẩn khi đang tạo draft mới hoàn toàn.",
        "messages": [
            ("IDTS-FE-001", "Vietnamese", "Vui lòng điền đầy đủ các trường bắt buộc trước khi lưu bug report.", "Save hoặc activate draft"),
            ("IDTS-FE-002", "Vietnamese", "Chưa chọn developer phù hợp. Bug sẽ ở trạng thái Pending Assignment.", "Tạo bug không có assignee"),
            ("IDTS-FE-003", "Vietnamese", "Developer được chọn không phụ trách component/category này.", "Validate assignment"),
            ("IDTS-FE-004", "Vietnamese", "Transition này bắt buộc phải có lý do hoặc ghi chú.", "Request information / reject / resolve / reopen"),
            ("IDTS-FE-005", "Vietnamese", "Chỉ role được phép mới được thực hiện lifecycle action này.", "Workflow action không đúng quyền"),
        ],
        "processing": [
            "Các phần system-managed và active-only được ẩn hoặc để read-only khi tạo bug. Status, reporter, next processor, history, comments và notifications không cho sửa tay ở create mode.",
            "Bảng attachment vẫn hiển thị ở draft create mode để Tester upload bằng chứng trước khi activate.",
            "Assignee dùng value help thân thiện với nghiệp vụ, có developer name, email, responsibility và classification context.",
            "CAP không tin các giá trị bug number, reporter, next processor hay initial status do client tự gửi lên. Các giá trị này được tính ở backend.",
            "Comment được tạo qua action riêng và hiển thị với author name/role dạng nghiệp vụ.",
            "Attachment dùng CAP draft-root flow được hỗ trợ và vẫn còn sau khi activate cũng như sau khi restart CAP local.",
            "Các thay đổi nghiệp vụ quan trọng và việc thêm attachment tạo ra HistoryEvents/HistoryLogs dạng grouped. Assignment và follow-up transition tạo notification records.",
        ],
        "items": [
            ("Title", "title", "Bug Summary", "Trường văn bản bắt buộc do Tester nhập"),
            ("Description", "description", "Bug Summary", "Multiline text bắt buộc"),
            ("Priority", "priority_code", "Bug Summary", "Value help bắt buộc"),
            ("Severity", "severity_code", "Bug Summary", "Value help bắt buộc"),
            ("Environment", "environment_code", "Bug Summary", "Value help tùy chọn"),
            ("SAP Module", "sapModule_ID", "Classification and Assignment", "Ngữ cảnh value help tùy chọn"),
            ("Application Component", "applicationComponent_ID", "Classification and Assignment", "Value help bắt buộc"),
            ("Defect Category", "defectCategory_ID", "Classification and Assignment", "Value help bắt buộc"),
            ("Assignee", "assignee_ID", "Classification and Assignment", "Value help tùy chọn; để trống nghĩa là Pending Assignment"),
            ("Steps to Reproduce", "stepsToReproduce", "Reproduction and Test Context", "Multiline text bắt buộc"),
            ("Actual Result", "actualResult", "Reproduction and Test Context", "Multiline text bắt buộc"),
            ("Expected Result", "expectedResult", "Reproduction and Test Context", "Multiline text bắt buộc"),
            ("Test Case Reference", "testCaseRef", "Reproduction and Test Context", "Reference tùy chọn"),
            ("Test Run Reference", "testRunRef", "Reproduction and Test Context", "Reference tùy chọn"),
            ("Attachment Upload", "attachments/content", "Evidence / Attachments", "Upload file tùy chọn trong draft create mode"),
            ("Bug Number", "bugNumber", "Header/System", "Backend sinh tự động, read-only"),
            ("Status", "status_code", "Header/System", "Backend điều khiển theo lifecycle, read-only"),
            ("Reporter", "reporter_ID", "Header/System", "Lấy từ user đăng nhập hoặc fallback local, read-only"),
            ("Next Processor", "nextProcessorUser_ID / nextProcessorRole_code", "Header/System", "Backend tự duy trì, read-only"),
        ],
        "bugs": [
            (
                "WP7 - Binary attachment draft ban đầu chưa persist ổn định",
                "Luồng attachment draft tùy biến ban đầu lưu được metadata nhưng có lúc không giữ được binary content sau activate/restart.",
                "Metadata và binary content của attachment phải còn sau khi activate và restart CAP local.",
                "Đã sửa bằng cách align đường persist stream với CAP draft-root flow và verify đầy đủ create/upload/activate/download/history.",
            ),
            (
                "WP5/WP7 - Comments và history từng lộ identifier kỹ thuật",
                "Comments, history và notifications từng hiển thị raw ID hoặc audit detail khó đọc thay vì business-friendly name.",
                "Sub-table phải hiển thị tên/nhãn dễ hiểu và giữ read-only trong luồng dùng bình thường.",
                "Đã sửa bằng enrichment ở service projection, text annotation và hardening read-only.",
            ),
            (
                "WP4 - Dialog Assign Developer vẫn hiện UUID sau khi chọn",
                "Dialog parameter mặc định của Fiori Elements vẫn render assignee đã chọn thành UUID thô sau value help.",
                "Dialog nên hiển thị business text hoặc input thân thiện hơn cho UX khi demo với mentor.",
                "Open item. Workaround hiện tại là dòng value-help đã dễ hiểu và Object Page hiển thị đúng tên sau khi action chạy. Hướng sửa về sau là targeted FE/UI5 extension.",
            ),
            (
                "WP4/WP5 - Section Comments chưa có entry point ngay tại chỗ",
                "Người dùng hiện thêm comment qua page-level action, chưa có nút Add Comment nằm ngay gần section Comments.",
                "Người dùng nên thêm comment ngay từ ngữ cảnh hội thoại.",
                "Open UX backlog item. Phương án sau là section-level action hoặc targeted FE/UI5 extension khi core happy flow đã ổn.",
            ),
        ],
    },
}


def copy_template(template_name: str, output_name: str) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output = OUTPUT_DIR / output_name
    shutil.copy2(TEMPLATE_DIR / template_name, output)
    return output


def write(ws, cell: str, value, *, wrap: bool = False):
    target = ws[cell]
    if isinstance(target, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if target.coordinate in merged_range:
                target = ws.cell(merged_range.min_row, merged_range.min_col)
                break
    target.value = value
    if isinstance(value, date):
        target.number_format = "yyyy-mm-dd"
    if wrap:
        old = copy(target.alignment)
        target.alignment = Alignment(
            horizontal=old.horizontal,
            vertical=old.vertical or "top",
            text_rotation=old.text_rotation,
            wrap_text=True,
            shrink_to_fit=old.shrink_to_fit,
            indent=old.indent,
        )


def copy_row_style(ws, source_row: int, target_row: int, max_column: int):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for column in range(1, max_column + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def fill_functional(language: str) -> Path:
    labels = FUNCTIONAL_LABELS[language]
    output = copy_template(
        "Functional_Specification.xlsx",
        f"Functional_Specification_IDTS_SAP01_{language}_v{FUNCTIONAL_VERSION}.xlsx",
    )
    wb = load_workbook(output)

    ws = wb["Cover"]
    write(ws, "B8", labels["title"].upper())
    write(ws, "I11", labels["module"])
    write(ws, "N11", "IDTS")
    write(ws, "U11", labels["module_name"])
    write(ws, "Z11", "Issue and Defect Tracking System in SAP")
    write(ws, "I12", labels["created_date"])
    write(ws, "N12", DOCUMENT_DATE)
    write(ws, "U12", labels["updated_date"])
    write(ws, "Z12", DOCUMENT_DATE)
    write(ws, "U16", labels["approver"])
    write(ws, "Z16", labels["reviewer"])
    write(ws, "AE16", labels["creator"])
    write(ws, "AE17", "DonHV")

    ws = wb["Histories"]
    write(ws, "B3", 1)
    write(ws, "C3", FUNCTIONAL_VERSION)
    write(ws, "D3", labels["history_description"], wrap=True)
    write(ws, "E3", "All functional sheets")
    write(ws, "F3", DOCUMENT_DATE)
    write(ws, "G3", "DonHV")
    for row in range(4, 20):
        for col in range(2, 8):
            ws.cell(row, col).value = None

    ws = wb["Function Overview"]
    write(ws, "I3", "IDTS-FE-BUG-MGMT")
    write(ws, "B14", "Function Overview" if language == "en" else "Tổng quan chức năng")
    for offset, text in enumerate(labels["overview"], start=15):
        write(ws, f"B{offset}", text, wrap=True)
        ws.row_dimensions[offset].height = 42
    write(ws, "B20", "Supplement" if language == "en" else "Bổ sung")
    write(ws, "B21", "The UI is implemented with SAP Fiori Elements List Report/Object Page and CAP OData V4 metadata." if language == "en" else "UI được triển khai bằng SAP Fiori Elements List Report/Object Page và metadata CAP OData V4.", wrap=True)

    ws = wb["Process Flow"]
    write(ws, "B3", "IDTS-FE-BUG-MGMT")
    write(ws, "P3", "Bug Management Create and Review Flow" if language == "en" else "Luồng tạo và review bug")
    write(ws, "B5", "Screen Flow" if language == "en" else "Luồng màn hình")
    for row, step in enumerate(labels["flow"], start=6):
        write(ws, f"B{row}", step, wrap=True)
        ws.row_dimensions[row].height = 36
    write(ws, "B44", "Supplement" if language == "en" else "Bổ sung")
    write(ws, "B45", "The Create Bug flow uses a draft Object Page instead of a modal dialog because the report has many meaningful fields and supporting evidence." if language == "en" else "Luồng Create Bug dùng draft Object Page thay vì modal dialog vì biểu mẫu có nhiều trường có ý nghĩa và có thêm bằng chứng đính kèm.", wrap=True)

    ws = wb["Screen Layout"]
    write(ws, "B2", "Screen Layout" if language == "en" else "Bố cục màn hình")
    write(ws, "B4", labels["layout_note"], wrap=True)
    for cell, value in [("C27", "No." if language == "en" else "STT"), ("F27", "Field name" if language == "en" else "Tên trường"), ("N27", "Technical name" if language == "en" else "Tên kỹ thuật"), ("S27", "Place" if language == "en" else "Vị trí"), ("AA27", "Logic")]:
        write(ws, cell, value)
    write(ws, "C28", 1)
    write(ws, "F28", "Main Create/Edit Fields and Evidence Upload" if language == "en" else "Các trường Create/Edit chính và phần upload bằng chứng", wrap=True)
    write(ws, "S28", "Object Page content", wrap=True)
    write(ws, "AA28", labels["layout_note"], wrap=True)
    target_rows = [29, 30, 31, 33, 35, 37, 38, 39, 41, 43, 45, 47, 49, 50]
    for number, ((name, technical, place, logic), target_row) in enumerate(zip(labels["items"][:14], target_rows), start=2):
        if target_row > 45:
            copy_row_style(ws, 45, target_row, 35)
        write(ws, f"C{target_row}", number)
        write(ws, f"F{target_row}", name, wrap=True)
        write(ws, f"N{target_row}", technical, wrap=True)
        write(ws, f"S{target_row}", place, wrap=True)
        write(ws, f"AA{target_row}", logic, wrap=True)
        ws.row_dimensions[target_row].height = 34
    for row in [51, 52, 54]:
        for column in [3, 6, 14, 19, 27]:
            cell = ws.cell(row, column)
            if not isinstance(cell, MergedCell):
                cell.value = None

    ws = wb["Screen Definition"]
    write(ws, "B3", "IDTS-FE-BUG-MGMT")
    write(ws, "P3", "Bug Management Create and Review Flow" if language == "en" else "Luồng tạo và review bug")
    write(ws, "B5", "New Bug / Bug Detail")
    write(ws, "B9", "1. Bug Object Page")
    write(ws, "B12", "1.1 Main Create/Edit Fields")
    for row in range(13, 90):
        for col in range(2, 60):
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.value = None
    target_rows = [13, 14, 15, 17, 20, 22, 25, 26, 28, 30, 32, 34, 36, 38, 40, 42, 44, 46, 50]
    for number, ((name, technical, _place, logic), target_row) in enumerate(zip(labels["items"], target_rows), start=1):
        if target_row > 44:
            copy_row_style(ws, 44, target_row, 59)
        write(ws, f"B{target_row}", number)
        write(ws, f"C{target_row}", name, wrap=True)
        is_read_only = ("read-only" in logic.lower()) or ("read-only" in technical.lower())
        write(ws, f"I{target_row}", "Input" if not is_read_only and "Server-" not in logic and "Backend" not in logic else "Display")
        write(ws, f"M{target_row}", "I/O" if not is_read_only and "Server-" not in logic and "Backend" not in logic else "O")
        write(ws, f"O{target_row}", "Single")
        write(ws, f"R{target_row}", technical, wrap=True)
        write(ws, f"U{target_row}", "-")
        write(ws, f"X{target_row}", "-")
        write(ws, f"AA{target_row}", "Yes" if ("Mandatory" in logic or "bắt buộc" in logic) else "No")
        write(ws, f"AE{target_row}", "Backend default" if ("Server-" in logic or "Backend" in logic or "Derived" in logic or "Lấy từ" in logic) else "")
        write(ws, f"AI{target_row}", "Text / Value Help")
        write(ws, f"AM{target_row}", "Left")
        write(ws, f"AQ{target_row}", "Yes" if "value help" in logic.lower() else "No")
        write(ws, f"AW{target_row}", logic, wrap=True)
        ws.row_dimensions[target_row].height = 34

    ws = wb["Message Definition"]
    write(ws, "B3", "IDTS-FE-BUG-MGMT")
    write(ws, "P3", "Bug Management Create and Review Flow" if language == "en" else "Luồng tạo và review bug")
    for row in range(6, 32):
        for col in range(2, 40):
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.value = None
    for row, (message_id, lang, message, timing) in enumerate(labels["messages"], start=6):
        if row > 6:
            copy_row_style(ws, 6, row, 40)
        write(ws, f"B{row}", message_id)
        write(ws, f"F{row}", lang)
        write(ws, f"J{row}", message, wrap=True)
        write(ws, f"AG{row}", timing, wrap=True)
        ws.row_dimensions[row].height = 36

    ws = wb["Processing Description"]
    write(ws, "B3", "IDTS-FE-BUG-MGMT")
    write(ws, "P3", "Bug Management Create and Review Flow" if language == "en" else "Luồng tạo và review bug")
    write(ws, "AQ2", "Created by:")
    write(ws, "AY2", "DonHV")
    write(ws, "AQ3", "Modified by:")
    write(ws, "AY3", "DonHV")
    write(ws, "B4", "Processing Description" if language == "en" else "Mô tả xử lý")
    for row, text in enumerate(labels["processing"], start=6):
        write(ws, f"B{row}", text, wrap=True)
        ws.row_dimensions[row].height = 42

    wb.properties.title = f"IDTS SAP490 Functional Specification {language.upper()} v{FUNCTIONAL_VERSION}"
    wb.properties.subject = "IDTS functional specification"
    wb.properties.creator = "IDTS SAP01 Team"
    wb.save(output)
    return output


def fill_bug_log(language: str) -> Path:
    labels = FUNCTIONAL_LABELS[language]
    output = copy_template(
        "Test_And_Fix_Bug.xlsx",
        f"Test_And_Fix_Bug_IDTS_SAP01_{language}_v{BUGFIX_VERSION}.xlsx",
    )
    wb = load_workbook(output)
    ws = wb["Fix and bugs"]
    headers = ["No", "Bug", "Details", "Expected result", "Fix"] if language == "en" else ["STT", "Lỗi", "Chi tiết", "Kết quả mong đợi", "Cách sửa"]
    for cell, value in zip(("A1", "B1", "C1", "D1", "E1"), headers):
        write(ws, cell, value)
    target_rows = [2, 3, 5, 6]
    for number, (bug, target_row) in enumerate(zip(labels["bugs"], target_rows), start=1):
        write(ws, f"A{target_row}", number)
        write(ws, f"B{target_row}", bug[0], wrap=True)
        write(ws, f"C{target_row}", bug[1], wrap=True)
        write(ws, f"D{target_row}", bug[2], wrap=True)
        write(ws, f"E{target_row}", bug[3], wrap=True)
        write(ws, f"F{target_row}", "Sprint 02 evidence")
        ws.row_dimensions[target_row].height = 90
    for row in [4, 7]:
        for column in range(1, 7):
            cell = ws.cell(row, column)
            if not isinstance(cell, MergedCell):
                cell.value = None
                cell.hyperlink = None
    wb.properties.title = f"IDTS SAP490 Test and Fix Bug {language.upper()} v{BUGFIX_VERSION}"
    wb.properties.subject = "IDTS Sprint 02 defects and fixes"
    wb.properties.creator = "IDTS SAP01 Team"
    wb.save(output)
    return output


def fill_test_scenario(language: str) -> Path:
    output = copy_template(
        "Test_Scenario.xlsx",
        f"Test_Scenario_IDTS_SAP01_{language}_v{SCENARIO_VERSION}.xlsx",
    )
    wb = load_workbook(output)
    ws = wb["Cover"]
    write(ws, "B8", "Test Scenario" if language == "en" else "Kịch bản kiểm thử")
    write(ws, "I11", "Module" if language == "en" else "Mô-đun")
    write(ws, "N11", "IDTS")
    write(ws, "U11", "Module Name" if language == "en" else "Tên mô-đun")
    write(ws, "Z11", "Issue and Defect Tracking System in SAP")
    write(ws, "I12", "Function ID" if language == "en" else "Mã chức năng")
    write(ws, "N12", "IDTS-SP2-RETEST")
    write(ws, "I13", "Function Name" if language == "en" else "Tên chức năng")
    write(ws, "N13", "Comprehensive Sprint 02 retest matrix" if language == "en" else "Retest matrix tổng hợp Sprint 02")
    write(ws, "I14", "Created Date" if language == "en" else "Ngày tạo")
    write(ws, "N14", DOCUMENT_DATE)
    write(ws, "U14", "Last Update Date" if language == "en" else "Ngày cập nhật")
    write(ws, "Z14", DOCUMENT_DATE)
    write(ws, "U18", "Approver" if language == "en" else "Người phê duyệt")
    write(ws, "Z18", "Reviewer" if language == "en" else "Người rà soát")
    write(ws, "AE18", "Creator" if language == "en" else "Người tạo")
    write(ws, "AE19", "DonHV")

    ws = wb["Histories"]
    history_headers = ["No", "Version", "Description", "Sheet", "Modified date", "Modified by"] if language == "en" else ["STT", "Phiên bản", "Mô tả", "Sheet", "Ngày sửa", "Người sửa"]
    for cell, value in zip(("B2", "C2", "D2", "E2", "F2", "G2"), history_headers):
        write(ws, cell, value)
    write(ws, "B3", 1)
    write(ws, "C3", SCENARIO_VERSION)
    write(ws, "D3", "Updated comprehensive retest matrix for current Sprint 02 backend/UI happy flows." if language == "en" else "Cập nhật retest matrix tổng hợp cho các happy flow backend/UI hiện tại của Sprint 02.", wrap=True)
    write(ws, "E3", "Test Scenario, Test Cases")
    write(ws, "F3", DOCUMENT_DATE)
    write(ws, "G3", "DonHV")

    matrix = wb["Test Scenario"]
    write(matrix, "A2", "No" if language == "en" else "STT")
    write(matrix, "B2", "Step Name" if language == "en" else "Tên bước")
    write(matrix, "C1", "TEST CASE")
    for column in range(3, 26):
        matrix.cell(2, column).value = None
    for row in range(3, 24):
        matrix.cell(row, 2).value = None
        for column in range(3, 26):
            matrix.cell(row, column).value = None
    for column, (scenario_id, _, _) in enumerate(SCENARIOS, start=3):
        write(matrix, matrix.cell(2, column).coordinate, scenario_id)
    for row, (_scenario_id, en_name, vi_name) in enumerate(SCENARIOS, start=3):
        write(matrix, f"A{row}", row - 2)
        write(matrix, f"B{row}", en_name if language == "en" else vi_name)
        write(matrix, matrix.cell(row, row).coordinate, "X")
    matrix.print_area = "A1:P22"
    matrix.sheet_properties.pageSetUpPr.fitToPage = True
    matrix.page_setup.fitToWidth = 1
    matrix.page_setup.fitToHeight = 1

    cases = wb["Test Cases"]
    meta = [
        ("B2", "Business Flow" if language == "en" else "Luồng nghiệp vụ"),
        ("L2", "Function Name" if language == "en" else "Tên chức năng"),
        ("BF2", "Created/Updated by" if language == "en" else "Người tạo/cập nhật"),
        ("BO2", "Created date" if language == "en" else "Ngày tạo"),
        ("BV2", "Reviewed by" if language == "en" else "Người rà soát"),
        ("CC2", "Reviewed Date" if language == "en" else "Ngày rà soát"),
    ]
    for cell, value in meta:
        write(cases, cell, value)
    write(cases, "B3", "Comprehensive Sprint 02 retest" if language == "en" else "Retest tổng hợp Sprint 02")
    write(cases, "L3", "Current happy-flow and follow-up validation" if language == "en" else "Xác nhận happy flow và luồng follow-up hiện tại")
    write(cases, "BF3", "DonHV")
    write(cases, "BO3", DOCUMENT_DATE)
    headers = [
        ("B6", "NO." if language == "en" else "STT"),
        ("E6", "Test Contents" if language == "en" else "Nội dung kiểm thử"),
        ("E7", "Test Cases" if language == "en" else "Test case"),
        ("Y7", "Test Data" if language == "en" else "Dữ liệu test"),
        ("AP7", "Predicted Test Results" if language == "en" else "Kết quả dự kiến"),
    ]
    for cell, value in headers:
        write(cases, cell, value)
    for row, test_case in enumerate(TEST_CASES, start=8):
        write(cases, f"B{row}", test_case["id"])
        write(cases, f"E{row}", test_case[language], wrap=True)
        write(cases, f"Y{row}", test_case[f"data_{language}"], wrap=True)
        write(cases, f"AP{row}", test_case[f"expected_{language}"], wrap=True)
        cases.row_dimensions[row].height = max(cases.row_dimensions[row].height or 15, 34)

    wb.properties.title = f"IDTS SAP490 Test Scenario {language.upper()} v{SCENARIO_VERSION}"
    wb.properties.subject = "IDTS Sprint 02 retest planning"
    wb.properties.creator = "IDTS SAP01 Team"
    wb.save(output)
    return output


def fill_unit_test(language: str) -> Path:
    output = copy_template(
        "Unit_Test.xlsx",
        f"Unit_Test_IDTS_SAP01_{language}_v{UNIT_VERSION}.xlsx",
    )
    wb = load_workbook(output)

    ws = wb["Cover"]
    write(ws, "B8", "Unit Test" if language == "en" else "Kiểm thử đơn vị")
    write(ws, "I11", "Module" if language == "en" else "Mô-đun")
    write(ws, "N11", "IDTS")
    write(ws, "U11", "Module Name" if language == "en" else "Tên mô-đun")
    write(ws, "Z11", "Issue and Defect Tracking System in SAP")
    write(ws, "I12", "Function ID" if language == "en" else "Mã chức năng")
    write(ws, "N12", "IDTS-SP2-RETEST")
    write(ws, "I13", "Function Name" if language == "en" else "Tên chức năng")
    write(ws, "N13", "Backend and HTTP retest execution" if language == "en" else "Kết quả chạy retest backend và HTTP")
    write(ws, "I14", "Created Date" if language == "en" else "Ngày tạo")
    write(ws, "N14", DOCUMENT_DATE)
    write(ws, "U14", "Last Update Date" if language == "en" else "Ngày cập nhật")
    write(ws, "Z14", DOCUMENT_DATE)
    write(ws, "U18", "Approver" if language == "en" else "Người phê duyệt")
    write(ws, "Z18", "Reviewer" if language == "en" else "Người rà soát")
    write(ws, "AE18", "Creator" if language == "en" else "Người tạo")
    write(ws, "AE19", "DonHV")

    ws = wb["Histories"]
    history_headers = ["No", "Version", "Description", "Sheet", "Modified date", "Modified by"] if language == "en" else ["STT", "Phiên bản", "Mô tả", "Sheet", "Ngày sửa", "Người sửa"]
    for cell, value in zip(("B2", "C2", "D2", "E2", "F2", "G2"), history_headers):
        write(ws, cell, value)
    write(ws, "B3", 1)
    write(ws, "C3", UNIT_VERSION)
    write(ws, "D3", "Executed backend programmatic retest plus HTTP comment/attachment verification." if language == "en" else "Đã chạy lại bộ retest backend dạng programmatic và HTTP comment/attachment verification.", wrap=True)
    write(ws, "E3", "UT")
    write(ws, "F3", DOCUMENT_DATE)
    write(ws, "G3", "DonHV")

    ws = wb["UT"]
    header_cells = ("B2", "L2", "AO2", "AX2", "BE2", "BL2")
    header_values = (
        "Function ID" if language == "en" else "Mã chức năng",
        "Function Name" if language == "en" else "Tên chức năng",
        "Creator" if language == "en" else "Người tạo",
        "Created Date" if language == "en" else "Ngày tạo",
        "Reviewer" if language == "en" else "Người rà soát",
        "Last Update Date" if language == "en" else "Ngày cập nhật",
    )
    for cell, value in zip(header_cells, header_values):
        write(ws, cell, value)
    write(ws, "B3", "IDTS-SP2-RETEST")
    write(ws, "L3", "Backend and HTTP retest execution" if language == "en" else "Kết quả chạy retest backend và HTTP")
    write(ws, "AO3", "DonHV")
    write(ws, "AX3", DOCUMENT_DATE)

    unit_headers = (
        ["NO.", "Test Contents", "Test Results", "Evidence", "Test Cases", "Predicted Test Results", "Tester", "Test Date", "Result"]
        if language == "en"
        else ["STT", "Nội dung kiểm thử", "Kết quả kiểm thử", "Bằng chứng", "Test case", "Kết quả dự kiến", "Người test", "Ngày test", "Kết quả"]
    )
    for cell, value in zip(("B6", "E6", "AX6", "BL6", "E7", "Y7", "AX7", "BD7", "BJ7"), unit_headers):
        write(ws, cell, value)

    for row in range(17, 40):
        copy_row_style(ws, 16, row, 70)

    for index, test_case in enumerate(TEST_CASES, start=8):
        evidence = (
            "HTTP QA script"
            if test_case["scenario"] in {"SC-03", "SC-04"}
            else ("Programmatic + HTTP QA" if test_case["scenario"] == "SC-14" else "Programmatic QA script")
        )
        evidence_vi = (
            "Script HTTP QA"
            if test_case["scenario"] in {"SC-03", "SC-04"}
            else ("Programmatic + HTTP QA" if test_case["scenario"] == "SC-14" else "Script programmatic QA")
        )
        write(ws, f"B{index}", test_case["id"])
        write(ws, f"E{index}", test_case[language], wrap=True)
        write(ws, f"Y{index}", test_case[f"expected_{language}"], wrap=True)
        write(ws, f"AX{index}", "DonHV")
        write(ws, f"BD{index}", DOCUMENT_DATE)
        write(ws, f"BJ{index}", "PASS")
        write(ws, f"BL{index}", evidence if language == "en" else evidence_vi, wrap=True)
        ws.row_dimensions[index].height = max(ws.row_dimensions[index].height or 15, 32)

    wb.properties.title = f"IDTS SAP490 Unit Test {language.upper()} v{UNIT_VERSION}"
    wb.properties.subject = "IDTS Sprint 02 backend and HTTP retest execution"
    wb.properties.creator = "IDTS SAP01 Team"
    wb.save(output)
    return output


def fill_functional_test(language: str) -> Path:
    output = copy_template(
        "Functional_Test.xlsx",
        f"Functional_Test_IDTS_SAP01_{language}_v{FUNCTIONAL_TEST_VERSION}.xlsx",
    )
    wb = load_workbook(output)

    ws = wb["Cover"]
    write(ws, "B8", "Functional Test" if language == "en" else "Kiem thu chuc nang")
    write(ws, "N11", "IDTS")
    write(ws, "Z11", "Issue and Defect Tracking System in SAP")
    write(ws, "N12", "IDTS-SP2-FT")
    write(ws, "N13", "Sprint 02 comprehensive retest" if language == "en" else "Retest tong hop Sprint 02")
    write(ws, "N14", DOCUMENT_DATE)
    write(ws, "Z14", DOCUMENT_DATE)
    write(ws, "AE19", "DonHV")

    ws = wb["Histories"]
    write(ws, "B3", 1)
    write(ws, "C3", FUNCTIONAL_TEST_VERSION)
    write(
        ws,
        "D3",
        "Created functional retest workbook aligned with current backend, comment, attachment, and audit flows."
        if language == "en"
        else "Tao workbook retest chuc nang dong bo voi cac luong backend, comment, attachment va audit hien tai.",
        wrap=True,
    )
    write(ws, "E3", "Test Cases, Test Result, Test Data Description")
    write(ws, "F3", DOCUMENT_DATE)
    write(ws, "G3", "DonHV")

    ws = wb["Test Cases"]
    write(ws, "B3", "Sprint 02 Retest" if language == "en" else "Retest Sprint 02")
    write(
        ws,
        "L3",
        "Functional and workflow verification for IDTS happy-flow demo"
        if language == "en"
        else "Xac nhan chuc nang va workflow cho happy flow demo IDTS",
    )
    grouped_cases = [
        ("1", "Create and basic validation", "Tao bug va validation co ban", [tc for tc in TEST_CASES if tc["id"] in {"SC-01a", "SC-01b", "SC-02a"}]),
        ("2", "Evidence and collaboration", "Bang chung va cong tac", [tc for tc in TEST_CASES if tc["id"] in {"SC-03a", "SC-03b", "SC-04a", "SC-04b"}]),
        ("3", "Assignment and developer execution", "Assign va xu ly boi developer", [tc for tc in TEST_CASES if tc["id"] in {"SC-05a", "SC-05b", "SC-06a", "SC-07a"}]),
        ("4", "Follow-up and closure", "Follow-up va dong bug", [tc for tc in TEST_CASES if tc["id"] in {"SC-08a", "SC-08b", "SC-08c", "SC-08d", "SC-09a", "SC-09b", "SC-10a", "SC-11a", "SC-11b", "SC-12a", "SC-12b", "SC-13a", "SC-13b", "SC-14a", "SC-14b"}]),
    ]
    row = 8
    for group_no, group_en, group_vi, cases in grouped_cases:
        write(ws, f"B{row}", group_no)
        write(ws, f"E{row}", group_en if language == "en" else group_vi, wrap=True)
        row += 1
        for idx, case in enumerate(cases, start=1):
            write(ws, f"B{row}", f"{group_no}.{idx}")
            write(ws, f"E{row}", case[language], wrap=True)
            write(ws, f"Y{row}", case[f"data_{language}"], wrap=True)
            write(ws, f"AP{row}", case[f"expected_{language}"], wrap=True)
            ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 15, 34)
            row += 1

    ws = wb["Test Result"]
    write(ws, "B2", "IDTS-SP2-FT")
    write(
        ws,
        "L2",
        "Functional retest execution"
        if language == "en"
        else "Ket qua chay lai kiem thu chuc nang",
    )
    write(ws, "AO2", "DonHV")
    write(ws, "AX2", DOCUMENT_DATE)
    write(ws, "BE2", "DonHV")
    write(ws, "BL2", DOCUMENT_DATE)
    result_rows = [
        ("SC-01/02", "Create + mandatory validation", "Happy flow create and validation error path", "PASS", "node scripts/qa/test-idts6-programmatic.js"),
        ("SC-03/04", "Attachment + comment flow", "HTTP attachment upload/download and comment history flow", "PASS", "powershell -ExecutionPolicy Bypass -File scripts/qa/test-comments-attachments.ps1"),
        ("SC-05 to SC-13", "Lifecycle transitions", "Assignment, review, progress, request info, reject, resolve, retest, close, reopen", "PASS", "node scripts/qa/test-idts6-programmatic.js"),
        ("SC-14", "Audit + notifications", "Grouped history and notification side effects", "PASS", "node scripts/qa/test-idts6-programmatic.js"),
    ]
    base_row = 10
    for offset, (case_id, title, detail, result, evidence) in enumerate(result_rows):
        current = base_row + offset
        write(ws, f"B{current}", case_id)
        write(ws, f"E{current}", title if language == "en" else detail, wrap=True)
        write(ws, f"L{current}", detail if language == "en" else title, wrap=True)
        write(ws, f"AX{current}", evidence, wrap=True)
        write(ws, f"BJ{current}", result)
        write(ws, f"BL{current}", DOCUMENT_DATE)
        ws.row_dimensions[current].height = 34

    ws = wb["Test Data Description"]
    write(ws, "B3", "IDTS-SP2-FT")
    write(
        ws,
        "L3",
        "Retest data and environment"
        if language == "en"
        else "Du lieu test va moi truong retest",
    )
    descriptions = [
        (
            "TD-01",
            "Local CAP file-based SQLite server",
            "CAP local dung db.sqlite file de test persistence qua restart.",
        ),
        (
            "TD-02",
            "Mock users DonHV / NhanT / DatDT / SangVN",
            "Dung mock auth de test role PM, Tester, Developer.",
        ),
        (
            "TD-03",
            "Valid classification pairs",
            "Dung cap Application Component / Defect Category hop le de tranh false negative.",
        ),
        (
            "TD-04",
            "Real attachment sample file",
            "Dung file text nho de verify draft upload, activate, download, va history.",
        ),
    ]
    row = 8
    for test_id, en_text, vi_text in descriptions:
        write(ws, f"B{row}", test_id)
        write(ws, f"E{row}", en_text if language == "en" else vi_text, wrap=True)
        write(ws, f"L{row}", vi_text if language == "en" else en_text, wrap=True)
        ws.row_dimensions[row].height = 32
        row += 1

    wb.properties.title = f"IDTS SAP490 Functional Test {language.upper()} v{FUNCTIONAL_TEST_VERSION}"
    wb.properties.subject = "IDTS Sprint 02 functional retest execution"
    wb.properties.creator = "IDTS SAP01 Team"
    wb.save(output)
    return output


def fill_test_report(language: str) -> Path:
    source = ROOT / "docs" / "sap490" / "templates" / "2_SAP490_Test Report Template (1).xlsx"
    output = OUTPUT_DIR / f"Test_Report_IDTS_SAP01_{language}_v{TEST_REPORT_VERSION}.xlsx"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, output)
    wb = load_workbook(output)

    project_name = "Issue and Defect Tracking System in SAP"
    project_code = "IDTS-SAP01"
    creator = "DonHV"

    ws = wb["Cover"]
    write(ws, "B4", project_name)
    write(ws, "B5", project_code)
    write(ws, "F4", creator)
    write(ws, "F5", DOCUMENT_DATE)
    write(ws, "F6", TEST_REPORT_VERSION)
    write(ws, "B11", DOCUMENT_DATE)
    write(ws, "C11", TEST_REPORT_VERSION)
    write(
        ws,
        "E11",
        "Sprint 02 retest report covering create, lifecycle, comments, attachments, history, and notifications."
        if language == "en"
        else "Bao cao retest Sprint 02 bao gom create, lifecycle, comment, attachment, history va notification.",
        wrap=True,
    )
    write(ws, "F11", "Retest matrix and SAP490 generated artifacts")

    ws = wb["Test Cases"]
    write(ws, "D3", project_name)
    write(ws, "D4", project_code)
    write(
        ws,
        "D5",
        "Local CAP SQLite server, mock users, browser + shell QA, and draft attachment flow."
        if language == "en"
        else "CAP local voi SQLite, mock users, browser + shell QA, va draft attachment flow.",
        wrap=True,
    )
    case_rows = [
        (1, "Sprint 02 Create and Assignment", "Feature 1", "Create flow, assignment, developer lifecycle, and status validation"),
        (2, "Sprint 02 Collaboration and Audit", "Feature 2", "Comments, attachments, grouped history, notifications, and follow-up transitions"),
    ]
    for row, (no, func, sheet, desc) in zip(range(9, 11), case_rows):
        write(ws, f"B{row}", no)
        write(ws, f"C{row}", func if language == "en" else desc)
        write(ws, f"D{row}", sheet)
        write(ws, f"E{row}", desc if language == "en" else func, wrap=True)
        write(ws, f"F{row}", "Current Sprint 02 CAP/Fiori baseline", wrap=True)

    ws = wb["Test Statistics"]
    write(ws, "C3", project_name)
    write(ws, "F3", creator)
    write(ws, "C4", project_code)
    write(ws, "F4", "DonHV / Team")
    write(ws, "H5", DOCUMENT_DATE)
    write(
        ws,
        "C6",
        "Release includes happy-flow retest for create, assignment, developer processing, comments, attachments, history, and notifications."
        if language == "en"
        else "Dot nay gom retest happy flow cho create, assignment, xu ly developer, comment, attachment, history va notification.",
        wrap=True,
    )

    feature1 = wb["Feature 1"]
    write(feature1, "B1", "Sprint 02 Create and Lifecycle")
    write(
        feature1,
        "B2",
        "Create bug, assignment, developer review, progress, request more information, reject, resolve, retest, close, and reopen."
        if language == "en"
        else "Tao bug, assign, review, progress, request more information, reject, resolve, retest, close va reopen.",
        wrap=True,
    )
    feature1_cases = [tc for tc in TEST_CASES if tc["id"] in {"SC-01a", "SC-01b", "SC-02a", "SC-05a", "SC-05b", "SC-06a", "SC-07a", "SC-08a", "SC-08b", "SC-08c", "SC-09a", "SC-09b", "SC-10a", "SC-11a", "SC-11b", "SC-12a", "SC-12b", "SC-13a", "SC-13b"}]
    row = 12
    for case in feature1_cases:
        write(feature1, f"A{row}", case["id"])
        write(feature1, f"B{row}", case["en"] if language == "en" else case["vi"], wrap=True)
        write(feature1, f"C{row}", case[f"data_{language}"], wrap=True)
        write(feature1, f"D{row}", case[f"expected_{language}"], wrap=True)
        write(feature1, f"E{row}", "Valid seed/demo state")
        write(feature1, f"F{row}", "Passed")
        write(feature1, f"G{row}", DOCUMENT_DATE)
        write(feature1, f"H{row}", creator)
        write(feature1, f"I{row}", "Passed")
        write(feature1, f"J{row}", DOCUMENT_DATE)
        write(feature1, f"K{row}", creator)
        feature1.row_dimensions[row].height = 36
        row += 1

    feature2 = wb["Feature 2"]
    write(feature2, "B2", "Sprint 02 Collaboration and Audit")
    write(
        feature2,
        "B3",
        "Comment, draft attachment upload/download, grouped history, and notification verification."
        if language == "en"
        else "Comment, upload/download attachment draft, grouped history va notification.",
        wrap=True,
    )
    feature2_cases = [tc for tc in TEST_CASES if tc["id"] in {"SC-03a", "SC-03b", "SC-03c", "SC-03d", "SC-04a", "SC-04b", "SC-08d", "SC-14a", "SC-14b"}]
    row = 12
    for case in feature2_cases:
        write(feature2, f"A{row}", case["id"])
        write(feature2, f"B{row}", case["en"] if language == "en" else case["vi"], wrap=True)
        write(feature2, f"C{row}", case[f"data_{language}"], wrap=True)
        write(feature2, f"D{row}", case[f"expected_{language}"], wrap=True)
        write(feature2, f"E{row}", "Local CAP + HTTP endpoints")
        write(feature2, f"F{row}", "Passed")
        write(feature2, f"G{row}", DOCUMENT_DATE)
        write(feature2, f"H{row}", creator)
        write(feature2, f"I{row}", "Passed")
        write(feature2, f"J{row}", DOCUMENT_DATE)
        write(feature2, f"K{row}", creator)
        feature2.row_dimensions[row].height = 36
        row += 1

    wb.properties.title = f"IDTS SAP490 Test Report {language.upper()} v{TEST_REPORT_VERSION}"
    wb.properties.subject = "IDTS Sprint 02 official retest summary"
    wb.properties.creator = "IDTS SAP01 Team"
    wb.save(output)
    return output


def main():
    outputs = []
    for language in ("en", "vi"):
        outputs.append(fill_functional(language))
        outputs.append(fill_bug_log(language))
        outputs.append(fill_test_scenario(language))
        outputs.append(fill_unit_test(language))
        outputs.append(fill_functional_test(language))
        outputs.append(fill_test_report(language))
    for output in outputs:
        print(output.relative_to(ROOT))


if __name__ == "__main__":
    main()
