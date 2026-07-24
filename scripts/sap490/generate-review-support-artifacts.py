"""Generate template-derived SAP490 technical and review-preparation artifacts.

These files deliberately distinguish planned UAT/change evidence from executed
evidence.  They copy school templates first and never write secrets.
"""

from copy import copy
from datetime import date
from pathlib import Path
import shutil

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment


ROOT = Path(__file__).resolve().parents[2]
TEMPLATES = ROOT / "docs" / "sap490" / "templates" / "Deliverable_template"
OUT = ROOT / "docs" / "sap490" / "generated"
DATE = date(2026, 7, 24)
VERSION = "0.3"
SYSTEM = "IDTS-SAP01"
NAME = "Issue and Defect Tracking System in SAP"


LABELS = {
    "en": {
        "technical_name": "CAP/Fiori MVP Technical Specification",
        "history": "Mentor-ready refresh aligned with deployed OData paths, draft and active write events, exact workflow action audit codes, expanded developer demo data, and the evidence-backed advisory AI boundary.",
        "intro": "IDTS is a SAP CAP Node.js and SAP Fiori Elements application. CAP CDS defines the domain and OData V4 services; Node.js handlers enforce authorization, validation, audit/history and notification side effects.",
        "scope": "Included: bug reporting, classification, responsibility-aware assignment, lifecycle actions, comments, draft attachments, notifications, audit/history, PM monitoring, and review-only AI suggestions. Excluded: autonomous workflow decisions, credentials, private endpoints, and source-code management.",
        "assumptions": "SQLite supports local development; PostgreSQL/HANA portability remains a deployment direction. Real AI is disabled unless private configuration is approved. AI input is allowlisted and AI output is advisory only.",
        "requirements": "AuthService is exposed at /odata/v4/auth/ and BugService at /odata/v4/bug/. Fiori draft creation uses NEW, field changes use PATCH, and SAVE activates the draft through CREATE; editing an active Bug uses EDIT, PATCH, and UPDATE. Bound actions remain server-authorized. Backend remains authoritative for catalog, responsibility, and lifecycle validation.",
        "design": "Core artifacts: db/schema.cds, srv/service.cds, srv/service.js, srv/bug-service/, srv/ai/, and app/bug-management-ui. Data includes Bugs, 14 QA users (12 Developers), 30 DeveloperResponsibilities, master data, exact action history/audit, notifications, PostgreSQL attachment metadata, S3 binary objects, and sanitized AiSuggestions rows that currently remain PENDING.",
        "standards": "Use CAP CDS/OData patterns, server-side authorization/validation, readable Fiori labels, safe error messages, no secrets in source/evidence, and matching docs/knowledge mirrors for changed app/srv/db files.",
        "screen": "Fiori Elements List Report and Object Page: Bug Summary, Classification and Assignment, Reproduction, Evidence/Attachments, Comments, History, Notifications, PM monitoring, and context-local review-only AI actions.",
        "messages": [
            ("IDTS-TECH-001", "Validation", "Required or invalid business data is rejected by the backend.", "Create/update/action validation"),
            ("IDTS-TECH-002", "Authorization", "Only the allowed role can execute this lifecycle action.", "Action authorization"),
            ("IDTS-TECH-003", "AI", "AI suggestions require review and do not change the bug automatically.", "AI suggestion result"),
        ],
        "uat_history": "Prepared UAT plan for mentor/user execution. No test result or sign-off is claimed in this version.",
        "uat_cases": [
            ("UAT-01", "Create a valid bug and verify classification/assignment", "Tester creates a complete bug; system validates fields and stores Assigned or Pending Assignment."),
            ("UAT-02", "Developer review and lifecycle handling", "Developer reviews an assigned bug, requests information or progresses/resolves it with the required reason/note."),
            ("UAT-03", "Tester/PM follow-up and closure", "Tester or PM resubmits, sends to retest, closes, or reopens while history and next processor remain readable."),
            ("UAT-04", "Comment and draft attachment evidence", "User adds a comment and uploads/downloads allowed draft evidence without losing context."),
            ("UAT-05", "PM monitoring", "PM identifies queues, workload, overdue state, and current action ownership."),
            ("UAT-06", "Advisory AI safety", "User opens an AI review suggestion; normal workflow remains usable and no suggestion changes the bug automatically."),
        ],
        "config_items": [
            ("Authentication", "AuthService /odata/v4/auth/ login/logout/me and hashed AuthSessions bearer-token records", "Configured; server-side authorization remains authoritative"),
            ("Bug service and draft", "BugService /odata/v4/bug/; NEW/PATCH/SAVE for create and EDIT/PATCH/UPDATE for active edits", "Configured and verified without changing the public contract"),
            ("Developer demo data", "14 Users, including 12 Developers, with 12 profiles and 30 DeveloperResponsibilities", "Configured for shared QA; values are demo data, not production identities"),
            ("Hosting and database", "Render shared-QA baseline with PostgreSQL; SQLite remains local development only", "Configured; no private connection data in this workbook"),
            ("Attachments", "@cap-js/attachments metadata in PostgreSQL with external S3 object storage for shared QA", "Configured; bucket and credentials remain private"),
            ("Notifications", "Notifications plus NotificationDeliveries email outbox and worker; SMTP/Brevo provider is private configuration", "Configured; provider failure does not roll back the bug workflow"),
            ("Fiori UI", "Fiori Elements Object Page/List Report and safe UI5 extensions", "Configured / verified by QA"),
            ("Workflow audit", "11 exact workflow action types map one-to-one to bound commands", "Configured and verified; legacy audit categories remain for historical compatibility"),
            ("AI provider", "AiSuggestions audit plus optional provider seam; live provider disabled and human-review-only", "NOT ACCEPTED live; mock/fallback/no-mutation PASS; source-linked rows remain PENDING"),
            ("Security evidence", "Secrets, private endpoints, tokens, attachment content and raw provider output are excluded", "Required for every review and deployment handoff"),
        ],
        "change_rows": [
            (1, "DonHV", "IDTS", "Review baseline", "CR-001", "N/A", "Synchronize the CAP/Fiori MVP and SAP490 mentor-review artifacts", "Mentor review", "Pending mentor review", "N/A", "N/A"),
            (2, "DonHV", "IDTS", "Shared-QA configuration", "CR-002", "CR-001", "Document AuthService, Render/PostgreSQL, S3 attachments, notification outbox, and disabled-by-default AI without secrets", "Repository and review pack", "Prepared", "N/A", "N/A"),
            (3, "Team", "IDTS", "QA evidence", "CR-003", "CR-001", "Separate executed Passed evidence from Pending tests and preserve product-only defect reporting", "Mentor review", "Prepared", "N/A", "N/A"),
        ],
    },
    "vi": {
        "technical_name": "Đặc tả kỹ thuật CAP/Fiori MVP",
        "history": "Bản cập nhật mentor-ready đồng bộ endpoint OData đã deploy, sự kiện ghi draft/active, exact workflow action audit, dữ liệu demo Developer mở rộng và ranh giới AI tư vấn có bằng chứng.",
        "intro": "IDTS là ứng dụng SAP CAP Node.js và SAP Fiori Elements. CAP CDS định nghĩa domain và OData V4 service; Node.js handler enforce authorization, validation, audit/history và notification side effect.",
        "scope": "Bao gồm: bug reporting, classification, assignment theo responsibility, lifecycle action, comment, draft attachment, notification, audit/history, PM monitoring và AI suggestion chỉ để review. Loại trừ: quyết định workflow tự động, credential, private endpoint và source-code management.",
        "assumptions": "SQLite dùng cho local development; PostgreSQL/HANA là hướng deployment. Real AI mặc định tắt nếu chưa có private configuration được duyệt. AI input được allowlist và AI output chỉ là advisory.",
        "requirements": "AuthService được expose tại /odata/v4/auth/ và BugService tại /odata/v4/bug/. Luồng tạo draft dùng NEW, thay đổi field dùng PATCH và SAVE kích hoạt draft qua CREATE; sửa Bug active dùng EDIT, PATCH và UPDATE. Bound action luôn được kiểm quyền ở server. Backend quyết định cuối cùng cho catalog, responsibility và lifecycle validation.",
        "design": "Artifact chính: db/schema.cds, srv/service.cds, srv/service.js, srv/bug-service/, srv/ai/ và app/bug-management-ui. Dữ liệu gồm Bugs, 14 user QA (12 Developer), 30 DeveloperResponsibilities, master data, history/audit exact action, notification, metadata attachment trong PostgreSQL, binary trên S3 và AiSuggestions đã làm sạch hiện giữ ở PENDING.",
        "standards": "Dùng CAP CDS/OData pattern, authorization/validation ở server, Fiori label dễ đọc, error message an toàn, không có secret trong source/evidence và cập nhật docs/knowledge mirror cho app/srv/db thay đổi.",
        "screen": "Fiori Elements List Report và Object Page: Bug Summary, Classification and Assignment, Reproduction, Evidence/Attachments, Comments, History, Notifications, PM monitoring và AI action review-only tại đúng context.",
        "messages": [
            ("IDTS-TECH-001", "Validation", "Backend từ chối business data thiếu hoặc không hợp lệ.", "Create/update/action validation"),
            ("IDTS-TECH-002", "Authorization", "Chỉ role được phép mới chạy lifecycle action này.", "Action authorization"),
            ("IDTS-TECH-003", "AI", "AI suggestion cần review và không tự đổi bug.", "AI suggestion result"),
        ],
        "uat_history": "UAT plan đã chuẩn bị cho mentor/user chạy. Version này không tuyên bố test result hoặc sign-off.",
        "uat_cases": [
            ("UAT-01", "Tạo bug hợp lệ và kiểm tra classification/assignment", "Tester tạo bug đủ dữ liệu; hệ thống validate field và lưu Assigned hoặc Pending Assignment."),
            ("UAT-02", "Developer review và lifecycle", "Developer review bug được assign, request information hoặc progress/resolve với reason/note cần thiết."),
            ("UAT-03", "Tester/PM follow-up và closure", "Tester hoặc PM resubmit, retest, close hoặc reopen trong khi history và next processor vẫn dễ đọc."),
            ("UAT-04", "Comment và draft attachment evidence", "User thêm comment và upload/download evidence draft được phép mà không mất context."),
            ("UAT-05", "PM monitoring", "PM nhận diện queue, workload, overdue state và current action owner."),
            ("UAT-06", "AI advisory safety", "User mở AI review suggestion; normal workflow vẫn dùng được và suggestion không tự đổi bug."),
        ],
        "config_items": [
            ("Xác thực", "AuthService /odata/v4/auth/ gồm login/logout/me và bản ghi bearer-token AuthSessions đã băm", "Đã cấu hình; authorization phía server vẫn là nguồn quyết định"),
            ("Bug service và draft", "BugService /odata/v4/bug/; NEW/PATCH/SAVE khi tạo và EDIT/PATCH/UPDATE khi sửa active Bug", "Đã cấu hình và xác minh mà không đổi public contract"),
            ("Dữ liệu demo Developer", "14 Users gồm 12 Developer, 12 profile và 30 DeveloperResponsibilities", "Đã cấu hình cho shared QA; đây là dữ liệu demo, không phải danh tính production"),
            ("Hosting và cơ sở dữ liệu", "Baseline shared QA trên Render với PostgreSQL; SQLite chỉ dùng cho local development", "Đã cấu hình; workbook không chứa connection data riêng tư"),
            ("Tệp đính kèm", "Metadata @cap-js/attachments trong PostgreSQL và object storage S3 bên ngoài cho shared QA", "Đã cấu hình; bucket và credential được giữ riêng tư"),
            ("Thông báo", "Notifications cùng email outbox NotificationDeliveries và worker; SMTP/Brevo là cấu hình riêng tư", "Đã cấu hình; lỗi provider không rollback workflow bug"),
            ("Fiori UI", "Fiori Elements Object Page/List Report và UI5 extension an toàn", "Đã cấu hình / QA đã verify"),
            ("Workflow audit", "11 exact workflow action type map một-một với các bound command", "Đã cấu hình và xác minh; category cũ chỉ giữ để tương thích lịch sử"),
            ("AI provider", "AiSuggestions audit và provider seam tùy chọn; live provider đang tắt và chỉ hỗ trợ human review", "Live NOT ACCEPTED; mock/fallback/no-mutation PASS; source-linked row giữ ở PENDING"),
            ("Bằng chứng bảo mật", "Loại trừ secret, private endpoint, token, nội dung attachment và raw provider output", "Bắt buộc cho mọi review và deployment handoff"),
        ],
        "change_rows": [
            (1, "DonHV", "IDTS", "Baseline review", "CR-001", "N/A", "Đồng bộ CAP/Fiori MVP và artifact SAP490 cho mentor review", "Mentor review", "Đang chờ mentor review", "N/A", "N/A"),
            (2, "DonHV", "IDTS", "Cấu hình shared QA", "CR-002", "CR-001", "Mô tả AuthService, Render/PostgreSQL, S3 attachment, notification outbox và AI mặc định tắt mà không ghi secret", "Repository và review pack", "Đã chuẩn bị", "N/A", "N/A"),
            (3, "Team", "IDTS", "Bằng chứng QA", "CR-003", "CR-001", "Tách evidence Passed đã thực thi khỏi test Pending và giữ defect log chỉ gồm lỗi sản phẩm", "Mentor review", "Đã chuẩn bị", "N/A", "N/A"),
        ],
    },
}


def copy_template(template, output):
    OUT.mkdir(parents=True, exist_ok=True)
    path = OUT / output
    shutil.copy2(TEMPLATES / template, path)
    return path


def writable_cell(ws, coordinate):
    cell = ws[coordinate]
    if not isinstance(cell, MergedCell):
        return cell
    for merged in ws.merged_cells.ranges:
        if coordinate in merged:
            return ws.cell(merged.min_row, merged.min_col)
    raise ValueError(f"No writable anchor for {ws.title}!{coordinate}")


def write(ws, coordinate, value, wrap=True):
    cell = writable_cell(ws, coordinate)
    cell.value = value
    if wrap:
        current = copy(cell.alignment)
        cell.alignment = Alignment(
            horizontal=current.horizontal,
            vertical=current.vertical or "top",
            wrap_text=True,
            text_rotation=current.text_rotation,
            shrink_to_fit=current.shrink_to_fit,
            indent=current.indent,
        )


def remove_template_defined_names(wb):
    """Static review copies do not use legacy template defined names."""
    wb.defined_names.clear()


def fill_cover(ws, title, function_id, function_name):
    write(ws, "B8", title)
    write(ws, "N11", "IDTS")
    write(ws, "Z11", NAME)
    write(ws, "N12", function_id)
    write(ws, "N13", function_name)
    write(ws, "N14", DATE)
    write(ws, "Z14", DATE)
    write(ws, "AE19", "DonHV")


def fill_metadata(ws):
    for cell, value in (("AT2", "DonHV"), ("BC2", DATE), ("AT3", "DonHV"), ("BC3", DATE), ("AT4", "Mentor / Supervisor"), ("BC4", "Pending review")):
        write(ws, cell, value)


def technical(language):
    labels = LABELS[language]
    output = copy_template("Technical_Specification.xlsx", f"Technical_Specification_IDTS_SAP01_{language}_v{VERSION}.xlsx")
    wb = load_workbook(output)
    fill_cover(wb["Cover"], "Technical Specification" if language == "en" else "Đặc tả kỹ thuật", "IDTS-TECH-REVIEW", labels["technical_name"])

    history = wb["Histories"]
    for cell, value in zip(("B3", "C3", "D3", "E3", "F3", "G3"), (1, VERSION, labels["history"], "All mapped sheets", DATE, "DonHV")):
        write(history, cell, value)

    pages = {
        "Introduction": labels["intro"],
        "Scope": labels["scope"],
        "Assumptions": labels["assumptions"],
        "Functional Requirements": labels["requirements"],
        "Technical Design": labels["design"],
        "Development Standards": labels["standards"],
        "Screen Layout": labels["screen"],
        "Screen Definition": labels["screen"],
    }
    for name, text in pages.items():
        ws = wb[name]
        fill_metadata(ws)
        write(ws, "B6", text)
        ws.row_dimensions[6].height = 72

    messages = wb["Message Definition"]
    fill_metadata(messages)
    for row, record in enumerate(labels["messages"], 6):
        for cell, value in zip((f"B{row}", f"H{row}", f"M{row}", f"AQ{row}"), record):
            write(messages, cell, value)
        messages.row_dimensions[row].height = 42

    implementation = wb["Technical Implementation"]
    write(implementation, "B4", "IDTS-TECH-REVIEW")
    write(implementation, "L4", labels["technical_name"])
    write(implementation, "B5", "Repository evidence: CAP compile, selected programmatic regression, secret scan, and template-derived SAP490 artifacts.")
    wb.properties.title = f"IDTS SAP490 Technical Specification {language.upper()} v{VERSION}"
    wb.properties.subject = "CAP/Fiori implementation review; no secrets"
    wb.properties.creator = "IDTS SAP01 Team"
    wb.save(output)
    return output


def uat(language):
    labels = LABELS[language]
    output = copy_template("UAT.xlsx", f"UAT_IDTS_SAP01_{language}_prepared_v{VERSION}.xlsx")
    wb = load_workbook(output)
    remove_template_defined_names(wb)
    fill_cover(wb["Cover"], "User Acceptance Test (Prepared)" if language == "en" else "User Acceptance Test (Đã chuẩn bị)", "IDTS-UAT-REVIEW", "UAT preparation - not executed" if language == "en" else "Chuẩn bị UAT - chưa thực thi")
    history = wb["Histories"]
    for cell, value in zip(("B3", "C3", "D3", "E3", "F3", "G3"), (1, VERSION, labels["uat_history"], "Test Scenario, Test Cases, Test Result", DATE, "DonHV")):
        write(history, cell, value)

    scenarios = wb["Test Scenario"]
    for row, (case_id, title, expected) in enumerate(labels["uat_cases"], 4):
        write(scenarios, f"A{row}", row - 3)
        write(scenarios, f"B{row}", case_id)
        write(scenarios, f"C{row}", "Fiori List Report / Object Page")
        write(scenarios, f"D{row}", "N/A")
        write(scenarios, f"E{row}", f"{title}: {expected}")
        scenarios.row_dimensions[row].height = 48

    cases = wb["Test Cases"]
    for cell, value in (("B3", "IDTS end-to-end UAT"), ("L3", "Mentor/user execution required"), ("BF3", "DonHV"), ("BO3", DATE), ("BV3", "Mentor / Supervisor"), ("CC3", "Pending")):
        write(cases, cell, value)
    for row, (case_id, title, expected) in enumerate(labels["uat_cases"], 8):
        write(cases, f"B{row}", case_id)
        write(cases, f"E{row}", title)
        write(cases, f"Y{row}", "Valid seeded role/data; no credentials in workbook")
        write(cases, f"AP{row}", expected)
        cases.row_dimensions[row].height = 48

    results = wb["Test Result"]
    write(results, "A7", "Prepared UAT only - execution, actual result, defects, acceptance decision, and sign-off are pending mentor/user testing." if language == "en" else "Chỉ là kế hoạch UAT đã chuẩn bị - việc thực thi, actual result, defect, quyết định chấp nhận và sign-off vẫn đang chờ mentor/user kiểm thử.")
    results.row_dimensions[7].height = 60
    wb.properties.title = f"IDTS SAP490 UAT {language.upper()} prepared v{VERSION}"
    wb.properties.subject = "Prepared only; no UAT result or sign-off claimed"
    wb.properties.creator = "IDTS SAP01 Team"
    wb.save(output)
    return output


def configuration(language):
    labels = LABELS[language]
    output = copy_template("Configuration_Note.xlsx", f"Configuration_Note_IDTS_SAP01_{language}_v{VERSION}.xlsx")
    wb = load_workbook(output)
    remove_template_defined_names(wb)
    cover = wb["Cover"]
    write(cover, "I19", SYSTEM)
    write(cover, "I20", VERSION)
    write(cover, "I21", DATE)
    changes = wb["Record of change"]
    for cell, value in zip(("A4", "B4", "C4", "D4", "E4", "F4", "G4"), (1, DATE, VERSION, "Initial CAP/Fiori review configuration note", "SAP490 review evidence", "Mentor / Supervisor", "Pending")):
        write(changes, cell, value)
    checklist = wb["Checklist"]
    for row, (item, detail, status) in enumerate(labels["config_items"], 4):
        for cell, value in zip((f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}"), ("IDTS", row - 3, item, detail, "N/A", "Prepared", status)):
            write(checklist, cell, value)
    sheet_notes = {
        "4": "N/A - IDTS does not use classic SAP customizing transactions; runtime configuration is managed through CAP profiles and private environment variables.",
        "5": "Intentionally unused - no transport request or classic SAP client configuration applies to this CAP/Fiori review artifact.",
    }
    if language == "vi":
        sheet_notes = {
            "4": "N/A - IDTS không dùng transaction customizing SAP cổ điển; cấu hình runtime được quản lý bằng CAP profile và biến môi trường riêng tư.",
            "5": "Cố ý không sử dụng - transport request hoặc cấu hình SAP client cổ điển không áp dụng cho artifact CAP/Fiori này.",
        }
    for name in ("4", "5"):
        ws = wb[name]
        write(ws, "A4", "IDTS CAP/Fiori configuration guidance" if language == "en" else "Hướng dẫn cấu hình IDTS CAP/Fiori")
        write(ws, "C5", "Repository configuration is documented without credentials; use private environment configuration for secrets." if language == "en" else "Cấu hình repository không ghi credential; secret phải dùng cấu hình môi trường riêng tư.")
        write(ws, "C6", sheet_notes[name])
    wb.properties.title = f"IDTS SAP490 Configuration Note {language.upper()} v{VERSION}"
    wb.properties.subject = "Secret-free CAP/Fiori configuration review"
    wb.properties.creator = "IDTS SAP01 Team"
    wb.save(output)
    return output


def change_tracker(language):
    labels = LABELS[language]
    output = copy_template("TR_Management.xlsx", f"TR_Management_IDTS_SAP01_{language}_v{VERSION}.xlsx")
    wb = load_workbook(output)
    remove_template_defined_names(wb)
    ws = wb["Sheet1"]
    changes = labels["change_rows"]
    for row, record in enumerate(changes, 2):
        for column, value in enumerate(record, 1):
            write(ws, ws.cell(row, column).coordinate, value)
    wb.properties.title = f"IDTS SAP490 Change Tracker {language.upper()} v{VERSION}"
    wb.properties.subject = "CAP/Fiori change management adaptation; not an SAP transport log"
    wb.properties.creator = "IDTS SAP01 Team"
    wb.save(output)
    return output


def main():
    outputs = []
    for language in ("en", "vi"):
        outputs.extend((technical(language), uat(language), configuration(language), change_tracker(language)))
    for output in outputs:
        print(output.relative_to(ROOT))


if __name__ == "__main__":
    main()
