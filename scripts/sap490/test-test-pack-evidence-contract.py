"""Focused regression checks for SAP490 test-pack layout and evidence links."""

from __future__ import annotations

import sys
from pathlib import Path
from urllib.parse import unquote

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


ROOT = Path(__file__).resolve().parents[2]
OUTPUT = ROOT / "docs" / "sap490" / "generated"


def hyperlink_target_exists(workbook, hyperlink) -> bool:
    if getattr(hyperlink, "target", None):
        return False
    location = unquote(str(
        getattr(hyperlink, "location", None)
        or ""
    )).lstrip("#")
    if not location or "!" not in location:
        return bool(location)
    sheet_name, coordinate = location.rsplit("!", 1)
    sheet_name = sheet_name.strip("'")
    if sheet_name not in workbook.sheetnames:
        return False
    return workbook[sheet_name][coordinate.replace("$", "")].value not in (None, "")


def has_hyperlink_appearance(cell) -> bool:
    color = cell.font.color
    rgb = getattr(color, "rgb", None) if color else None
    return bool(cell.font.underline) or (isinstance(rgb, str) and rgb[-6:] in {"0563C1", "0000FF"})


def validate_functional(language: str, errors: list[str]) -> None:
    path = OUTPUT / f"Functional_Test_IDTS_SAP01_{language}_v0.3.xlsx"
    workbook = load_workbook(path, data_only=False)
    sheet = workbook["Test Result"]
    first_run_row = next(
        (
            row
            for row in range(5, sheet.max_row + 1)
            if str(sheet[f"B{row}"].value or "").startswith("REG-")
        ),
        None,
    )
    if first_run_row != 5:
        errors.append(
            f"{path.name}: first execution run must start in the official row block at row 5; got {first_run_row}"
        )
    workbook.close()


def validate_report(language: str, errors: list[str]) -> None:
    path = OUTPUT / f"Test_Report_IDTS_SAP01_{language}_v0.4.xlsx"
    workbook = load_workbook(path, data_only=False)
    sheet = workbook["Test Cases"]
    for row in range(9, sheet.max_row + 1):
        link_cell = sheet[f"D{row}"]
        plain_cell = sheet[f"E{row}"]
        if not link_cell.hyperlink or not hyperlink_target_exists(workbook, link_cell.hyperlink):
            errors.append(f"{path.name}/{sheet.title}!D{row}: broken or blank hyperlink target")
        if has_hyperlink_appearance(plain_cell) and not plain_cell.hyperlink:
            errors.append(f"{path.name}/{sheet.title}!E{row}: hyperlink styling without a hyperlink")
    workbook.close()


def validate_unit(language: str, errors: list[str]) -> None:
    path = OUTPUT / f"Unit_Test_IDTS_SAP01_{language}_v0.4.xlsx"
    workbook = load_workbook(path, data_only=False)
    sheet = workbook["UT"]
    for row in range(8, 13):
        cell = sheet[f"BL{row}"]
        if not cell.hyperlink:
            errors.append(f"{path.name}/{sheet.title}!BL{row}: missing case-specific evidence link")
        elif not hyperlink_target_exists(workbook, cell.hyperlink):
            errors.append(f"{path.name}/{sheet.title}!BL{row}: evidence link points to a blank or missing target")
    workbook.close()


def main() -> int:
    errors: list[str] = []
    for language in ("en", "vi"):
        validate_functional(language, errors)
        validate_report(language, errors)
        validate_unit(language, errors)
    if errors:
        print("SAP490 evidence contract: FAIL")
        for error in errors:
            print(f"- {error}")
        return 1
    print("SAP490 evidence contract: PASS")
    return 0


if __name__ == "__main__":
    sys.exit(main())
