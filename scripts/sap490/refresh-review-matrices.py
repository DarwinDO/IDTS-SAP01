"""Refresh the local SAP490 PM and team matrices without external dependencies."""

from copy import copy
from datetime import date
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
GENERATED = ROOT / "docs/sap490/generated"
RUN_DATE = "2026-07-24"


MEMBERS = {
    "donhv": ("DonHV", "BA/PM consolidation; CAP backend and integration"),
    "datdt": ("DatDT", "SAP Fiori/UI5 lead"),
    "sangvn": ("SangVN", "SAP Fiori/UI5 support"),
    "nhant": ("NhanT", "Backend verification and QA"),
}


def clone_row_style(ws, source_row, target_row):
    for column in range(1, ws.max_column + 1):
        src = ws.cell(source_row, column)
        dst = ws.cell(target_row, column)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def configure_print_layout(workbook):
    """Print only populated cells and fit every worksheet to one page wide."""
    for ws in workbook.worksheets:
        populated = [
            cell
            for row in ws.iter_rows()
            for cell in row
            if cell.value not in (None, "")
        ]
        if not populated:
            continue
        max_row = max(cell.row for cell in populated)
        max_column = max(cell.column for cell in populated)
        ws.print_area = f"A1:{get_column_letter(max_column)}{max_row}"
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.scale = None
        ws.page_setup.orientation = "landscape" if max_column > 6 else "portrait"
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5


def extract_status_entries(member_id):
    path = ROOT / f"docs/pm/status/{member_id}.md"
    text = path.read_text(encoding="utf-8")
    matches = list(re.finditer(r"^##\s+(\d{4}-\d{2}-\d{2})\s*[-—:]?\s*(.+)$", text, re.M))
    entries = []
    for index, match in enumerate(matches):
        if match.group(1) < "2026-07-10":
            continue
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        body = text[match.end():end]
        detail = next((line.strip(" -*") for line in body.splitlines() if line.strip() and not line.lstrip().startswith(("#", "|"))), "")
        summary = re.sub(r"\s+", " ", match.group(2)).strip()
        jira = next(iter(re.findall(r"IDTS-\d+", summary + " " + detail)), "—")
        pr = next(iter(re.findall(r"(?:PR\s*#|pull/)(\d+)", summary + " " + detail, re.I)), None)
        entries.append((match.group(1), jira, summary, detail[:800], pr))
    return entries


def refresh_contributions():
    output = GENERATED / "Team_Contribution_Matrix_IDTS_SAP01_20260724.xlsx"
    wb = load_workbook(output)
    summary = wb["Summary"]
    summary["A1"] = "SU26SAP01_GSU26SAP01 — Team Contribution Matrix — 2026-07-24"
    summary["A2"] = "Current mentor-review matrix refreshed from member status files. Attribution remains evidence-based; mentor approval and human UAT sign-off are outside this workbook."

    ws = wb["Contributions"]
    existing = {(str(ws.cell(r, 1).value), str(ws.cell(r, 3).value), str(ws.cell(r, 6).value)) for r in range(2, ws.max_row + 1)}
    for member_id, (name, role) in MEMBERS.items():
        for entry_date, jira, work, detail, pr in extract_status_entries(member_id):
            key = (name, entry_date, work)
            if key in existing:
                continue
            row = ws.max_row + 1
            clone_row_style(ws, max(2, row - 1), row)
            values = [
                name, role, entry_date, jira,
                f"https://dutassociation.atlassian.net/browse/{jira}" if jira != "—" else "—",
                work, detail or "—", "Documentation / QA / implementation",
                "Primary recorded contribution" if member_id == "donhv" else "Owner / support contribution",
                f"https://github.com/DarwinDO/IDTS-SAP01/pull/{pr}" if pr else "—",
                f"docs/pm/status/{member_id}.md", "Status heading", "Review evidence before final attribution",
            ]
            for column, value in enumerate(values, 1):
                ws.cell(row, column, value)
            existing.add(key)

    evidence = wb["Evidence Index"]
    known = {str(evidence.cell(r, 2).value) for r in range(2, evidence.max_row + 1)}
    for issue in ("IDTS-82", "IDTS-89", "IDTS-90", "IDTS-100"):
        if issue in known:
            continue
        row = evidence.max_row + 1
        clone_row_style(evidence, max(2, row - 1), row)
        values = ["Jira", issue, f"https://dutassociation.atlassian.net/browse/{issue}", "DonHV / team", "docs/pm/status/donhv.md", "docs/pm/evidence"]
        for column, value in enumerate(values, 1):
            evidence.cell(row, column, value)

    configure_print_layout(wb)
    wb.properties.title = "IDTS SAP490 Team Contribution Matrix 2026-07-24"
    wb.save(output)
    return output


def parse_risk_rows():
    text = (ROOT / "docs/pm/risk-decision-log.md").read_text(encoding="utf-8")
    rows = []
    for line in text.splitlines():
        if not re.match(r"^\|\s*(DEC|RISK)-\d+\s*\|", line):
            continue
        cells = [cell.strip() for cell in line.strip().strip("|").split("|")]
        kind = "Decision" if cells[0].startswith("DEC-") else "Risk"
        rows.append([kind, *cells[:6], "docs/pm/risk-decision-log.md"])
    return rows


def refresh_pm_matrices():
    output = GENERATED / "SAP490_Review_Matrices_IDTS_SAP01_20260724.xlsx"
    wb = load_workbook(output)
    summary = wb["Summary"]
    summary["B2"] = RUN_DATE
    summary["B4"] = "Repository Markdown + current IDTS-100 Shared QA evidence"
    summary["B5"] = "SAP490 mentor/team review workbook; current result: 21 PASSED cases and 6 PREPARED human UAT cases"

    ws = wb["Risk Decision Log"]
    for row in range(2, ws.max_row + 1):
        for column in range(1, 8):
            ws.cell(row, column).value = None
    for row, values in enumerate(parse_risk_rows(), 2):
        if row > ws.max_row:
            clone_row_style(ws, row - 1, row)
        for column, value in enumerate(values[:7], 1):
            ws.cell(row, column, value)
    ws.auto_filter.ref = f"A1:G{ws.max_row}"
    configure_print_layout(wb)
    wb.properties.title = "IDTS SAP490 Review Matrices 2026-07-24"
    wb.save(output)
    return output


def main():
    print(refresh_contributions().relative_to(ROOT))
    print(refresh_pm_matrices().relative_to(ROOT))


if __name__ == "__main__":
    main()
