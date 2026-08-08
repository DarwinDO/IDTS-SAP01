#!/usr/bin/env python3
"""Validate structural and critical-range fidelity for an IDTS SAP490 XLSX candidate."""

from __future__ import annotations

import argparse
import json
import math
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Side
from openpyxl.utils.cell import range_boundaries


VISIBLE_BORDER_STYLES = {
    "dashDot", "dashDotDot", "dashed", "dotted", "double", "hair",
    "medium", "mediumDashDot", "mediumDashDotDot", "mediumDashed",
    "slantDashDot", "thick", "thin",
}


def normalize_rgb(value: str | None) -> str | None:
    if not value:
        return None
    value = value.upper()
    return f"FF{value}" if len(value) == 6 else value


def border_visible(side: Side) -> bool:
    return bool(side and side.style in VISIBLE_BORDER_STYLES)


def border_color(side: Side) -> tuple[str | None, bool]:
    if not border_visible(side):
        return None, False
    color = side.color
    if color is None:
        return None, True
    if color.type == "rgb":
        return normalize_rgb(color.rgb), False
    return None, True


def cells_in_range(sheet, address: str):
    min_col, min_row, max_col, max_row = range_boundaries(address)
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            yield cell


def compare_page_setup(reference, candidate, errors: list[str]) -> None:
    attrs = ("orientation", "paperSize", "fitToWidth", "fitToHeight", "scale")
    for attr in attrs:
        before = getattr(reference.page_setup, attr)
        after = getattr(candidate.page_setup, attr)
        if before != after:
            errors.append(f"{candidate.title}: page_setup.{attr} changed from {before!r} to {after!r}")
    margin_attrs = ("left", "right", "top", "bottom", "header", "footer")
    for attr in margin_attrs:
        before = getattr(reference.page_margins, attr)
        after = getattr(candidate.page_margins, attr)
        if not math.isclose(float(before), float(after), rel_tol=0, abs_tol=1e-6):
            errors.append(f"{candidate.title}: page margin {attr} changed from {before!r} to {after!r}")


def validate_frame(sheet, address: str, errors: list[str]) -> None:
    min_col, min_row, max_col, max_row = range_boundaries(address)
    for col in range(min_col, max_col + 1):
        top = sheet.cell(min_row, col)
        bottom = sheet.cell(max_row, col)
        if not border_visible(top.border.top):
            errors.append(f"{sheet.title}!{top.coordinate}: missing top border for frame {address}")
        if not border_visible(bottom.border.bottom):
            errors.append(f"{sheet.title}!{bottom.coordinate}: missing bottom border for frame {address}")
    for row in range(min_row, max_row + 1):
        left = sheet.cell(row, min_col)
        right = sheet.cell(row, max_col)
        if not border_visible(left.border.left):
            errors.append(f"{sheet.title}!{left.coordinate}: missing left border for frame {address}")
        if not border_visible(right.border.right):
            errors.append(f"{sheet.title}!{right.coordinate}: missing right border for frame {address}")


def validate_no_grid(sheet, address: str, errors: list[str]) -> None:
    for cell in cells_in_range(sheet, address):
        if isinstance(cell, MergedCell):
            continue
        for side_name in ("top", "bottom", "left", "right", "diagonal", "vertical", "horizontal"):
            side = getattr(cell.border, side_name, None)
            if border_visible(side):
                errors.append(f"{sheet.title}!{cell.coordinate}: unexpected visible border {side_name} in borderless range {address}")


def validate_style_rule(sheet, rule: dict[str, Any], errors: list[str]) -> None:
    address = rule["range"]
    expected_font = rule.get("font", {})
    expected_fill = normalize_rgb(rule.get("fill_rgb"))
    expected_border = normalize_rgb(rule.get("border_rgb"))
    for cell in cells_in_range(sheet, address):
        if isinstance(cell, MergedCell):
            continue
        for key in ("name", "bold", "italic", "underline"):
            if key in expected_font and getattr(cell.font, key) != expected_font[key]:
                errors.append(f"{sheet.title}!{cell.coordinate}: font {key}={getattr(cell.font, key)!r}, expected {expected_font[key]!r}")
        if "size" in expected_font and not math.isclose(float(cell.font.sz or 0), float(expected_font["size"]), abs_tol=0.01):
            errors.append(f"{sheet.title}!{cell.coordinate}: font size={cell.font.sz!r}, expected {expected_font['size']!r}")
        if expected_fill:
            fill_rgb = normalize_rgb(cell.fill.fgColor.rgb) if cell.fill.fill_type == "solid" and cell.fill.fgColor.type == "rgb" else None
            if fill_rgb != expected_fill:
                errors.append(f"{sheet.title}!{cell.coordinate}: fill={fill_rgb!r}, expected solid {expected_fill}")
        if expected_border:
            for side_name in ("top", "bottom", "left", "right"):
                side = getattr(cell.border, side_name)
                color, implicit = border_color(side)
                if implicit:
                    errors.append(f"{sheet.title}!{cell.coordinate}: {side_name} uses automatic/theme border color")
                elif border_visible(side) and color != expected_border:
                    errors.append(f"{sheet.title}!{cell.coordinate}: {side_name} border color={color!r}, expected {expected_border}")


def validate(reference_path: Path | None, candidate_path: Path, policy_path: Path) -> int:
    policy = json.loads(policy_path.read_text(encoding="utf-8"))
    candidate = load_workbook(candidate_path, data_only=False)
    reference = load_workbook(reference_path, data_only=False) if reference_path else None
    errors: list[str] = []

    if reference:
        if policy.get("preserve_sheet_order", True) and reference.sheetnames != candidate.sheetnames:
            errors.append(f"Workbook sheet order changed: {reference.sheetnames!r} -> {candidate.sheetnames!r}")
        if policy.get("preserve_sheet_visibility", True):
            for name in set(reference.sheetnames) & set(candidate.sheetnames):
                if reference[name].sheet_state != candidate[name].sheet_state:
                    errors.append(f"{name}: sheet visibility changed from {reference[name].sheet_state!r} to {candidate[name].sheet_state!r}")

    for sheet_name, sheet_policy in policy.get("sheets", {}).items():
        if sheet_name not in candidate.sheetnames:
            errors.append(f"Missing required sheet: {sheet_name}")
            continue
        sheet = candidate[sheet_name]
        if "show_gridlines" in sheet_policy and sheet.sheet_view.showGridLines != sheet_policy["show_gridlines"]:
            errors.append(f"{sheet_name}: showGridLines={sheet.sheet_view.showGridLines!r}, expected {sheet_policy['show_gridlines']!r}")
        if "print_area" in sheet_policy and str(sheet.print_area) != sheet_policy["print_area"]:
            errors.append(f"{sheet_name}: print_area={str(sheet.print_area)!r}, expected {sheet_policy['print_area']!r}")
        if reference and sheet_name in reference.sheetnames:
            ref_sheet = reference[sheet_name]
            if sheet_policy.get("preserve_page_setup", True):
                compare_page_setup(ref_sheet, sheet, errors)
            if sheet_policy.get("preserve_merges", True):
                before = sorted(str(item) for item in ref_sheet.merged_cells.ranges)
                after = sorted(str(item) for item in sheet.merged_cells.ranges)
                if before != after:
                    errors.append(f"{sheet_name}: merged ranges changed without policy approval")
        for address in sheet_policy.get("closed_frames", []):
            validate_frame(sheet, address, errors)
        for address in sheet_policy.get("no_grid_ranges", []):
            validate_no_grid(sheet, address, errors)
        for rule in sheet_policy.get("style_rules", []):
            validate_style_rule(sheet, rule, errors)

    if errors:
        print(f"FAIL: {len(errors)} XLSX fidelity issue(s)")
        for item in errors:
            print(f"- {item}")
        return 1
    print("PASS: XLSX structure and critical-range fidelity policy satisfied")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)
    validate_parser = subparsers.add_parser("validate")
    validate_parser.add_argument("--reference", type=Path)
    validate_parser.add_argument("--candidate", type=Path, required=True)
    validate_parser.add_argument("--policy", type=Path, required=True)
    args = parser.parse_args()
    if args.command == "validate":
        return validate(args.reference, args.candidate, args.policy)
    return 2


if __name__ == "__main__":
    sys.exit(main())
