import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Border, Color, Font, PatternFill, Side


SCRIPT = Path(__file__).with_name("audit_xlsx_fidelity.py")


def make_workbook(path: Path, *, broken: bool = False) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Technical Design"
    sheet.sheet_view.showGridLines = False
    sheet.print_area = "B2:F8"
    sheet.page_setup.orientation = "landscape"

    black = Side(style="thin", color="FF000000")
    auto = Side(style="thin", color=Color(auto=True))
    for cell in sheet["B2:F2"][0]:
        cell.fill = PatternFill("solid", fgColor="FFFFFFFF")
        cell.font = Font(name="Times New Roman", size=12, bold=True, italic=True)
        cell.border = Border(
            top=auto if broken else black,
            bottom=black,
            left=black if cell.column == 2 else Side(),
            right=Side() if broken or cell.column != 6 else black,
        )
    sheet["B2"] = "Tables"

    for row in sheet.iter_rows(min_row=3, max_row=6, min_col=2, max_col=6):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="FFFFFFFF")
            cell.font = Font(name="Times New Roman", size=12)
            cell.border = Border(bottom=black) if broken else Border()
    sheet["B3"] = "idts_cap_Bugs"
    sheet["D3"] = "Authoritative Bug record."
    workbook.save(path)


class AuditXlsxFidelityTest(unittest.TestCase):
    def run_audit(self, workbook: Path, policy: Path) -> subprocess.CompletedProcess[str]:
        return subprocess.run(
            [sys.executable, str(SCRIPT), "validate", "--candidate", str(workbook), "--policy", str(policy)],
            text=True,
            capture_output=True,
            check=False,
        )

    def test_valid_frame_body_and_typography_pass(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook = root / "candidate.xlsx"
            policy = root / "policy.json"
            make_workbook(workbook)
            policy.write_text(json.dumps({
                "sheets": {
                    "Technical Design": {
                        "show_gridlines": False,
                        "print_area": "'Technical Design'!$B$2:$F$8",
                        "closed_frames": ["B2:F2"],
                        "no_grid_ranges": ["B3:F6"],
                        "style_rules": [{
                            "range": "B2:F2",
                            "font": {"name": "Times New Roman", "size": 12, "bold": True, "italic": True},
                            "fill_rgb": "FFFFFFFF",
                            "border_rgb": "FF000000"
                        }]
                    }
                }
            }), encoding="utf-8")
            result = self.run_audit(workbook, policy)
            self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
            self.assertIn("PASS", result.stdout)

    def test_missing_right_edge_auto_border_and_body_grid_fail(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook = root / "candidate.xlsx"
            policy = root / "policy.json"
            make_workbook(workbook, broken=True)
            policy.write_text(json.dumps({
                "sheets": {
                    "Technical Design": {
                        "closed_frames": ["B2:F2"],
                        "no_grid_ranges": ["B3:F6"],
                        "style_rules": [{
                            "range": "B2:F2",
                            "font": {"name": "Times New Roman", "size": 12, "bold": True, "italic": True},
                            "fill_rgb": "FFFFFFFF",
                            "border_rgb": "FF000000"
                        }]
                    }
                }
            }), encoding="utf-8")
            result = self.run_audit(workbook, policy)
            self.assertNotEqual(result.returncode, 0)
            self.assertIn("missing right border", result.stdout)
            self.assertIn("automatic/theme border color", result.stdout)
            self.assertIn("unexpected visible border", result.stdout)


if __name__ == "__main__":
    unittest.main()
